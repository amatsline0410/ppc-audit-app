"""Listing Optimizer (Competitor Research / Indexed Keywords / SEO tracker).

Replaces the "Project Snore x Competitor Research / Indexed Keywords / SEO"
Google Sheet (mapping: research_tracker/MAPPING.md — the approved mapping wins).

  migrate(db, path, ...)   one-time sheet import (Main grid + X-ray + Listing
                           Audit markers + Listing Copy + Search Terms keywords)
  create_project / import_xray / set_listing_copy / listing_audit
                           the RAW-data flow: blank project + raw Cerebro/X-ray
                           exports + pasted copy; the app computes the analysis
  import_cerebro(...)      ongoing snapshot import (Cerebro export, 1+ ASINs)
  matrix / scorecard / movers / suggest / export_matrix
                           the sheet's views, recomputed live

Rank semantics (per mapping): integer = organic rank; '-' and 0 = not ranked
(stored None on a snapshot row = checked-but-unranked); NO row = unchecked.
Snapshots append-only — re-importing a date replaces that date only.
"""
from __future__ import annotations
import io
import math
import re
from datetime import date
import pandas as pd
from sqlalchemy import text as _sqltext
from . import workbook
from sqlalchemy.orm import Session
from .. import models as md

PAGE_SIZE = 48                      # desktop results/page (mapping open-q #7 default)
ELEMENTS = ["title", "bullet_points", "aplus", "description", "search_terms", "alt_text",
            "sp_broad", "sp_phrase", "sp_exact", "sb_broad", "sb_phrase", "sb_exact"]

# ---- pure metrics ------------------------------------------------------------

def page(rank: int | None) -> int | None:
    """Result page for an organic rank (48/page)."""
    if rank is None or rank <= 0:
        return None
    return math.ceil(rank / PAGE_SIZE)


def index_rate(ranked: int, tracked: int) -> float | None:
    """ranked/tracked, None-safe."""
    return round(ranked / tracked, 4) if tracked else None


def rank_delta(cur: int | None, prev: int | None) -> int | None:
    """current - previous; negative = improved. None if either side unranked."""
    if cur is None or prev is None:
        return None
    return cur - prev


_norm_re = re.compile(r"\s+")

def _norm(s: str) -> str:
    return _norm_re.sub(" ", (s or "").lower()).strip()


def match_kind(copy_text: str, keyword: str) -> str | None:
    """'exact' if the whole phrase appears (case/space folded); 'broad' if every
    word appears somewhere; None otherwise. (Sheet's Comparison-tab markers.)"""
    text, kw = _norm(copy_text), _norm(keyword)
    if not text or not kw:
        return None
    if kw in text:
        return "exact"
    words = set(text.split())
    if all(w in words for w in kw.split()):
        return "broad"
    return None


def count_exact(copy_text: str, keywords: list) -> int:
    """Sheet's 'No. of exact' — tracked keywords appearing verbatim in the text."""
    return sum(1 for k in keywords if match_kind(copy_text, k.keyword) == "exact")


def total_exact_sv(copy_text: str, keywords: list) -> int:
    """Sheet's 'Total exact SV' — search volume over the exact-matched keywords."""
    return sum((k.search_volume or 0) for k in keywords
               if match_kind(copy_text, k.keyword) == "exact")


# ---- tolerant cell coercion --------------------------------------------------

def _int(v) -> int | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip().replace(",", "").replace("$", "")
    if s in ("", "-", "nan", "None"):
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _num(v) -> float | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip().replace(",", "").replace("$", "").replace("%", "")
    if s in ("", "-", "nan", "None") or s.startswith("#"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _rank(v) -> int | None:
    """Rank cell: '-' and 0 both mean not-ranked (mapping open-q #1 default)."""
    r = _int(v)
    return r if r and r > 0 else None


def _yes(v) -> bool | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip().lower()
    if s in ("", "nan"):
        return None
    return s in ("yes", "y", "true", "1")


def _s(v) -> str | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    return s if s and s.lower() != "nan" and not s.startswith("#") else None


# ---- sheet migration ---------------------------------------------------------

_ASIN_RE = re.compile(r"^B0[A-Z0-9]{8}$")


def purge_project(db: Session, proj: "md.TrackerProject") -> None:
    """Delete a project + all children explicitly (don't rely on SQLite FK
    pragma being on for the session — belt and braces)."""
    kw_ids = [k.id for k in db.query(md.TrackedKeyword.id)
              .filter(md.TrackedKeyword.project_id == proj.id)]
    if kw_ids:
        db.query(md.RankSnapshot).filter(
            md.RankSnapshot.keyword_id.in_(kw_ids)).delete(synchronize_session=False)
        db.query(md.KeywordUsage).filter(
            md.KeywordUsage.keyword_id.in_(kw_ids)).delete(synchronize_session=False)
        db.query(md.TrackedKeyword).filter(
            md.TrackedKeyword.project_id == proj.id).delete(synchronize_session=False)
    db.query(md.TrackedCompetitor).filter(
        md.TrackedCompetitor.project_id == proj.id).delete(synchronize_session=False)
    db.query(md.ListingCopy).filter(
        md.ListingCopy.project_id == proj.id).delete(synchronize_session=False)
    db.delete(proj)

# Main-tab attribute label -> TrackedCompetitor field
_ATTR = {"price": "price", "sales": "sales", "revenue": "revenue",
         "review count": "review_count", "star ratings": "rating",
         "fulfillment": "fulfillment", "size tier": "size_tier",
         "seller country/region": "seller_country", "creation date": "creation_date",
         "listing health score": "listing_health_score"}
_ATTR_YES = {"pdp images": "pdp_images", "pdp videos": "pdp_videos",
             "brand story": "brand_story", "generic/premium a+": "aplus",
             "crawlable text": "crawlable_text", "alt text": "alt_text",
             "comparison table": "comparison_table",
             "amazon's badge/highlight": "amazon_badge"}
# X-ray header (normalized: BOM-stripped, whitespace-collapsed, lowercased) ->
# field. Loose aliases: covers both the sheet's X-ray tabs and raw Helium10
# X-ray CSV exports ("Price  $", "Fees  $", "ASIN Sales", "Active Sellers",
# "Image URL", ... — exports drift, never match exact headers).
_XRAY = {"product details": "title", "asin": "asin", "brand": "brand",
         "price $": "price", "price": "price",
         "sales": "sales", "asin sales": "sales",
         "revenue": "revenue", "asin revenue": "revenue",
         "bsr": "bsr", "seller country/region": "seller_country",
         "fba fees $": "fba_fees", "fees $": "fba_fees", "fba fees": "fba_fees",
         "active sellers #": "active_sellers", "active sellers": "active_sellers",
         "ratings": "rating", "rating": "rating",
         "review count": "review_count", "images": "images",
         "review velocity": "review_velocity", "buy box": "buy_box",
         "category": "category", "size tier": "size_tier", "fulfillment": "fulfillment",
         "dimensions": "dimensions", "weight": "weight",
         "creation date": "creation_date",
         "image source url": "image_url", "image url": "image_url"}
_INT_F = {"sales", "bsr", "active_sellers", "review_count", "images", "review_velocity"}
_NUM_F = {"price", "revenue", "fba_fees", "rating", "weight", "listing_health_score"}
# Listing Audit column label -> KeywordUsage element
_LA_ELEM = {"title": "title", "bullet points": "bullet_points",
            "a+ / brand story": "aplus", "a+ /brand story": "aplus", "a+/brand story": "aplus",
            "description": "description", "search terms": "search_terms",
            "alt text": "alt_text",
            "sp targeting broad": "sp_broad", "sp targeting phrase": "sp_phrase",
            "sp targeting exact": "sp_exact", "sb targeting broad": "sb_broad",
            "sb targeting phrase": "sb_phrase", "sb targeting exact": "sb_exact"}
_LC_ELEM = {"title": "title", "bullet points": "bullet_points", "a+ texts": "aplus",
            "description": "description", "search terms": "search_terms",
            "alt text": "alt_text"}


def _norm_label(v) -> str:
    return _norm(str(v or "").replace("\ufeff", "").replace("\n", " "))


def _iso_date(v) -> str | None:
    """Creation-date cell -> ISO string. Exports drift: 'Mar 22, 2011',
    '2011-03-22', datetimes — pandas parses them all."""
    s = _s(v)
    if not s:
        return None
    ts = pd.to_datetime(s, errors="coerce")
    return None if pd.isna(ts) else ts.date().isoformat()


def _xray_rows(df: pd.DataFrame) -> list[dict]:
    """Parse an X-ray tab/export (header row 0) into TrackedCompetitor field
    dicts. Headers matched loosely (normalized + aliases) — covers the sheet's
    X-ray tabs and raw Helium10 X-ray CSVs."""
    if df.empty:
        return []
    heads = [_norm_label(h) if pd.notna(h) else "" for h in df.iloc[0]]
    out = []
    for i in range(1, len(df)):
        row = df.iloc[i]
        rec = {}
        for c, h in enumerate(heads):
            f = _XRAY.get(h)
            if not f or rec.get(f) is not None:   # first non-empty alias wins
                continue
            v = row[c]
            if f in _INT_F:
                rec[f] = _int(v)
            elif f in _NUM_F:
                rec[f] = _num(v)
            elif f == "creation_date":
                rec[f] = _iso_date(v)
            else:
                rec[f] = _s(v)
        if rec.get("asin") and _ASIN_RE.match(rec["asin"]):
            out.append(rec)
    return out


def migrate(db: Session, path: str, name: str | None = None,
            snapshot_date: date | None = None) -> dict:
    """One-time sheet import per the approved MAPPING.md. Idempotent-ish: a
    project with the same name is replaced (cascade wipes its children)."""
    snap = snapshot_date or date.today()
    S = pd.read_excel(path, sheet_name=None, dtype=str, header=None, engine=workbook.excel_engine())
    S = {k.strip().lower(): v for k, v in S.items()}
    main = S.get("main")
    if main is None:
        raise ValueError("No 'Main' tab — this doesn't look like the competitor research sheet.")

    # -- primary + competitors from the X-ray tabs
    your = _xray_rows(S.get("your x-ray", pd.DataFrame()))
    comps = _xray_rows(S.get("comp x-ray", pd.DataFrame()))
    primary_asin = your[0]["asin"] if your else None

    pname = name or "Imported project"
    old = db.query(md.TrackerProject).filter(md.TrackerProject.name == pname).first()
    if old:
        purge_project(db, old)
        db.flush()
    proj = md.TrackerProject(name=pname, primary_asin=primary_asin)
    db.add(proj)
    db.flush()

    comp_rows: dict[str, md.TrackedCompetitor] = {}
    for rec, prim in [(r, True) for r in your] + [(r, False) for r in comps]:
        c = md.TrackedCompetitor(project_id=proj.id, is_primary=prim, **rec)
        db.add(c)
        comp_rows[rec["asin"]] = c

    # -- Main tab: locate the ASIN header row of the attribute matrix
    n_attr_health = 0
    label_col, first_val = None, None
    for i in range(min(60, len(main))):
        for c in range(0, min(8, main.shape[1] - 1)):
            if _s(main.iat[i, c]) == "ASIN":
                cand = _s(main.iat[i, c + 1]) or ""
                if _ASIN_RE.match(cand):
                    label_col, first_val, asin_row = c, c + 1, i
                    break
        if label_col is not None:
            break
    if label_col is not None:
        asins_by_col = {c: _s(main.iat[asin_row, c]) for c in range(first_val, main.shape[1])
                        if _s(main.iat[asin_row, c]) and _ASIN_RE.match(_s(main.iat[asin_row, c]))}
        # attribute rows below, until the CEREBRO DATA divider
        for i in range(asin_row + 1, min(asin_row + 40, len(main))):
            label = _norm_label(main.iat[i, label_col])
            if not label:
                continue
            if "cerebro" in label:
                break
            f = _ATTR.get(label)
            fy = _ATTR_YES.get(label)
            if not f and not fy:
                continue
            for c, asin in asins_by_col.items():
                comp = comp_rows.get(asin)
                if comp is None:
                    continue
                v = main.iat[i, c]
                if fy:
                    setattr(comp, fy, _yes(v))
                    n_attr_health += 1
                elif f in _NUM_F:
                    if getattr(comp, f, None) is None:
                        setattr(comp, f, _num(v))
                elif f == "creation_date":
                    if getattr(comp, f, None) is None:
                        setattr(comp, f, (_s(v) or "")[:10] or None)

    # -- Main tab: keyword grid. Header row: col0 'Search Terms', ASINs from col 4+
    hdr = None
    for i in range(min(80, len(main))):
        if _s(main.iat[i, 0]) == "Search Terms" and _s(main.iat[i, 1]) == "Search Volume":
            hdr = i
            break
    if hdr is None:
        raise ValueError("Main tab: couldn't find the 'Search Terms' keyword grid header.")
    grid_asins = {c: _s(main.iat[hdr, c]) for c in range(3, main.shape[1])
                  if _s(main.iat[hdr, c]) and _ASIN_RE.match(_s(main.iat[hdr, c]))}
    # data starts after the 'Organic Rank'/'Filter' chrome rows
    start = hdr + 1
    while start < len(main) and _s(main.iat[start, 0]) is None:
        start += 1

    kw_rows: dict[str, md.TrackedKeyword] = {}
    snaps = []
    for i in range(start, len(main)):
        kw = _s(main.iat[i, 0])
        if kw is None:
            continue
        key = _norm(kw)
        if key in kw_rows:
            continue
        k = md.TrackedKeyword(project_id=proj.id, keyword=kw,
                              search_volume=_int(main.iat[i, 1]),
                              relevancy=_int(main.iat[i, 2]),
                              keyword_sales=_int(main.iat[i, 3]),
                              source="cerebro")
        db.add(k)
        kw_rows[key] = k
        for c, asin in grid_asins.items():
            r = _rank(main.iat[i, c])
            if r is not None:
                snaps.append((k, asin, r))
    db.flush()
    for k, asin, r in snaps:
        db.add(md.RankSnapshot(keyword_id=k.id, asin=asin, checked_at=snap,
                               organic_rank=r, method="migrate"))

    # -- Listing Audit markers -> KeywordUsage
    n_usage = 0
    la = S.get("listing audit")
    if la is not None and len(la) > 3:
        lhdr = None
        for i in range(min(10, len(la))):
            if _s(la.iat[i, 0]) == "Search Terms":
                lhdr = i
                break
        if lhdr is not None:
            elem_cols = {}
            for c in range(2, la.shape[1]):
                e = _LA_ELEM.get(_norm_label(la.iat[lhdr, c]))
                if e:
                    elem_cols[c] = e
            for i in range(lhdr + 1, len(la)):
                kw = _s(la.iat[i, 0])
                if kw is None:
                    continue
                k = kw_rows.get(_norm(kw))
                if k is None:
                    continue
                for c, e in elem_cols.items():
                    mk = _s(la.iat[i, c])
                    if mk and mk.lower() in ("exact", "broad"):
                        db.add(md.KeywordUsage(keyword_id=k.id, element=e, match=mk.lower()))
                        n_usage += 1

    # -- Listing Copy: Current/Proposed text blocks
    n_copy = 0
    lc = S.get("listing copy")
    if lc is not None and len(lc) > 2:
        variant = None
        for c in range(lc.shape[1]):
            v0 = _s(lc.iat[0, c])
            if v0 in ("Current", "Proposed"):
                variant = v0.lower()
            e = _LC_ELEM.get(_norm_label(lc.iat[2, c]))
            if variant and e:
                text = _s(lc.iat[1, c]) or (_s(lc.iat[1, c + 1]) if c + 1 < lc.shape[1] else None)
                if text and not db.query(md.ListingCopy).filter_by(
                        project_id=proj.id, variant=variant, element=e).first():
                    db.add(md.ListingCopy(project_id=proj.id, variant=variant, element=e, text=text))
                    n_copy += 1

    # -- BP Comparison row 1 -> competitor bullet_points text
    bp = S.get("bp comparison")
    if bp is not None and len(bp) > 4:
        arow = None
        for i in range(min(8, len(bp))):
            vals = [_s(bp.iat[i, c]) for c in range(3, bp.shape[1])]
            if any(v and _ASIN_RE.match(v) for v in vals):
                arow = i
                break
        if arow is not None:
            for c in range(3, bp.shape[1]):
                asin = _s(bp.iat[arow, c])
                if asin and _ASIN_RE.match(asin) and asin in comp_rows:
                    comp_rows[asin].bullet_points = _s(bp.iat[1, c])

    # -- Search Terms tab: extra keywords (grid confirmed empty -> list only)
    n_st = 0
    st = S.get("search terms")
    if st is not None:
        for i in range(len(st)):
            kw = _s(st.iat[i, 0])
            if kw is None or _norm(kw) in kw_rows or _int(st.iat[i, 1]) is None:
                continue
            k = md.TrackedKeyword(project_id=proj.id, keyword=kw,
                                  search_volume=_int(st.iat[i, 1]),
                                  relevancy=_int(st.iat[i, 2]), source="search_terms")
            db.add(k)
            kw_rows[_norm(kw)] = k
            n_st += 1

    db.commit()
    return {"project_id": proj.id, "name": proj.name, "primary_asin": primary_asin,
            "competitors": len(comp_rows), "keywords": len(kw_rows),
            "rank_cells": len(snaps), "usage_markers": n_usage,
            "copy_blocks": n_copy, "search_terms_added": n_st,
            "snapshot_date": snap.isoformat()}


# ---- raw-data project flow ----------------------------------------------------
# The Listing Optimizer path: no pre-computed sheet needed. Create a blank
# project, upload RAW exports (Cerebro keywords/ranks, X-ray competitors),
# paste raw listing copy — the app computes the analysis (listing-audit
# match markers, exact counts, SV coverage) itself via match_kind().

COPY_ELEMENTS = ELEMENTS[:6]   # the pasteable listing elements (no sp_/sb_ targeting)
TOP_COMPETITORS = 10           # active competitor cap (top by revenue; primary always on)


def _tracked_asins(db: Session, proj: "md.TrackerProject") -> set[str]:
    """ASINs the project actually tracks = every ASIN with a Cerebro/manual rank
    snapshot, plus the primary. The X-ray import matches against this set."""
    kw_ids = [k.id for k in db.query(md.TrackedKeyword.id)
              .filter(md.TrackedKeyword.project_id == proj.id)]
    asins = set()
    if kw_ids:
        asins = {r[0] for r in db.query(md.RankSnapshot.asin).distinct()
                 .filter(md.RankSnapshot.keyword_id.in_(kw_ids))}
    if proj.primary_asin:
        asins.add(proj.primary_asin)
    return asins


def create_project(db: Session, name: str, primary_asin: str | None = None) -> dict:
    """Blank project for the raw-data flow (Cerebro/X-ray/copy uploaded after)."""
    pname = (name or "").strip()[:80]
    if not pname:
        raise ValueError("Project name required.")
    if db.query(md.TrackerProject).filter(md.TrackerProject.name == pname).first():
        raise ValueError(f"A project named '{pname}' already exists.")
    asin = (primary_asin or "").strip().upper() or None
    if asin and not _ASIN_RE.match(asin):
        raise ValueError("Primary ASIN must look like B0XXXXXXXX.")
    proj = md.TrackerProject(name=pname, primary_asin=asin)
    db.add(proj)
    db.commit()
    return {"project_id": proj.id, "name": proj.name, "primary_asin": proj.primary_asin}


def _apply_matching(db: Session, proj: "md.TrackerProject") -> int:
    """Re-sync competitor `active` flags to the tracked ASIN set. Matched
    non-primary competitors capped to the top TOP_COMPETITORS by revenue;
    unmatched rows are KEPT but inactive (hidden from every view) so upload
    order doesn't matter — a later Cerebro upload re-activates them. Returns
    the number of active rows."""
    tracked = _tracked_asins(db, proj)
    comps = db.query(md.TrackedCompetitor).filter_by(project_id=proj.id).all()
    for c in comps:
        c.is_primary = (c.asin == proj.primary_asin)
    matched = sorted((c for c in comps if c.asin in tracked and not c.is_primary),
                     key=lambda c: c.revenue or 0, reverse=True)
    keep = {c.id for c in matched[:TOP_COMPETITORS]}
    n = 0
    for c in comps:
        c.active = bool(c.is_primary or c.id in keep)
        n += c.active
    return n


def import_xray(db: Session, project_id: int, path: str) -> dict:
    """Raw Helium10 X-ray export (csv/xlsx) -> TrackedCompetitor upsert.

    ALL rows are stored, but Cerebro is the source of truth for WHICH ASINs are
    displayed: only rows whose ASIN appears in the project's rank snapshots (or
    is the primary) go active — the rest of the (often 80+) X-ray rows stay
    hidden until a Cerebro upload tracks them (upload order doesn't matter).
    Imported attributes overwrite; manual audit fields (Yes/None rows) are
    preserved. Of the matched, only the top TOP_COMPETITORS by revenue stay
    active (primary always)."""
    proj = db.get(md.TrackerProject, project_id)
    if not proj:
        raise ValueError("Unknown tracker project.")
    df = (pd.read_csv(path, dtype=str, header=None) if str(path).lower().endswith(".csv")
          else pd.read_excel(path, dtype=str, header=None, engine=workbook.excel_engine()))
    recs = _xray_rows(df)
    if not recs:
        raise ValueError("No ASIN rows found — is this an X-ray export?")
    existing = {c.asin: c for c in db.query(md.TrackedCompetitor)
                .filter(md.TrackedCompetitor.project_id == project_id)}
    added = updated = 0
    for rec in recs:
        c = existing.get(rec["asin"])
        if c is None:
            c = md.TrackedCompetitor(project_id=proj.id,
                                     is_primary=(rec["asin"] == proj.primary_asin), **rec)
            db.add(c)
            existing[rec["asin"]] = c
            added += 1
        else:
            for f, v in rec.items():
                if f != "asin" and v is not None:
                    setattr(c, f, v)
            updated += 1
    db.flush()
    tracked = _tracked_asins(db, proj)
    matched = sum(1 for r in recs if r["asin"] in tracked)
    active = _apply_matching(db, proj)
    db.commit()
    return {"project_id": proj.id, "matched": matched, "skipped": len(recs) - matched,
            "added": added, "updated": updated, "active": active}


# pasteable competitor elements — no search_terms (no competitor backend data)
COMP_COPY_ELEMENTS = [e for e in COPY_ELEMENTS if e != "search_terms"]


def _ensure_listing_schema(db: Session) -> None:
    """Older base dbs predate ListingCopy.asin — ALTER it in (idempotent)."""
    cols = [r[1] for r in db.execute(_sqltext("PRAGMA table_info(listing_copy)"))]
    if cols and "asin" not in cols:
        db.execute(_sqltext("ALTER TABLE listing_copy ADD COLUMN asin VARCHAR"))
        db.commit()


def set_listing_copy(db: Session, project_id: int, element: str, text: str | None,
                     variant: str = "current", asin: str | None = None) -> dict:
    """Save raw listing copy for one element (empty text deletes the block).
    asin None = our listing; asin set = that competitor's copy (manual paste)."""
    _ensure_listing_schema(db)
    if not db.get(md.TrackerProject, project_id):
        raise ValueError("Unknown tracker project.")
    if element not in COPY_ELEMENTS:
        raise ValueError(f"Element must be one of: {', '.join(COPY_ELEMENTS)}")
    if variant not in ("current", "proposed"):
        raise ValueError("Variant must be 'current' or 'proposed'.")
    if asin:
        asin = asin.strip().upper()
        if element not in COMP_COPY_ELEMENTS:
            raise ValueError("No search-term data for competitors — paste Title, "
                             "Bullet Points, A+, Description or Alt Text.")
        known = db.query(md.TrackedCompetitor).filter_by(
            project_id=project_id, asin=asin).first()
        if not known:
            raise ValueError(f"{asin} is not a tracked competitor of this project.")
    row = db.query(md.ListingCopy).filter_by(
        project_id=project_id, variant=variant, element=element,
        asin=asin or None).first()
    text = (text or "").strip() or None
    if text is None:
        if row:
            db.delete(row)
    else:
        if row is None:
            row = md.ListingCopy(project_id=project_id, variant=variant,
                                 element=element, asin=asin or None)
            db.add(row)
        row.text = text
    db.commit()
    return {"project_id": project_id, "variant": variant, "element": element,
            "asin": asin or None, "chars": len(text) if text else 0}


def listing_audit(db: Session, project_id: int, variant: str = "current") -> dict:
    """The computed listing analysis (replaces the sheet's hand-marked Listing
    Audit tab): for every tracked keyword x element, match markers are computed
    LIVE from the raw copy via match_kind(). Elements without pasted copy fall
    back to sheet-imported KeywordUsage markers (e.g. sp_/sb_ targeting)."""
    _ensure_listing_schema(db)
    proj = db.get(md.TrackerProject, project_id)
    if not proj:
        raise ValueError("Unknown tracker project.")
    all_copy = (db.query(md.ListingCopy)
                .filter_by(project_id=project_id, variant=variant).all())
    copies = {c.element: (c.text or "") for c in all_copy
              if c.asin is None and (c.text or "").strip()}
    comp_copy: dict[str, dict[str, str]] = {}      # asin -> element -> text
    for c in all_copy:
        if c.asin and (c.text or "").strip():
            comp_copy.setdefault(c.asin, {})[c.element] = c.text
    kws = (db.query(md.TrackedKeyword)
           .filter(md.TrackedKeyword.project_id == project_id)
           .order_by(md.TrackedKeyword.search_volume.desc().nullslast()).all())
    usage: dict[tuple[int, str], str] = {}
    kw_ids = [k.id for k in kws]
    if kw_ids:
        for u in db.query(md.KeywordUsage).filter(md.KeywordUsage.keyword_id.in_(kw_ids)):
            usage[(u.keyword_id, u.element)] = u.match
    # normalize each copy text ONCE (match_kind re-folds the full text per call —
    # 10k keywords x 6 elements made that quadratic-slow on real Cerebro lists)
    pre = {}
    for e, txt in copies.items():
        t = _norm(txt)
        pre[e] = (t, set(t.split()))

    def _match(prenorm, kw_norm: str) -> str | None:
        t, words = prenorm
        if not t or not kw_norm:
            return None
        if kw_norm in t:
            return "exact"
        if all(w in words for w in kw_norm.split()):
            return "broad"
        return None

    def _mk(e: str, kw_norm: str) -> str | None:
        return _match(pre[e], kw_norm)

    # competitor copy, same normalize-once treatment (comparison columns)
    comp_pre = {a: {e: ((t := _norm(txt)), set(t.split()))
                    for e, txt in elems.items()}
                for a, elems in comp_copy.items()}
    comp_stats = {a: {"covered": 0,
                      "elements": {e: {"exact": 0, "broad": 0, "sv": 0} for e in pres}}
                  for a, pres in comp_pre.items()}

    rows = []
    stats = {e: {"exact": 0, "broad": 0, "sv": 0} for e in ELEMENTS}
    for k in kws:
        kwn = _norm(k.keyword)
        marks = {}
        for e in ELEMENTS:
            # sheet-imported usage markers describe the LIVE listing — they only
            # back-fill the 'current' variant; a proposed draft scores pasted copy only
            mk = _mk(e, kwn) if e in copies else (
                usage.get((k.id, e)) if variant == "current" else None)
            if mk:
                marks[e] = mk
                if mk == "exact":
                    stats[e]["exact"] += 1
                    stats[e]["sv"] += k.search_volume or 0
                elif mk == "broad":
                    stats[e]["broad"] += 1
        for a, pres in comp_pre.items():
            hit = False
            cs = comp_stats[a]
            for e, pn in pres.items():
                mk = _match(pn, kwn)
                if mk:
                    hit = True
                    if mk == "exact":
                        cs["elements"][e]["exact"] += 1
                        cs["elements"][e]["sv"] += k.search_volume or 0
                    else:
                        cs["elements"][e]["broad"] += 1
            if hit:
                cs["covered"] += 1
        rows.append({"keyword_id": k.id, "keyword": k.keyword, "sv": k.search_volume,
                     "marks": marks,
                     "n_exact": sum(1 for m in marks.values() if m == "exact")})
    elements = [{
        "element": e, "has_copy": e in copies,
        "chars": len(copies[e]) if e in copies else 0,
        "exact": stats[e]["exact"], "broad": stats[e]["broad"],
        "total_exact_sv": stats[e]["sv"],
    } for e in ELEMENTS]
    covered = sum(1 for r in rows if r["marks"])
    uncovered = [{"keyword": r["keyword"], "sv": r["sv"]}
                 for r in rows if not r["marks"]][:25]
    # competitor comparison columns — every active non-primary competitor shows
    # (empty ones are the paste targets); search_terms excluded by design
    comps = (db.query(md.TrackedCompetitor)
             .filter(md.TrackedCompetitor.project_id == project_id,
                     md.TrackedCompetitor.active == True,          # noqa: E712
                     md.TrackedCompetitor.is_primary == False)     # noqa: E712
             .order_by(md.TrackedCompetitor.revenue.desc().nullslast()).all())
    competitors = [{
        "asin": c.asin, "brand": c.brand, "image_url": c.image_url,
        "covered": comp_stats.get(c.asin, {}).get("covered", 0),
        "elements": [{
            "element": e,
            "has_copy": e in comp_copy.get(c.asin, {}),
            "chars": len(comp_copy.get(c.asin, {}).get(e, "")),
            "exact": comp_stats.get(c.asin, {}).get("elements", {}).get(e, {}).get("exact", 0),
            "broad": comp_stats.get(c.asin, {}).get("elements", {}).get(e, {}).get("broad", 0),
            "total_exact_sv": comp_stats.get(c.asin, {}).get("elements", {}).get(e, {}).get("sv", 0),
        } for e in COMP_COPY_ELEMENTS],
    } for c in comps]
    return {"project": {"id": proj.id, "name": proj.name, "primary_asin": proj.primary_asin},
            "variant": variant,
            "copy": {e: copies.get(e, "") for e in COPY_ELEMENTS},
            "comp_copy": comp_copy,
            "elements": elements, "rows": rows, "competitors": competitors,
            "coverage": {"keywords": len(rows), "covered": covered,
                         "uncovered_top": uncovered}}


# ---- listing sanitizer (Amazon banned keywords) -------------------------------
# User maintains the banned list; the checker scans OUR OWN listing copy only
# (asin NULL — competitors are never scanned) and flags every banned phrase
# present per element. Word-boundary matching on punctuation-folded text, so
# "free" flags "BPA-Free!" but not "freedom".

_san_fold_re = re.compile(r"[^a-z0-9]+")

def _san_norm(s: str) -> str:
    return _san_fold_re.sub(" ", (s or "").lower()).strip()


def set_banned(db: Session, text: str | None) -> dict:
    """Replace the banned-keyword list. Accepts newline/comma-separated phrases;
    blanks and duplicates (case/punctuation-folded) dropped. Empty text clears."""
    phrases, seen = [], set()
    for raw in re.split(r"[\n,;]+", text or ""):
        p = raw.strip()
        key = _san_norm(p)
        if p and key and key not in seen:
            seen.add(key)
            phrases.append(p)
    db.query(md.BannedKeyword).delete()
    for p in phrases:
        db.add(md.BannedKeyword(phrase=p))
    db.commit()
    return {"count": len(phrases)}


def get_banned(db: Session) -> list[str]:
    return [b.phrase for b in db.query(md.BannedKeyword).order_by(md.BannedKeyword.phrase)]


_kw_delim_re = re.compile(r"[,;\n]+")


_KW_SOUP_WORDS = 6      # a "phrase" longer than this is Amazon word-soup


def _catalog_keywords(search_terms: list[str] | None) -> list[str]:
    """Tokenize Category-Listings-Report generic-keyword blobs into trackable
    keywords. Split on explicit delimiters (comma/semicolon/newline) first; any
    segment still longer than _KW_SOUP_WORDS words is Amazon's undelimited
    word-soup (a stray comma inside soup must not yield a 400-char "phrase") and
    splits into single words — Amazon indexes the field word-by-word anyway.
    Short segments stay whole phrases."""
    out, seen = [], set()

    def add(s):
        s = s.strip()
        if len(s) < 2 or s.isdigit():
            return
        key = _norm(s)
        if key and key not in seen:
            seen.add(key)
            out.append(s)

    for blob in search_terms or []:
        for seg in _kw_delim_re.split((blob or "").strip()):
            seg = seg.strip()
            if not seg:
                continue
            words = seg.split()
            if len(words) > _KW_SOUP_WORDS:
                for w in words:
                    add(w)
            else:
                add(seg)
    return out


def import_catalog_copy(db: Session, project_id: int, product: dict) -> dict:
    """Prefill the project's OWN listing copy from a Product Benchmark catalog
    product (Category Listings Report row): title / bullet points / description /
    search terms. Empty catalog elements are SKIPPED (never wipe a manual paste);
    non-empty ones replace the current-variant own copy. The product's search
    terms are also added as TRACKED keywords (source='search_terms') so the SEO
    tab covers them — deduped against the project's existing keyword list.
    A project without a primary ASIN adopts the product's ASIN (never overwrites
    an existing one)."""
    proj = db.get(md.TrackerProject, project_id)
    if not proj:
        raise ValueError("Unknown tracker project.")
    asin = (product.get("asin") or "").strip().upper()
    if asin and not proj.primary_asin:
        proj.primary_asin = asin
        db.commit()
    elements = {
        "title": (product.get("title") or "").strip(),
        "bullet_points": "\n".join(product.get("bullets") or []).strip(),
        "description": (product.get("description") or "").strip(),
        "search_terms": " ".join(product.get("search_terms") or []).strip(),
    }
    imported = []
    for el, text in elements.items():
        if text:
            set_listing_copy(db, project_id, el, text)   # own copy (asin None)
            imported.append(el)

    kw_added = kw_known = 0
    words = _catalog_keywords(product.get("search_terms"))
    if words:
        existing = {_norm(k.keyword) for k in db.query(md.TrackedKeyword)
                    .filter(md.TrackedKeyword.project_id == project_id)}
        for w in words:
            if _norm(w) in existing:
                kw_known += 1
            else:
                db.add(md.TrackedKeyword(project_id=project_id, keyword=w,
                                         source="search_terms"))
                existing.add(_norm(w))
                kw_added += 1
        db.commit()
    return {"project_id": project_id, "sku": product.get("sku"),
            "asin": product.get("asin"), "title": product.get("title"),
            "imported": imported,
            "skipped": [e for e in elements if e not in imported],
            "keywords_added": kw_added, "keywords_known": kw_known}


def find_banned(text: str, phrases: list[str]) -> list[str]:
    """Banned phrases present in `text` (whole-word, punctuation/case folded)."""
    hay = f" {_san_norm(text)} "
    if not hay.strip():
        return []
    return [p for p in phrases if (n := _san_norm(p)) and f" {n} " in hay]


def sanitize(db: Session, project_id: int, variant: str = "current") -> dict:
    """Scan OUR listing copy for banned keywords. Own copy only (asin NULL) —
    title, bullet points, A+, description, search terms, alt text."""
    _ensure_listing_schema(db)
    proj = db.get(md.TrackerProject, project_id)
    if not proj:
        raise ValueError("Unknown tracker project.")
    phrases = get_banned(db)
    copies = {c.element: c.text for c in db.query(md.ListingCopy)
              .filter_by(project_id=project_id, variant=variant)
              if c.asin is None and (c.text or "").strip()}
    elements = []
    for e in COPY_ELEMENTS:
        if e in copies:
            flags = find_banned(copies[e], phrases)
            elements.append({"element": e, "checked": True,
                             "flagged": len(flags), "phrases": flags})
        else:
            elements.append({"element": e, "checked": False,
                             "flagged": 0, "phrases": []})
    return {"project_id": project_id, "banned_count": len(phrases),
            "total_flagged": sum(x["flagged"] for x in elements),
            "elements": elements}


# ---- ongoing snapshot import (Cerebro export) --------------------------------

def import_cerebro(db: Session, project_id: int, path: str,
                   checked_at: date | None = None, asin: str | None = None) -> dict:
    """Cerebro CSV/xlsx -> RankSnapshot rows. Single-ASIN export (Organic Rank /
    Position (Rank) column, needs `asin`) or multi-ASIN (per-ASIN rank columns).
    Unknown keywords are added to the master list. Re-import of the same date
    replaces that date's rows for the touched ASINs only."""
    proj = db.get(md.TrackerProject, project_id)
    if not proj:
        raise ValueError("Unknown tracker project.")
    snap = checked_at or date.today()
    df = (pd.read_csv(path, dtype=str) if str(path).lower().endswith(".csv")
          else pd.read_excel(path, dtype=str, engine=workbook.excel_engine()))
    df.columns = [str(c).replace("\ufeff", "").strip() for c in df.columns]
    cols = {c.lower(): c for c in df.columns}
    kw_col = cols.get("keyword phrase") or cols.get("keyword") or cols.get("search terms")
    if not kw_col:
        raise ValueError("No 'Keyword Phrase' column — is this a Cerebro export?")
    sv_col = cols.get("search volume")
    sales_col = cols.get("keyword sales")
    rank_col = cols.get("organic rank") or cols.get("position (rank)") or cols.get("rank")
    asin_cols = {c: c for c in df.columns if _ASIN_RE.match(str(c).strip())}
    if rank_col and not asin:
        asin = proj.primary_asin       # rank column = the exported ASIN's own ranks
    if not asin_cols and not (rank_col and asin):
        raise ValueError("No per-ASIN rank columns, and no target ASIN given for the "
                         "single-ASIN rank column.")

    kws = {_norm(k.keyword): k for k in db.query(md.TrackedKeyword)
           .filter(md.TrackedKeyword.project_id == project_id)}
    touched_asins = set(asin_cols) | ({asin} if rank_col and asin else set())
    touched_kw_ids = [k.id for k in kws.values()]
    if touched_kw_ids:
        db.query(md.RankSnapshot).filter(
            md.RankSnapshot.keyword_id.in_(touched_kw_ids),
            md.RankSnapshot.checked_at == snap,
            md.RankSnapshot.asin.in_(list(touched_asins))).delete(synchronize_session=False)

    added_kw, n_rows = 0, 0
    for _, r in df.iterrows():
        kw = _s(r[kw_col])
        if kw is None:
            continue
        k = kws.get(_norm(kw))
        if k is None:
            k = md.TrackedKeyword(project_id=project_id, keyword=kw,
                                  search_volume=_int(r[sv_col]) if sv_col else None,
                                  keyword_sales=_int(r[sales_col]) if sales_col else None,
                                  source="cerebro")
            db.add(k)
            db.flush()
            kws[_norm(kw)] = k
            added_kw += 1
        else:
            if sv_col and _int(r[sv_col]) is not None:
                k.search_volume = _int(r[sv_col])
            if sales_col and _int(r[sales_col]) is not None:
                k.keyword_sales = _int(r[sales_col])
        if rank_col and asin:
            rk = _rank(r[rank_col])
            if rk is not None:
                db.add(md.RankSnapshot(keyword_id=k.id, asin=asin, checked_at=snap,
                                       organic_rank=rk, method="cerebro"))
                n_rows += 1
        for ac in asin_cols:
            rk = _rank(r[ac])
            if rk is not None:
                db.add(md.RankSnapshot(keyword_id=k.id, asin=str(ac).strip(),
                                       checked_at=snap, organic_rank=rk, method="cerebro"))
                n_rows += 1
    db.flush()
    _apply_matching(db, proj)   # newly tracked ASINs re-activate stored X-ray rows
    db.commit()
    return {"snapshot_date": snap.isoformat(), "asins": sorted(touched_asins),
            "rank_rows": n_rows, "keywords_added": added_kw}


# ---- views -------------------------------------------------------------------

def _dates(db: Session, project_id: int) -> list[date]:
    kw_ids = [k.id for k in db.query(md.TrackedKeyword.id)
              .filter(md.TrackedKeyword.project_id == project_id)]
    if not kw_ids:
        return []
    rows = (db.query(md.RankSnapshot.checked_at).distinct()
            .filter(md.RankSnapshot.keyword_id.in_(kw_ids)).all())
    return sorted(r[0] for r in rows)


def _ranks_asof(db: Session, project_id: int, when: date) -> dict[tuple[int, str], int | None]:
    """(keyword_id, asin) -> rank from the LATEST snapshot at/before `when`."""
    kw_ids = [k.id for k in db.query(md.TrackedKeyword.id)
              .filter(md.TrackedKeyword.project_id == project_id)]
    if not kw_ids:
        return {}
    rows = (db.query(md.RankSnapshot)
            .filter(md.RankSnapshot.keyword_id.in_(kw_ids),
                    md.RankSnapshot.checked_at <= when)
            .order_by(md.RankSnapshot.checked_at).all())
    out = {}
    for r in rows:                     # later dates overwrite earlier (ordered asc)
        out[(r.keyword_id, r.asin)] = r.organic_rank
    return out


def _listing_texts(db: Session, proj: "md.TrackerProject",
                   comps: list) -> dict[str, str]:
    """Normalized listing text per active ASIN — Title + Bullet Points from the
    X-ray/sheet import; the primary additionally gets its pasted Listing Copy
    (title, bullet points, description). Feeds the computed REL column."""
    texts = {}
    for c in comps:
        t = " ".join(p for p in (c.title, c.bullet_points) if p)
        if t.strip():
            texts[c.asin] = _norm(t)
    copies = [c.text for c in db.query(md.ListingCopy)
              .filter(md.ListingCopy.project_id == proj.id,
                      md.ListingCopy.variant == "current",
                      md.ListingCopy.element.in_(["title", "bullet_points", "description"]))
              if (c.text or "").strip()]
    if proj.primary_asin and copies:
        texts[proj.primary_asin] = _norm(
            " ".join([texts.get(proj.primary_asin, ""), *copies]))
    return texts


def relevancy_counts(texts: dict[str, str], keywords: list) -> dict[int, int]:
    """REL per keyword = how many ASINs' listing text (title / bullet points /
    description) contains the keyword as an exact phrase."""
    out = {}
    for k in keywords:
        kw = _norm(k.keyword)
        out[k.id] = sum(1 for t in texts.values() if kw in t) if kw else 0
    return out


def matrix(db: Session, project_id: int, when: date | None = None) -> dict:
    """The sheet's core view: keyword rows x ASIN columns, rank per cell.
    Cell: int rank | None (checked, unranked at import time it isn't stored —
    so None here means 'no data'); page-1/2-3 heat computed client-side.
    REL is computed live: count of ASINs whose listing text uses the keyword."""
    proj = db.get(md.TrackerProject, project_id)
    if not proj:
        raise ValueError("Unknown tracker project.")
    dates = _dates(db, project_id)
    when = when or (dates[-1] if dates else date.today())
    comps = (db.query(md.TrackedCompetitor)
             .filter(md.TrackedCompetitor.project_id == project_id,
                     md.TrackedCompetitor.active == True)  # noqa: E712
             .order_by(md.TrackedCompetitor.is_primary.desc(),
                       md.TrackedCompetitor.revenue.desc().nullslast()).all())
    asins = [c.asin for c in comps]
    ranks = _ranks_asof(db, project_id, when)
    kws = (db.query(md.TrackedKeyword)
           .filter(md.TrackedKeyword.project_id == project_id)
           .order_by(md.TrackedKeyword.search_volume.desc().nullslast()).all())
    rel = relevancy_counts(_listing_texts(db, proj, comps), kws)
    rows = [{"keyword_id": k.id, "keyword": k.keyword, "sv": k.search_volume,
             "relevancy": rel.get(k.id, 0), "kw_sales": k.keyword_sales,
             "ranks": [ranks.get((k.id, a)) for a in asins]} for k in kws]
    return {"project": {"id": proj.id, "name": proj.name, "primary_asin": proj.primary_asin},
            "date": when.isoformat(), "dates": [d.isoformat() for d in dates],
            "asins": asins,
            "columns": [{"asin": c.asin, "brand": c.brand, "is_primary": c.is_primary}
                        for c in comps],
            "rows": rows}


def scorecard(db: Session, project_id: int) -> dict:
    """Per-ASIN SEO rollups at the latest snapshot + per-date trend series."""
    proj = db.get(md.TrackerProject, project_id)
    if not proj:
        raise ValueError("Unknown tracker project.")
    dates = _dates(db, project_id)
    comps = (db.query(md.TrackedCompetitor)
             .filter(md.TrackedCompetitor.project_id == project_id,
                     md.TrackedCompetitor.active == True)  # noqa: E712
             .order_by(md.TrackedCompetitor.is_primary.desc(),
                       md.TrackedCompetitor.revenue.desc().nullslast()).all())
    tracked = db.query(md.TrackedKeyword).filter(
        md.TrackedKeyword.project_id == project_id).count()
    series = {c.asin: [] for c in comps}
    cards = []
    for d in dates:
        ranks = _ranks_asof(db, project_id, d)
        for c in comps:
            rs = [r for (kid, a), r in ranks.items() if a == c.asin and r is not None]
            series[c.asin].append({
                "date": d.isoformat(), "ranked": len(rs),
                "page1": sum(1 for r in rs if r <= PAGE_SIZE),
                "top10": sum(1 for r in rs if r <= 10),
                "avg_rank": round(sum(rs) / len(rs), 1) if rs else None})
    best_p1 = 0
    for c in comps:
        last = series[c.asin][-1] if series[c.asin] else None
        if last and not c.is_primary:
            best_p1 = max(best_p1, last["page1"])
    for c in comps:
        last = series[c.asin][-1] if series[c.asin] else None
        cards.append({
            "asin": c.asin, "brand": c.brand, "is_primary": c.is_primary,
            "kw_tracked": tracked,
            "kw_ranked": last["ranked"] if last else 0,
            "index_rate": index_rate(last["ranked"], tracked) if last else None,
            "page1_count": last["page1"] if last else 0,
            "top10_count": last["top10"] if last else 0,
            "avg_rank": last["avg_rank"] if last else None,
            "coverage_vs_best": (round(last["page1"] / best_p1, 4)
                                 if last and c.is_primary and best_p1 else None),
            "series": series[c.asin]})
    # sheet-header KPIs, recomputed live
    top = [c for c in comps if not c.is_primary][:10]
    total_rev = sum(c.revenue or 0 for c in top)
    prim = next((c for c in comps if c.is_primary), None)
    return {"project": {"id": proj.id, "name": proj.name, "primary_asin": proj.primary_asin},
            "dates": [d.isoformat() for d in dates], "cards": cards,
            "kpis": {"total_revenue_top10": round(total_rev, 2),
                     "market_share": (round((prim.revenue or 0) / total_rev, 6)
                                      if prim and total_rev else None),
                     "avg_reviews_top10": (round(sum(c.review_count or 0 for c in top)
                                                 / len(top), 1) if top else None)}}


def movers(db: Session, project_id: int, top: int = 25) -> dict:
    """Primary-ASIN rank deltas between the two latest snapshot dates."""
    proj = db.get(md.TrackerProject, project_id)
    if not proj:
        raise ValueError("Unknown tracker project.")
    dates = _dates(db, project_id)
    if len(dates) < 2:
        return {"dates": [d.isoformat() for d in dates], "climbers": [], "decliners": [],
                "new": [], "lost": [], "note": "Need 2+ snapshot dates to compare."}
    prev_d, cur_d = dates[-2], dates[-1]
    prev = _ranks_asof(db, project_id, prev_d)
    cur = _ranks_asof(db, project_id, cur_d)
    asin = proj.primary_asin
    kws = {k.id: k for k in db.query(md.TrackedKeyword)
           .filter(md.TrackedKeyword.project_id == project_id)}
    climbers, decliners, new, lost = [], [], [], []
    for kid, k in kws.items():
        p, c = prev.get((kid, asin)), cur.get((kid, asin))
        base = {"keyword": k.keyword, "sv": k.search_volume, "prev": p, "cur": c}
        if p is None and c is not None:
            new.append(base)                       # absent before -> "new", no fake delta
        elif p is not None and c is None:
            lost.append(base)
        else:
            d = rank_delta(c, p)
            if d is None or d == 0:
                continue
            (climbers if d < 0 else decliners).append({**base, "delta": d})
    climbers.sort(key=lambda x: x["delta"])
    decliners.sort(key=lambda x: -x["delta"])
    new.sort(key=lambda x: (x["cur"] or 9999))
    return {"prev_date": prev_d.isoformat(), "cur_date": cur_d.isoformat(),
            "dates": [d.isoformat() for d in dates],
            "climbers": climbers[:top], "decliners": decliners[:top],
            "new": new[:top], "lost": lost[:top]}


def set_cell(db: Session, keyword_id: int, asin: str, rank: int | None,
             when: date | None = None) -> dict:
    """Manual grid edit (the sheet habit): upsert today's snapshot row."""
    k = db.get(md.TrackedKeyword, keyword_id)
    if not k:
        raise ValueError("Unknown keyword.")
    when = when or date.today()
    row = (db.query(md.RankSnapshot)
           .filter_by(keyword_id=keyword_id, asin=asin, checked_at=when).first())
    if row is None:
        row = md.RankSnapshot(keyword_id=keyword_id, asin=asin, checked_at=when,
                              method="manual")
        db.add(row)
    row.organic_rank = rank if rank and rank > 0 else None
    row.method = "manual"
    db.commit()
    return {"keyword_id": keyword_id, "asin": asin, "date": when.isoformat(),
            "organic_rank": row.organic_rank}


def suggest(db: Session, project_id: int, max_page: int = 3, min_sv: int = 500) -> dict:
    """PPC bridge: rank-support keywords (primary ranked page 2..max_page, decent
    volume -> push with exact targets) + competitor ASINs as PT suggestions."""
    proj = db.get(md.TrackerProject, project_id)
    if not proj:
        raise ValueError("Unknown tracker project.")
    dates = _dates(db, project_id)
    if not dates:
        return {"rank_support": [], "product_targets": []}
    ranks = _ranks_asof(db, project_id, dates[-1])
    kws = {k.id: k for k in db.query(md.TrackedKeyword)
           .filter(md.TrackedKeyword.project_id == project_id)}
    support = []
    for kid, k in kws.items():
        r = ranks.get((kid, proj.primary_asin))
        p = page(r)
        if p is not None and 2 <= p <= max_page and (k.search_volume or 0) >= min_sv:
            support.append({"keyword": k.keyword, "sv": k.search_volume, "rank": r,
                            "page": p, "reason": f"page {p} organic — exact-target to push to page 1"})
    support.sort(key=lambda x: -(x["sv"] or 0))
    pts = [{"asin": c.asin, "brand": c.brand, "review_count": c.review_count,
            "rating": c.rating, "price": c.price,
            "reason": "competitor — product-target its detail page"}
           for c in db.query(md.TrackedCompetitor)
           .filter(md.TrackedCompetitor.project_id == project_id,
                   md.TrackedCompetitor.active == True,   # noqa: E712
                   md.TrackedCompetitor.is_primary == False)  # noqa: E712
           .order_by(md.TrackedCompetitor.revenue.desc().nullslast())]
    return {"date": dates[-1].isoformat(), "rank_support": support, "product_targets": pts}


# ---- SEO + backend search-term recommendations --------------------------------
_ST_MAX_BYTES = 249          # Amazon indexes the generic-keywords field up to 249 bytes
_TITLE_MAX = 200             # hard cap on most categories
_TITLE_THIN = 80             # under this the title is leaving keyword room unused
_SEV_ORDER = {"high": 0, "medium": 1, "low": 2}


def seo_recommend(db: Session, project_id: int, variant: str = "current") -> dict:
    """SEO recommendations + a backend search-term recommendation, computed from
    the Listing Audit (tracked keywords vs pasted copy) + the banned list.

    Recommendations = prioritized actions: missing copy, top-SV keywords not
    exact in the title / uncovered anywhere, length checks (title chars,
    search-terms bytes, bullet count), banned phrases, wasted backend words
    (already indexed by visible copy).

    Search-term recommendation = a ready-to-paste generic-keywords line: the
    highest-SV tracked-keyword WORDS not already in the visible copy (Amazon
    indexes title/bullets/description word-by-word, so repeating them back
    there wastes bytes), deduped, banned-free, packed to the 249-byte limit.
    """
    audit = listing_audit(db, project_id, variant)
    copy = audit["copy"]
    rows = audit["rows"]                      # ordered by SV desc (listing_audit)
    recs: list[dict] = []

    def rec(severity, area, title, action, keywords=None):
        recs.append({"severity": severity, "area": area, "title": title,
                     "action": action, "keywords": keywords or []})

    # 1) copy elements never pasted — nothing to audit / index there
    has = {e["element"]: e for e in audit["elements"]}
    for e, sev in (("title", "high"), ("search_terms", "high"),
                   ("bullet_points", "medium"), ("description", "medium")):
        if e in has and not has[e]["has_copy"]:
            rec(sev, e, f"No {e.replace('_', ' ')} pasted",
                "Paste it in the Listing Audit tab so coverage (and these recommendations) can see it.")

    if rows:
        # 2) top keywords not EXACT in the title (title = strongest ranking field)
        if has.get("title", {}).get("has_copy"):
            miss_title = [{"keyword": r["keyword"], "sv": r["sv"]}
                          for r in rows if (r["sv"] or 0) > 0 and r["marks"].get("title") != "exact"][:5]
            if miss_title:
                rec("high", "title", "Top keywords not exact in the title",
                    "Work these phrases verbatim into the title — exact title matches carry the most rank weight.",
                    miss_title)
        # 3) keywords with no coverage anywhere
        uncov = [u for u in audit["coverage"]["uncovered_top"] if (u["sv"] or 0) > 0][:8]
        if uncov:
            rec("high", "coverage", "High-volume keywords not covered anywhere",
                "These tracked keywords appear in no listing element — add them to a bullet, the description or the backend search terms.",
                uncov)

    # 4) length checks
    title_txt = copy.get("title", "")
    if title_txt:
        if len(title_txt) > _TITLE_MAX:
            rec("medium", "title", f"Title is {len(title_txt)} chars (over {_TITLE_MAX})",
                "Most categories cap the title at 200 chars — Amazon may suppress it. Trim it.")
        elif len(title_txt) < _TITLE_THIN:
            rec("low", "title", f"Title is only {len(title_txt)} chars",
                f"There's room up to ~{_TITLE_MAX} — add your top keyword phrases.")
    bullets_txt = copy.get("bullet_points", "")
    if bullets_txt:
        n_bul = len([b for b in bullets_txt.splitlines() if b.strip()])
        if n_bul < 5:
            rec("medium", "bullet_points", f"Only {n_bul} bullet point{'s' if n_bul != 1 else ''}",
                "Amazon gives you 5 — each is indexed. Add the missing ones with uncovered keywords.")
    st_txt = copy.get("search_terms", "")
    st_bytes = len(st_txt.strip().encode("utf-8"))
    if st_txt and st_bytes > _ST_MAX_BYTES:
        rec("high", "search_terms", f"Backend search terms are {st_bytes} bytes (over {_ST_MAX_BYTES})",
            "Amazon ignores the whole field beyond 249 bytes — cut it below the limit.")

    # 5) banned phrases in our copy
    san = sanitize(db, project_id, variant)
    if san["total_flagged"]:
        flagged = sorted({p for el in san["elements"] for p in el["phrases"]})
        rec("high", "compliance", f"{san['total_flagged']} banned phrase hit{'s' if san['total_flagged'] != 1 else ''} in your copy",
            "Remove or replace these before Amazon suppresses the listing (see the Sanitizer on Product Overview).",
            [{"keyword": p, "sv": None} for p in flagged[:10]])

    # 6) backend search-term recommendation (visible copy is already indexed —
    #    the field should spend its 249 bytes on words the copy DOESN'T have)
    visible: set[str] = set()
    for e in ("title", "bullet_points", "description", "aplus", "alt_text"):
        visible |= set(_norm(copy.get(e, "")).split())
    banned_words: set[str] = set()
    for p in get_banned(db):
        banned_words |= set(_san_norm(p).split())
    chosen, seen = [], set()
    for r in rows:
        for w in _norm(r["keyword"]).split():
            if len(w) < 2 or w.isdigit() or w in visible or w in seen or w in banned_words:
                continue
            seen.add(w)
            chosen.append({"word": w, "sv": r["sv"] or 0})
    line, used, size = [], [], 0
    for c in chosen:
        need = len(c["word"].encode("utf-8")) + (1 if line else 0)
        if size + need > _ST_MAX_BYTES:
            continue                     # skip; a shorter later word may still fit
        line.append(c["word"]); used.append(c); size += need

    cur_words = list(dict.fromkeys(_norm(st_txt).split()))
    wasted = [w for w in cur_words if w in visible]
    if wasted:
        rec("medium", "search_terms",
            f"{len(wasted)} backend word{'s' if len(wasted) != 1 else ''} already indexed by your visible copy",
            "Words in the title/bullets/description are already indexed — replace them in the backend field "
            "with the suggested uncovered words below.",
            [{"keyword": w, "sv": None} for w in wasted[:10]])

    recs.sort(key=lambda r: _SEV_ORDER.get(r["severity"], 9))
    counts = {s: sum(1 for r in recs if r["severity"] == s) for s in ("high", "medium", "low")}
    return {"project": audit["project"], "variant": variant,
            "recommendations": recs, "counts": counts,
            "search_terms": {
                "suggested": " ".join(line), "bytes": size, "max_bytes": _ST_MAX_BYTES,
                "words": used, "candidates": len(chosen),
                "current_bytes": st_bytes, "wasted_words": wasted[:15],
            }}


# ---- Product Optimization exec report (xlsx with native Excel charts) ---------
def report_xlsx(db: Session, project_id: int) -> bytes:
    """One client-ready workbook mirroring the three Product Optimization views,
    with native Excel charts: Overview (scorecards + page-1 bar + market-share
    pie), SEO (rank distribution pie, page-1 trend line, movers), Listing Audit
    (element coverage bar + recommendations + the 249-byte backend line) and
    Product Overview (competitor table + revenue bar)."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill

    proj = db.get(md.TrackerProject, project_id)
    if not proj:
        raise ValueError("Unknown tracker project.")
    sc = scorecard(db, project_id)
    mx = matrix(db, project_id)
    mv = movers(db, project_id)
    la = listing_audit(db, project_id)
    rec = seo_recommend(db, project_id)
    comps = [c for c in competitors(db, project_id)["competitors"] if c["active"]]

    TITLE = Font(bold=True, size=14)
    H = Font(bold=True, color="1F2329")
    HFILL = PatternFill("solid", fgColor="FCD535")     # brand yellow headers
    SUB = Font(bold=True, size=11)

    def head(ws, row, labels, widths=None):
        for i, lab in enumerate(labels, start=1):
            c = ws.cell(row=row, column=i, value=lab)
            c.font = H; c.fill = HFILL
            if widths and i <= len(widths) and widths[i - 1]:
                ws.column_dimensions[c.column_letter].width = widths[i - 1]
        return row + 1

    def title(ws, row, text):
        ws.cell(row=row, column=1, value=text).font = SUB
        return row + 1

    wb = Workbook()

    # ---- Overview -------------------------------------------------------------
    ws = wb.active
    ws.title = "Overview"
    ws.cell(row=1, column=1, value=f"Product Optimization report — {proj.name}").font = TITLE
    ws.cell(row=2, column=1, value=f"Primary ASIN {proj.primary_asin or '—'} · generated from the latest snapshot"
            + (f" ({sc['dates'][-1]})" if sc["dates"] else " (no rank snapshots yet)"))
    kpis = sc["kpis"]
    r = 4
    for lab, val in [("Tracked keywords", len(mx["rows"])),
                     ("Competitors (active)", sum(1 for c in comps if not c["is_primary"])),
                     ("Rank snapshots", len(sc["dates"])),
                     ("Top-10 revenue / mo", kpis.get("total_revenue_top10")),
                     ("Market share", kpis.get("market_share")),
                     ("Avg reviews (top 10)", kpis.get("avg_reviews_top10")),
                     ("Copy coverage", f"{la['coverage']['covered']}/{la['coverage']['keywords']} keywords used somewhere"),
                     ("SEO recommendations", f"{rec['counts']['high']} high · {rec['counts']['medium']} medium · {rec['counts']['low']} low")]:
        ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 44

    # scorecard table + page-1 bar chart
    r += 1
    r = title(ws, r, "ASIN scorecards (latest snapshot)")
    hdr_row = r
    r = head(ws, r, ["Brand / ASIN", "You?", "Indexed %", "Page 1", "Top 10", "Avg rank"],
             widths=[28, 6, 11, 9, 9, 10])
    first_data = r
    for c in sc["cards"]:
        ws.cell(row=r, column=1, value=(c["brand"] or c["asin"]))
        ws.cell(row=r, column=2, value="★" if c["is_primary"] else "")
        ws.cell(row=r, column=3, value=c["index_rate"])
        ws.cell(row=r, column=4, value=c["page1_count"])
        ws.cell(row=r, column=5, value=c["top10_count"])
        ws.cell(row=r, column=6, value=c["avg_rank"])
        r += 1
    if sc["cards"]:
        ch = BarChart(); ch.type = "col"; ch.title = "Page-1 keywords by ASIN"; ch.height = 8; ch.width = 18
        ch.add_data(Reference(ws, min_col=4, min_row=hdr_row, max_row=r - 1), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=first_data, max_row=r - 1))
        ws.add_chart(ch, f"H{hdr_row}")
        rev_rows = [c for c in comps if (c["revenue"] or 0) > 0]
        if rev_rows:
            base = r + 2
            ws.cell(row=base - 1, column=1, value="Revenue / mo (X-ray)").font = SUB
            for i, c in enumerate(rev_rows):
                ws.cell(row=base + i, column=1, value=(c["brand"] or c["asin"]) + (" ★" if c["is_primary"] else ""))
                ws.cell(row=base + i, column=2, value=c["revenue"])
            pie = PieChart(); pie.title = "Revenue share"; pie.height = 8; pie.width = 12
            pie.add_data(Reference(ws, min_col=2, min_row=base, max_row=base + len(rev_rows) - 1))
            pie.set_categories(Reference(ws, min_col=1, min_row=base, max_row=base + len(rev_rows) - 1))
            ws.add_chart(pie, f"H{hdr_row + 18}")

    # ---- SEO ------------------------------------------------------------------
    ws = wb.create_sheet("SEO")
    ws.cell(row=1, column=1, value="SEO — rank visibility").font = TITLE
    prim_col = next((i for i, c in enumerate(mx["columns"]) if c["is_primary"]), None)
    buckets = {"Page 1": 0, "Page 2-3": 0, "Deeper": 0, "Unranked": 0}
    if prim_col is not None:
        for row_ in mx["rows"]:
            p = page(row_["ranks"][prim_col])
            if p is None:
                buckets["Unranked"] += 1
            elif p == 1:
                buckets["Page 1"] += 1
            elif p <= 3:
                buckets["Page 2-3"] += 1
            else:
                buckets["Deeper"] += 1
    r = title(ws, 3, f"Your rank distribution · {len(mx['rows'])} tracked keywords · as of {mx['date'] or '—'}")
    r = head(ws, r, ["Bucket", "Keywords"], widths=[16, 12])
    first_data = r
    for k, v in buckets.items():
        ws.cell(row=r, column=1, value=k); ws.cell(row=r, column=2, value=v); r += 1
    pie = PieChart(); pie.title = "Rank distribution (you)"; pie.height = 8; pie.width = 12
    pie.add_data(Reference(ws, min_col=2, min_row=first_data, max_row=r - 1))
    pie.set_categories(Reference(ws, min_col=1, min_row=first_data, max_row=r - 1))
    ws.add_chart(pie, "E3")

    # page-1 trend line (all ASINs over snapshots)
    if len(sc["dates"]) > 1:
        r += 1
        r = title(ws, r, "Page-1 keywords over snapshots")
        hdr_row = r
        labels = ["Date"] + [(c["brand"] or c["asin"]) + (" ★" if c["is_primary"] else "") for c in sc["cards"]]
        r = head(ws, r, labels, widths=[12] + [16] * len(sc["cards"]))
        first_data = r
        by_date = {c["asin"]: {s["date"]: s["page1"] for s in c["series"]} for c in sc["cards"]}
        for d_ in sc["dates"]:
            ws.cell(row=r, column=1, value=d_)
            for j, c in enumerate(sc["cards"], start=2):
                ws.cell(row=r, column=j, value=by_date[c["asin"]].get(d_))
            r += 1
        ln = LineChart(); ln.title = "Page-1 trend"; ln.height = 9; ln.width = 20
        ln.add_data(Reference(ws, min_col=2, max_col=1 + len(sc["cards"]),
                              min_row=hdr_row, max_row=r - 1), titles_from_data=True)
        ln.set_categories(Reference(ws, min_col=1, min_row=first_data, max_row=r - 1))
        ws.add_chart(ln, "E22")

    # movers
    if mv.get("climbers") is not None:
        r += 1
        r = title(ws, r, f"Movers · {mv.get('prev_date', '—')} → {mv.get('cur_date', '—')}")
        r = head(ws, r, ["Direction", "Keyword", "SV", "Prev", "Cur", "Δ"], widths=[10, 34, 10, 8, 8, 8])
        for lab, rows_ in (("▲ climber", mv.get("climbers", [])), ("▼ decliner", mv.get("decliners", [])),
                           ("new", mv.get("new", [])), ("lost", mv.get("lost", []))):
            for m in rows_[:15]:
                ws.cell(row=r, column=1, value=lab)
                ws.cell(row=r, column=2, value=m["keyword"])
                ws.cell(row=r, column=3, value=m["sv"])
                ws.cell(row=r, column=4, value=m["prev"])
                ws.cell(row=r, column=5, value=m["cur"])
                ws.cell(row=r, column=6, value=m.get("delta"))
                r += 1

    # ---- Listing Audit ----------------------------------------------------------
    ws = wb.create_sheet("Listing Audit")
    ws.cell(row=1, column=1, value="Listing Audit — copy coverage & recommendations").font = TITLE
    cov = la["coverage"]
    ws.cell(row=2, column=1, value=f"{cov['covered']} of {cov['keywords']} tracked keywords appear somewhere in the copy")
    r = title(ws, 4, "Element coverage")
    hdr_row = r
    r = head(ws, r, ["Element", "Has copy", "Chars", "Exact", "Broad", "Exact SV"],
             widths=[18, 10, 10, 9, 9, 12])
    first_data = r
    for e in la["elements"]:
        if e["element"] not in COPY_ELEMENTS:
            continue
        ws.cell(row=r, column=1, value=e["element"].replace("_", " "))
        ws.cell(row=r, column=2, value="yes" if e["has_copy"] else "")
        ws.cell(row=r, column=3, value=e["chars"])
        ws.cell(row=r, column=4, value=e["exact"])
        ws.cell(row=r, column=5, value=e["broad"])
        ws.cell(row=r, column=6, value=e["total_exact_sv"])
        r += 1
    ch = BarChart(); ch.type = "col"; ch.grouping = "stacked"; ch.overlap = 100
    ch.title = "Keyword matches per element (exact + broad)"; ch.height = 8; ch.width = 18
    ch.add_data(Reference(ws, min_col=4, max_col=5, min_row=hdr_row, max_row=r - 1), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=first_data, max_row=r - 1))
    ws.add_chart(ch, "H4")

    r += 1
    r = title(ws, r, f"SEO recommendations · {len(rec['recommendations'])}")
    r = head(ws, r, ["Severity", "Area", "Recommendation", "Action", "Keywords"],
             widths=[10, 14, 44, 60, 40])
    for x in rec["recommendations"]:
        ws.cell(row=r, column=1, value=x["severity"])
        ws.cell(row=r, column=2, value=x["area"])
        ws.cell(row=r, column=3, value=x["title"])
        c4 = ws.cell(row=r, column=4, value=x["action"]); c4.alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=5, value=", ".join(k["keyword"] for k in x["keywords"][:8]))
        r += 1
    st = rec["search_terms"]
    r += 1
    r = title(ws, r, f"Suggested backend search terms ({st['bytes']}/{st['max_bytes']} bytes; current field {st['current_bytes']}B)")
    ws.cell(row=r, column=1, value=st["suggested"]).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

    # uncovered top keywords
    if cov["uncovered_top"]:
        r += 2
        r = title(ws, r, "Top keywords not covered anywhere")
        r = head(ws, r, ["Keyword", "SV"], widths=[34, 12])
        for u in cov["uncovered_top"][:15]:
            ws.cell(row=r, column=1, value=u["keyword"]); ws.cell(row=r, column=2, value=u["sv"]); r += 1

    # ---- Product Overview -------------------------------------------------------
    ws = wb.create_sheet("Product Overview")
    ws.cell(row=1, column=1, value="Product Overview — competitor catalog").font = TITLE
    r = 3
    hdr_row = r
    r = head(ws, r, ["Brand", "ASIN", "You?", "Price", "Rating", "Reviews", "Revenue/mo",
                     "BSR", "Listing age (y)", "Health /10"],
             widths=[20, 14, 6, 9, 8, 10, 12, 10, 13, 10])
    first_data = r
    ordered = sorted(comps, key=lambda c: (not c["is_primary"], -(c["revenue"] or 0)))
    for c in ordered:
        ws.cell(row=r, column=1, value=c["brand"])
        ws.cell(row=r, column=2, value=c["asin"])
        ws.cell(row=r, column=3, value="★" if c["is_primary"] else "")
        ws.cell(row=r, column=4, value=c["price"])
        ws.cell(row=r, column=5, value=c["rating"])
        ws.cell(row=r, column=6, value=c["review_count"])
        ws.cell(row=r, column=7, value=c["revenue"])
        ws.cell(row=r, column=8, value=c["bsr"])
        ws.cell(row=r, column=9, value=c["listing_age"])
        ws.cell(row=r, column=10, value=c["listing_health_score"])
        r += 1
    if ordered:
        ch = BarChart(); ch.type = "col"; ch.title = "Revenue / mo by competitor"; ch.height = 9; ch.width = 20
        ch.add_data(Reference(ws, min_col=7, min_row=hdr_row, max_row=r - 1), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=first_data, max_row=r - 1))
        ws.add_chart(ch, f"L{hdr_row}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_matrix(db: Session, project_id: int, when: date | None = None) -> bytes:
    """Coverage matrix -> xlsx (the client deliverable that replaces the Sheet)."""
    m = matrix(db, project_id, when)
    heads = ["Keyword", "Search Volume", "Relevancy", "Keyword Sales"] + \
        [f"{c['brand'] or ''} {c['asin']}".strip() + (" (YOU)" if c["is_primary"] else "")
         for c in m["columns"]]
    data = [[r["keyword"], r["sv"], r["relevancy"], r["kw_sales"]]
            + [("-" if v is None else v) for v in r["ranks"]] for r in m["rows"]]
    df = pd.DataFrame(data, columns=heads)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name=f"Coverage {m['date']}")
    return buf.getvalue()


def listing_age_years(creation_date: str | None, today: date | None = None) -> float | None:
    """Years since listing creation (sheet's 'Listing Age' row), 1 decimal."""
    if not creation_date:
        return None
    try:
        created = date.fromisoformat(str(creation_date)[:10])
    except ValueError:
        return None
    days = ((today or date.today()) - created).days
    return round(days / 365.0, 1)


# manual competitor-audit fields the UI may edit (sheet's Yes/None dropdown rows)
_EDITABLE_BOOL = {"pdp_images", "pdp_videos", "brand_story", "aplus", "crawlable_text",
                  "alt_text", "comparison_table", "amazon_badge", "active"}

# Listing Health Score: COMPUTED, not manual — each audit check worth 1.25
# points, 8 checks = perfect 10. (Stored sheet scores are ignored.)
AUDIT_FIELDS = ["pdp_images", "pdp_videos", "brand_story", "aplus",
                "crawlable_text", "alt_text", "comparison_table", "amazon_badge"]
HEALTH_POINTS = 1.25


def health_score(c: "md.TrackedCompetitor") -> float:
    """1.25 x Yes-count over the 8 audit rows (None/False score nothing)."""
    return round(HEALTH_POINTS * sum(1 for f in AUDIT_FIELDS
                                     if getattr(c, f) is True), 2)


def competitors(db: Session, project_id: int) -> dict:
    """The sheet's Main-tab competitor matrix, data side: one record per ASIN
    (primary first, then revenue desc) with every imported attribute + computed
    listing age; plus the KPI header block."""
    proj = db.get(md.TrackerProject, project_id)
    if not proj:
        raise ValueError("Unknown tracker project.")
    comps = (db.query(md.TrackedCompetitor)
             .filter(md.TrackedCompetitor.project_id == project_id,
                     md.TrackedCompetitor.active == True)  # noqa: E712
             .order_by(md.TrackedCompetitor.is_primary.desc(),
                       md.TrackedCompetitor.revenue.desc().nullslast()).all())
    rows = []
    for c in comps:
        rows.append({
            "id": c.id, "asin": c.asin, "brand": c.brand, "title": c.title,
            "is_primary": c.is_primary, "active": c.active,
            "price": c.price, "sales": c.sales, "revenue": c.revenue, "bsr": c.bsr,
            "seller_country": c.seller_country, "fba_fees": c.fba_fees,
            "active_sellers": c.active_sellers, "rating": c.rating,
            "review_count": c.review_count, "images": c.images,
            "review_velocity": c.review_velocity, "buy_box": c.buy_box,
            "category": c.category, "size_tier": c.size_tier,
            "fulfillment": c.fulfillment, "dimensions": c.dimensions,
            "weight": c.weight, "creation_date": c.creation_date,
            "listing_age": listing_age_years(c.creation_date),
            "image_url": c.image_url,
            "listing_health_score": health_score(c),
            "pdp_images": c.pdp_images, "pdp_videos": c.pdp_videos,
            "brand_story": c.brand_story, "aplus": c.aplus,
            "crawlable_text": c.crawlable_text, "alt_text": c.alt_text,
            "comparison_table": c.comparison_table, "amazon_badge": c.amazon_badge,
        })
    top = [r for r in rows if not r["is_primary"]][:10]
    total_rev = sum(r["revenue"] or 0 for r in top)
    prim = next((r for r in rows if r["is_primary"]), None)
    return {"project": {"id": proj.id, "name": proj.name, "primary_asin": proj.primary_asin},
            "competitors": rows,
            "kpis": {"total_revenue_top10": round(total_rev, 2),
                     "market_share": (round((prim["revenue"] or 0) / total_rev, 6)
                                      if prim and total_rev else None),
                     "avg_reviews_top10": (round(sum(r["review_count"] or 0 for r in top)
                                                 / len(top), 1) if top else None)}}


def set_competitor_field(db: Session, comp_id: int, field: str, value) -> dict:
    """Manual audit edit (the sheet's Yes/None dropdown rows). The health score
    is computed from these — never set directly."""
    c = db.get(md.TrackedCompetitor, comp_id)
    if not c:
        raise ValueError("Unknown competitor.")
    if field not in _EDITABLE_BOOL:
        raise ValueError(f"Field '{field}' is not editable.")
    setattr(c, field, None if value is None else bool(value))
    db.commit()
    return {"id": c.id, "field": field, "value": getattr(c, field),
            "listing_health_score": health_score(c)}


def list_projects(db: Session) -> list[dict]:
    out = []
    for p in db.query(md.TrackerProject).order_by(md.TrackerProject.id):
        nkw = db.query(md.TrackedKeyword).filter(md.TrackedKeyword.project_id == p.id).count()
        ncomp = db.query(md.TrackedCompetitor).filter(
            md.TrackedCompetitor.project_id == p.id).count()
        out.append({"id": p.id, "name": p.name, "primary_asin": p.primary_asin,
                    "keywords": nkw, "competitors": ncomp,
                    "created_at": p.created_at.isoformat() if p.created_at else None})
    return out


# ---- keyword intake from other engines (Keywords tab / harvest / n-gram) --------
def add_keywords(db: Session, project_id: int, rows: list[dict], source: str = "manual") -> dict:
    """Merge keywords into a project's master list (dedupe by normalized phrase).
    rows: [{keyword, search_volume?, source?}]. Existing keywords keep their data;
    a duplicate only backfills a missing search_volume. Returns counts."""
    proj = db.get(md.TrackerProject, project_id)
    if not proj:
        raise ValueError("Unknown tracker project.")
    existing = {(k.keyword or "").strip().lower(): k
                for k in db.query(md.TrackedKeyword)
                .filter(md.TrackedKeyword.project_id == project_id)}
    added = dup = 0
    for r in rows or []:
        kw = str(r.get("keyword") or "").strip()
        if not kw or len(kw) > 200:
            continue
        key = kw.lower()
        sv = r.get("search_volume")
        hit = existing.get(key)
        if hit is not None:
            dup += 1
            if sv and not hit.search_volume:
                hit.search_volume = int(sv)
            continue
        k = md.TrackedKeyword(project_id=project_id, keyword=kw,
                              search_volume=int(sv) if sv else None,
                              source=str(r.get("source") or source)[:40])
        db.add(k)
        existing[key] = k
        added += 1
    db.commit()
    total = db.query(md.TrackedKeyword).filter(
        md.TrackedKeyword.project_id == project_id).count()
    return {"added": added, "duplicates": dup, "total": total,
            "project": {"id": proj.id, "name": proj.name, "primary_asin": proj.primary_asin}}


# ---- AI relevancy prompt --------------------------------------------------------
_PROMPT_ELEMENTS = [("title", "Title"), ("bullet_points", "Bullet Points"),
                    ("description", "Description"), ("search_terms", "Backend Search Terms"),
                    ("alt_text", "Alt Text"), ("aplus", "A+ Content")]


def relevancy_prompt(db: Session, project_id: int, max_keywords: int = 300) -> dict:
    """Copy-paste AI prompt: judge every tracked keyword's relevancy against the
    CURRENT listing copy and the PROPOSED listing copy, and say where each
    keyword should live. Pure text assembly — no LLM call here."""
    proj = db.get(md.TrackerProject, project_id)
    if not proj:
        raise ValueError("Unknown tracker project.")
    kws = (db.query(md.TrackedKeyword)
           .filter(md.TrackedKeyword.project_id == project_id)
           .order_by(md.TrackedKeyword.search_volume.desc().nullslast()).all())
    truncated = max(0, len(kws) - max_keywords)
    kws = kws[:max_keywords]

    def copy_block(variant: str) -> str:
        rows = db.query(md.ListingCopy).filter(
            md.ListingCopy.project_id == project_id,
            md.ListingCopy.variant == variant,
            md.ListingCopy.asin.is_(None)).all()
        by_el = {r.element: (r.text or "").strip() for r in rows}
        lines = []
        for el, label in _PROMPT_ELEMENTS:
            txt = by_el.get(el) or "(empty)"
            lines.append(f"{label}:\n{txt}")
        return "\n\n".join(lines)

    kw_lines = "\n".join(
        f"- {k.keyword}"
        + (f" (search volume {k.search_volume:,})" if k.search_volume else "")
        + (f" [source: {k.source}]" if k.source else "")
        for k in kws)

    prompt = f"""You are an Amazon listing SEO expert. Judge keyword relevancy for this product.

PRODUCT
ASIN: {proj.primary_asin or "(not set)"}
Project: {proj.name}

CURRENT LISTING DATA
{copy_block("current")}

PROPOSED LISTING DATA
{copy_block("proposed")}

KEYWORDS ({len(kws)}{f", top {max_keywords} by search volume — {truncated} more omitted" if truncated else ""})
{kw_lines}

TASK
For EVERY keyword above, output one markdown table with these columns:
| Keyword | Relevancy (1-5) | In Current? | In Proposed? | Best Placement | Reason |

Rules:
1. Relevancy 5 = exactly what this product is; 3 = related/complementary use case; 1 = wrong product or different intent. Judge against the CURRENT and PROPOSED listing data above, not generic category knowledge.
2. "In Current?" / "In Proposed?" = yes / partial / no, based on whether the keyword's meaningful words appear in that listing copy (title, bullets, description, backend).
3. Best Placement = Title, Bullet Points, Description, Backend Search Terms, or Drop (for relevancy 1-2 keywords that would dilute indexing).
4. Reason = one short sentence.
5. After the table, add a summary: (a) the 10 highest-value keywords missing from the PROPOSED copy, (b) any keywords the PROPOSED copy covers worse than the CURRENT copy, (c) irrelevant keywords to remove from the tracked list.
Keep every keyword's text exactly as written. Do not invent keywords."""

    return {"prompt": prompt, "keywords": len(kws), "truncated": truncated,
            "project": {"id": proj.id, "name": proj.name, "primary_asin": proj.primary_asin}}


def primary_seo(db: Session, project_id: int) -> dict | None:
    """The primary ASIN's SEO headline numbers (indexed % / page 1 / avg rank /
    tracked / ranked) at the latest snapshot — the Keywords tab compares this
    BEFORE vs AFTER a keyword push, so a denominator drop is visible, not silent."""
    try:
        sc = scorecard(db, project_id)
    except ValueError:
        return None
    card = next((c for c in sc["cards"] if c.get("is_primary")), None) \
        or (sc["cards"][0] if sc["cards"] else None)
    if not card:
        return None
    return {"asin": card["asin"], "indexed": card["index_rate"],
            "page1": card["page1_count"], "top10": card["top10_count"],
            "avg_rank": card["avg_rank"], "tracked": card["kw_tracked"],
            "ranked": card["kw_ranked"]}

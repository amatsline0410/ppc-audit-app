"""SKU-level transactions — Amazon Payments Date Range (transaction) report.

Feeds the Product Benchmark tab: upload the transaction report from Seller
Central > Payments > Reports Repository (Date Range report, transaction view),
then filter by date to see per-SKU order/refund economics (product sales,
selling fees, FBA fees, promos, net proceeds).

Scope: STORE-level (like the catalog) — persisted as `<store>/_transactions.json`,
shared by every audit in the store, deleted with the store dir. Uploads MERGE:
overlapping date ranges dedupe row-by-row, so monthly report re-uploads accumulate
into one continuous ledger.

File shape: several quoted preamble lines ("Includes Amazon Marketplace…",
definitions), then the real header row starting with "date/time". Amounts are
plain numbers (may carry $ or thousands commas); dates look like
"Jun 1, 2026 5:43:55 AM PDT". Rows without a SKU (Transfer, Service Fee,
Shipping Services, FBA Inventory Fee…) are account-level — kept, but reported
separately from the SKU tables.
"""
from __future__ import annotations

import csv
import json
import os
import re
from datetime import datetime

from .. import database as dbmod

_MAX_FILES = 6   # rolling upload-history length (matches other meta lines)


# ---- store-scoped file --------------------------------------------------------
def _store_path(store_id: str) -> str:
    return os.path.join(dbmod._store_dir(store_id), "_transactions.json")


def read_txn(store_id: str | None) -> dict:
    if not store_id:
        return {}
    p = _store_path(store_id)
    if os.path.exists(p):
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return {}


def write_txn(store_id: str, data: dict) -> None:
    with open(_store_path(store_id), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def delete_all(store_id: str) -> None:
    p = _store_path(store_id)
    if os.path.exists(p):
        os.unlink(p)


# ---- parsing -------------------------------------------------------------------
_num_re = re.compile(r"-?[\d,]*\.?\d+")

# loose header -> row key mapping (Amazon drifts wording/case across marketplaces)
_COLS = {
    "date/time": "datetime",
    "settlement id": "settlement_id",
    "type": "type",
    "order id": "order_id",
    "sku": "sku",
    "description": "description",
    "quantity": "quantity",
    "marketplace": "marketplace",
    "fulfillment": "fulfillment",
    "order city": "city",
    "order state": "state",
    "product sales": "product_sales",
    "shipping credits": "shipping_credits",
    "gift wrap credits": "gift_wrap",
    "promotional rebates": "promo",
    "marketplace withheld tax": "withheld_tax",
    "selling fees": "selling_fees",
    "fba fees": "fba_fees",
    "other transaction fees": "other_fees",
    "other": "other",
    "total": "total",
    "transaction status": "status",
}
_MONEY = ("product_sales", "shipping_credits", "gift_wrap", "promo", "withheld_tax",
          "selling_fees", "fba_fees", "other_fees", "other", "total")


def _s(v) -> str:
    return str(v).strip() if v is not None else ""


def _money(v) -> float:
    m = _num_re.search(_s(v).replace(",", ""))
    return float(m.group()) if m else 0.0


def _date_iso(v) -> str | None:
    """'Jun 1, 2026 5:43:55 AM PDT' -> '2026-06-01' (time/zone dropped)."""
    s = _s(v)
    m = re.match(r"([A-Za-z]{3})\.?\s+(\d{1,2}),?\s+(\d{4})", s)
    if not m:
        return None
    try:
        return datetime.strptime(f"{m.group(1)} {m.group(2)} {m.group(3)}", "%b %d %Y").date().isoformat()
    except ValueError:
        return None


def _rows_from_file(path: str):
    """Yield raw rows (lists of cells) from a .csv or .xlsx/.xlsm export."""
    if path.lower().endswith(".csv"):
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            yield from csv.reader(f)
        return
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            probe = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                probe.append(row)
                if i >= 12:
                    break
            if any(_s(r[0] if r else "").lower() == "date/time" for r in probe):
                yield from ws.iter_rows(values_only=True)
                return
        raise ValueError("no transaction sheet found")
    finally:
        wb.close()


def parse_txn(path: str) -> list[dict]:
    """Parse a Date Range transaction report. Skips the quoted preamble, keys
    columns loosely off the real header row, keeps every transaction row."""
    header = None   # {col index: row key}
    out = []
    for raw in _rows_from_file(path):
        cells = [_s(c) for c in (raw or [])]
        if header is None:
            if cells and cells[0].lower() == "date/time":
                header = {}
                for i, c in enumerate(cells):
                    key = _COLS.get(c.lower())
                    if key:
                        header[i] = key
            continue
        if not any(cells):
            continue
        r = {k: "" for k in _COLS.values()}
        for i, key in header.items():
            if i < len(cells):
                r[key] = cells[i]
        date = _date_iso(r["datetime"])
        if not date:
            continue                      # not a data row
        for k in _MONEY:
            r[k] = round(_money(r[k]), 2)
        qty = _num_re.search(_s(r["quantity"]))
        r["quantity"] = int(float(qty.group().replace(",", ""))) if qty else 0
        r["date"] = date
        out.append(r)
    if header is None:
        raise ValueError(
            "This doesn't look like an Amazon transaction report — no 'date/time' "
            "header row found. Download it from Seller Central > Payments > "
            "Reports Repository (Date Range report, transaction view).")
    if not out:
        raise ValueError("No transaction rows found in that file.")
    return out


# ---- merge (uploads accumulate; matching rows UPDATE in place) -------------------
def _key(r: dict) -> str:
    """Transaction identity — deliberately excludes the amounts, so a re-export of
    the same transaction with changed fields (status Deferred -> Released, fee
    correction) REPLACES the stored row instead of double-counting it."""
    return "|".join([r.get("datetime", ""), r.get("settlement_id", ""), r.get("type", ""),
                     r.get("order_id", ""), r.get("sku", "")])


def merge(data: dict, rows: list[dict], filename: str) -> tuple[dict, int, int, int]:
    """Upsert parsed rows into the store ledger: new transactions append, known
    ones with changed data update in place, identical ones skip.
    Returns (data, added, updated, dupes)."""
    data = data or {}
    existing = data.get("rows", [])
    by_key = {_key(r): i for i, r in enumerate(existing)}
    added = updated = dupes = 0
    for r in rows:
        k = _key(r)
        i = by_key.get(k)
        if i is None:
            by_key[k] = len(existing)
            existing.append(r)
            added += 1
        elif existing[i] == r:
            dupes += 1
        else:
            existing[i] = r
            updated += 1
    existing.sort(key=lambda r: (r.get("date") or "", r.get("datetime") or ""))
    files = [f for f in data.get("files", []) if f.get("name") != filename]
    files.append({"name": filename, "rows": len(rows),
                  "uploaded": datetime.now().isoformat(timespec="seconds")})
    data.update(rows=existing, files=files[-_MAX_FILES:],
                updated=datetime.now().isoformat(timespec="seconds"))
    return data, added, updated, dupes


# ---- date-filtered summary -------------------------------------------------------
def summary(data: dict, start: str | None = None, end: str | None = None,
            sku: str | None = None) -> dict:
    """Filter the ledger by [start, end] (ISO dates, inclusive) and roll up per SKU.
    `sku` additionally narrows the transaction drill-down list (not the rollup)."""
    rows = data.get("rows", []) if data else []
    dates = [r["date"] for r in rows if r.get("date")]
    lo, hi = (min(dates), max(dates)) if dates else (None, None)
    sel = [r for r in rows
           if (not start or r["date"] >= start) and (not end or r["date"] <= end)]

    sku_rows = [r for r in sel if r.get("sku")]
    other_rows = [r for r in sel if not r.get("sku")]

    per = {}
    for r in sku_rows:
        p = per.setdefault(r["sku"], {
            "sku": r["sku"], "description": r.get("description", ""),
            "orders": 0, "refunds": 0, "units": 0, "product_sales": 0.0,
            "promo": 0.0, "selling_fees": 0.0, "fba_fees": 0.0,
            "other_fees": 0.0, "net": 0.0,
        })
        t = (r.get("type") or "").lower()
        if t == "order":
            p["orders"] += 1
            p["units"] += r["quantity"]
        elif t == "refund":
            p["refunds"] += 1
            p["units"] -= r["quantity"]
        p["product_sales"] += r["product_sales"]
        p["promo"] += r["promo"]
        p["selling_fees"] += r["selling_fees"]
        p["fba_fees"] += r["fba_fees"]
        p["other_fees"] += r["other_fees"] + r["other"]
        p["net"] += r["total"]
    skus = sorted(per.values(), key=lambda p: p["net"], reverse=True)
    for p in skus:
        for k in ("product_sales", "promo", "selling_fees", "fba_fees", "other_fees", "net"):
            p[k] = round(p[k], 2)

    def tot(key, src):
        return round(sum(r[key] for r in src), 2)
    totals = {
        "transactions": len(sel),
        "orders": sum(1 for r in sku_rows if (r.get("type") or "").lower() == "order"),
        "refunds": sum(1 for r in sku_rows if (r.get("type") or "").lower() == "refund"),
        "units": sum(p["units"] for p in skus),
        "product_sales": tot("product_sales", sku_rows),
        "promo": tot("promo", sku_rows),
        "selling_fees": tot("selling_fees", sku_rows),
        "fba_fees": tot("fba_fees", sku_rows),
        "other_fees": round(sum(r["other_fees"] + r["other"] for r in sku_rows), 2),
        "net": tot("total", sku_rows),
        "account_other": tot("total", [r for r in other_rows
                                       if (r.get("type") or "").lower() != "transfer"]),
        "transfers": tot("total", [r for r in other_rows
                                   if (r.get("type") or "").lower() == "transfer"]),
    }

    drill = [r for r in sku_rows if not sku or r["sku"] == sku]
    daily = {}
    for r in sku_rows:
        d = daily.setdefault(r["date"], {"date": r["date"], "orders": 0, "refunds": 0,
                                         "units": 0, "product_sales": 0.0, "net": 0.0})
        tt = (r.get("type") or "").lower()
        if tt == "order":
            d["orders"] += 1; d["units"] += r["quantity"]
        elif tt == "refund":
            d["refunds"] += 1; d["units"] -= r["quantity"]
        d["product_sales"] += r["product_sales"]
        d["net"] += r["total"]
    days = sorted(daily.values(), key=lambda d: d["date"])
    for d in days:
        d["product_sales"] = round(d["product_sales"], 2); d["net"] = round(d["net"], 2)
    return {
        "range": {"min": lo, "max": hi, "start": start, "end": end},
        "totals": totals,
        "skus": skus,
        "transactions": drill,
        "account_rows": other_rows,
        "days": days,
        "files": (data or {}).get("files", []),
        "updated": (data or {}).get("updated"),
        "total_rows": len(rows),
    }


# ---- Export Report (xlsx with native Excel charts) --------------------------------
def report_xlsx(data: dict, start: str | None = None, end: str | None = None) -> bytes:
    """Client-ready SKU Transactions workbook for the selected date window:
    Overview (KPIs, net-by-SKU bar, fee-composition pie) + Daily Trend (sales/net
    line) + By SKU rollup + raw Transactions."""
    import io
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill

    s = summary(data, start, end)
    if not s["total_rows"]:
        raise ValueError("No transactions yet — upload a Payments Date Range report first.")

    TITLE = Font(bold=True, size=14)
    H = Font(bold=True, color="1F2329")
    HFILL = PatternFill("solid", fgColor="FCD535")
    SUB = Font(bold=True, size=11)

    def head(ws, row, labels, widths=None):
        for i, lab in enumerate(labels, start=1):
            c = ws.cell(row=row, column=i, value=lab)
            c.font = H; c.fill = HFILL
            if widths and i <= len(widths) and widths[i - 1]:
                ws.column_dimensions[c.column_letter].width = widths[i - 1]
        return row + 1

    t = s["totals"]
    window = f"{start or s['range']['min']} → {end or s['range']['max']}"
    wb = Workbook()

    # ---- Overview -----------------------------------------------------------------
    ws = wb.active
    ws.title = "Overview"
    ws.cell(row=1, column=1, value="SKU Transactions report").font = TITLE
    ws.cell(row=2, column=1,
            value=f"window {window} · {t['transactions']} transactions · "
                  f"{len(s['skus'])} SKUs (ledger {s['range']['min']} → {s['range']['max']})")
    r = 4
    for lab, val in [("Orders", t["orders"]), ("Refunds", t["refunds"]), ("Units", t["units"]),
                     ("Product sales", t["product_sales"]), ("Promo rebates", t["promo"]),
                     ("Selling fees", t["selling_fees"]), ("FBA fees", t["fba_fees"]),
                     ("Other fees", t["other_fees"]), ("Net proceeds", t["net"]),
                     ("Account-level other (no SKU)", t["account_other"]),
                     ("Bank transfers", t["transfers"])]:
        ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    # net proceeds by SKU (top 15) — bar
    r += 1
    ws.cell(row=r, column=1, value="Net proceeds by SKU (top 15)").font = SUB
    r += 1
    r = head(ws, r, ["SKU", "Net proceeds"], widths=[34, 14])
    first = r
    for p in s["skus"][:15]:
        ws.cell(row=r, column=1, value=p["sku"]); ws.cell(row=r, column=2, value=p["net"]); r += 1
    if r > first:
        bar = BarChart(); bar.type = "col"; bar.title = "Net proceeds by SKU"
        bar.height = 8; bar.width = 20; bar.legend = None
        bar.add_data(Reference(ws, min_col=2, min_row=first - 1, max_row=r - 1), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
        ws.add_chart(bar, "D4")

    # where the money goes — pie (absolute values of the deduction buckets)
    r += 1
    ws.cell(row=r, column=1, value="Deductions").font = SUB
    r += 1
    r = head(ws, r, ["Bucket", "Amount"], widths=None)
    first = r
    for lab, val in [("Selling fees", abs(t["selling_fees"])), ("FBA fees", abs(t["fba_fees"])),
                     ("Promo rebates", abs(t["promo"])), ("Other fees", abs(t["other_fees"]))]:
        ws.cell(row=r, column=1, value=lab); ws.cell(row=r, column=2, value=round(val, 2)); r += 1
    pie = PieChart(); pie.title = "Deductions"; pie.height = 8; pie.width = 12
    pie.add_data(Reference(ws, min_col=2, min_row=first, max_row=r - 1))
    pie.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
    ws.add_chart(pie, "D21")

    # ---- Daily Trend ----------------------------------------------------------------
    ws = wb.create_sheet("Daily Trend")
    ws.cell(row=1, column=1, value=f"Daily trend · {window}").font = TITLE
    r = head(ws, 3, ["Date", "Orders", "Refunds", "Units", "Product sales", "Net proceeds"],
             widths=[12, 8, 8, 8, 13, 13])
    first = r
    for d in s["days"]:
        ws.cell(row=r, column=1, value=d["date"]); ws.cell(row=r, column=2, value=d["orders"])
        ws.cell(row=r, column=3, value=d["refunds"]); ws.cell(row=r, column=4, value=d["units"])
        ws.cell(row=r, column=5, value=d["product_sales"]); ws.cell(row=r, column=6, value=d["net"])
        r += 1
    if r > first:
        ln = LineChart(); ln.title = "Product sales vs net proceeds"; ln.height = 9; ln.width = 26
        ln.add_data(Reference(ws, min_col=5, max_col=6, min_row=first - 1, max_row=r - 1),
                    titles_from_data=True)
        ln.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
        ws.add_chart(ln, "H3")

    # ---- By SKU ---------------------------------------------------------------------
    ws = wb.create_sheet("By SKU")
    ws.cell(row=1, column=1, value=f"Per-SKU rollup · {window}").font = TITLE
    r = head(ws, 3, ["SKU", "Product", "Orders", "Refunds", "Units", "Product sales",
                     "Promos", "Selling fees", "FBA fees", "Other", "Net proceeds"],
             widths=[26, 44, 8, 8, 8, 13, 10, 12, 10, 10, 13])
    for p in s["skus"]:
        ws.append([p["sku"], p["description"], p["orders"], p["refunds"], p["units"],
                   p["product_sales"], p["promo"], p["selling_fees"], p["fba_fees"],
                   p["other_fees"], p["net"]])

    # ---- Transactions -----------------------------------------------------------------
    ws = wb.create_sheet("Transactions")
    ws.cell(row=1, column=1, value=f"SKU-level transactions · {window}").font = TITLE
    head(ws, 3, ["Date", "Type", "Order ID", "SKU", "Qty", "City", "State",
                 "Product sales", "Promos", "Selling fees", "FBA fees", "Total", "Status"],
         widths=[12, 10, 22, 26, 6, 16, 8, 13, 10, 12, 10, 11, 10])
    for x in s["transactions"]:
        ws.append([x["date"], x["type"], x["order_id"], x["sku"], x["quantity"],
                   x["city"], x["state"], x["product_sales"], x["promo"],
                   x["selling_fees"], x["fba_fees"], x["total"], x["status"]])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

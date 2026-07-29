"""Product Ad performance — its OWN feature, on its OWN table (`ProductAdFact`),
**separate from the PPC Audit star schema** (DimAd/FactPerformance).

Upload a Sponsored Products bulk → the Product Ad rows (Entity='Product Ad') of its
Sponsored Products Campaigns sheet are parsed straight into `ProductAdFact` (ASIN +
SKU + real IDs + that row's metrics), replacing the prior snapshot. `summary()`
consolidates one row per (ASIN, SKU) — summing the underlying ads, recomputing rates
from raw counts — plus one account total. `detail()` drills one or more ASINs into
their ads + per-campaign rollups. Fully self-contained: it never reads the audit
pipeline, so Product Ads runs off its own dedicated upload.
"""
from __future__ import annotations
import pandas as pd
from sqlalchemy import text
from sqlalchemy.orm import Session
from .. import models as md
from .. import metrics as M
from . import bulkfmt, workbook


# add the campaign_type column to pre-existing product_ad_fact tables (SQLite ALTER).
def _ensure_schema(db: Session) -> None:
    cols = {r[1] for r in db.execute(text("PRAGMA table_info(product_ad_fact)")).fetchall()}
    if cols and "campaign_type" not in cols:
        db.execute(text("ALTER TABLE product_ad_fact ADD COLUMN campaign_type VARCHAR"))
        db.commit()


# ---- parse the bulk's Product Ad rows ---------------------------------------
def _num(v) -> float:
    if v is None:
        return 0.0
    s = str(v).strip().replace("$", "").replace(",", "").replace("%", "")
    if not s or s.lower() == "nan":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def _find_sheet(path: str) -> str | None:
    """The Sponsored Products Campaigns sheet by name, else any sheet with an
    'Entity' column carrying 'Product Ad' rows."""
    hit = workbook.find_sheet(path, "sponsored product", "campaign")
    if hit:
        return hit
    for sn in workbook.sheet_names(path):
        if any(str(c).strip().lower() == "entity" for c in workbook.columns(path, sn)):
            return sn
    return None


def _resolve(df: pd.DataFrame) -> dict[str, str]:
    norm = {str(c).strip().lower(): c for c in df.columns}

    def find(*preds):
        for pred in preds:
            for low, orig in norm.items():
                if pred(low):
                    return orig
        return None

    return {
        "entity": find(lambda s: s == "entity"),
        "campaign_id": find(lambda s: s == "campaign id"),
        "ad_group_id": find(lambda s: s == "ad group id"),
        "ad_id": find(lambda s: s == "ad id"),
        # Amazon fills the bare "Campaign Name" only on Campaign rows and "Ad Group
        # Name" only on Ad Group rows — both are blank on the Product Ad / Keyword
        # rows we actually read. The "(Informational only)" twins are filled on every
        # row, so prefer those and keep the bare ones as the fallback.
        "campaign_name": find(lambda s: s == "campaign name (informational only)",
                              lambda s: s.startswith("campaign name")),
        "ad_group_name": find(lambda s: s == "ad group name (informational only)",
                              lambda s: s.startswith("ad group name")),
        "campaign_name_own": find(lambda s: s == "campaign name"),
        "ad_group_name_own": find(lambda s: s == "ad group name"),
        "state": find(lambda s: s == "state"),
        "sku": find(lambda s: s == "sku"),
        "asin": find(lambda s: s.startswith("asin")),
        "impressions": find(lambda s: s == "impressions"),
        "clicks": find(lambda s: s == "clicks"),
        "spend": find(lambda s: s == "spend", lambda s: "spend" in s),
        "sales": find(lambda s: "7 day" in s and "sales" in s, lambda s: s == "sales", lambda s: "sales" in s),
        "orders": find(lambda s: "7 day" in s and "orders" in s, lambda s: s == "orders", lambda s: "orders" in s),
        "units": find(lambda s: "7 day" in s and "units" in s, lambda s: s == "units", lambda s: "units" in s),
    }


def _read_df(path: str) -> pd.DataFrame:
    if path.lower().endswith(".csv"):
        return workbook.read_str(path)
    sheet = _find_sheet(path)
    if sheet is None:
        raise ValueError("No 'Sponsored Products Campaigns' sheet found — Product Ads needs the "
                         "Amazon Sponsored Products bulk export.")
    return workbook.read_str(path, sheet)


def classify_campaigns(df: pd.DataFrame) -> dict[str, str]:
    """Map each Campaign ID → its targeting kind, read from the SAME bulk:
    'auto' (Targeting Type = auto), else 'keyword' (has positive Keyword entities),
    'product' (has positive Product Targeting entities), else 'manual' (type unknown).
    Auto campaigns win first, so their auto-clause Product Targeting rows don't
    mislabel them. Returns {} when the sheet carries no campaign/entity rows."""
    norm = {str(c).strip().lower(): c for c in df.columns}
    c_entity = norm.get("entity")
    c_cid = norm.get("campaign id")
    c_tt = norm.get("targeting type")
    if not c_entity or not c_cid:
        return {}
    tt_map: dict[str, str] = {}          # campaign_id -> auto/manual (from Campaign rows)
    has_kw: set[str] = set()
    has_pt: set[str] = set()
    for _, r in df.iterrows():
        ent = _str(r.get(c_entity)).lower()
        cid = bulkfmt.idstr(_str(r.get(c_cid)))
        if not cid:
            continue
        if ent == "campaign":
            tt = _str(r.get(c_tt)).lower() if c_tt else ""
            if tt:
                tt_map[cid] = tt
        elif ent == "keyword":          # positive keyword only (negatives are 'negative keyword')
            has_kw.add(cid)
        elif ent == "product targeting":
            has_pt.add(cid)
    out: dict[str, str] = {}
    for cid in set(tt_map) | has_kw | has_pt:
        tt = tt_map.get(cid, "")
        if tt.startswith("auto"):
            out[cid] = "auto"
        elif cid in has_kw:
            out[cid] = "keyword"
        elif cid in has_pt:
            out[cid] = "product"
        elif tt.startswith("manual"):
            out[cid] = "manual"
    return out


def name_maps(df: pd.DataFrame) -> tuple[dict[str, str], dict[str, str]]:
    """({campaign_id: name}, {ad_group_id: name}) read off the bulk's own Campaign
    and Ad Group entity rows — the last-resort fallback for exports that carry
    neither a populated "Name" nor an "(Informational only)" column on child rows."""
    cols = _resolve(df)
    c_ent, c_cid, c_agid = cols.get("entity"), cols.get("campaign_id"), cols.get("ad_group_id")
    c_cname, c_agname = cols.get("campaign_name_own"), cols.get("ad_group_name_own")
    camps: dict[str, str] = {}
    ags: dict[str, str] = {}
    if not c_ent:
        return camps, ags
    for _, r in df.iterrows():
        ent = _str(r.get(c_ent)).lower()
        if ent == "campaign" and c_cid and c_cname:
            cid = bulkfmt.idstr(_str(r.get(c_cid)))
            nm = _str(r.get(c_cname))
            if cid and nm:
                camps[cid] = nm
        elif ent == "ad group" and c_agid and c_agname:
            ag = bulkfmt.idstr(_str(r.get(c_agid)))
            nm = _str(r.get(c_agname))
            if ag and nm:
                ags[ag] = nm
    return camps, ags


def _pick_name(r, cols: dict, key: str, fallback: dict[str, str], ent_id: str | None) -> str | None:
    """A row's campaign / ad-group name: its own column, else the name the bulk's
    Campaign / Ad Group entity row carries for that ID. Never returns "" — callers
    fall back to the raw ID only when there is genuinely no name anywhere."""
    col = cols.get(key)
    own = _str(r.get(col)) if col else ""
    if not own:
        col2 = cols.get(f"{key}_own")
        own = _str(r.get(col2)) if col2 else ""
    return own or (fallback.get(ent_id) if ent_id else None) or None


def _rows_from_df(df: pd.DataFrame, ctypes: dict[str, str]) -> list[dict]:
    cols = _resolve(df)
    if not cols.get("entity"):
        raise ValueError("That sheet has no 'Entity' column — is it the Sponsored Products bulk?")
    camp_names, ag_names = name_maps(df)
    rows = []
    for _, r in df.iterrows():
        if _str(r.get(cols["entity"])).lower() != "product ad":
            continue
        asin = _str(r.get(cols["asin"])) if cols.get("asin") else ""
        sku = _str(r.get(cols["sku"])) if cols.get("sku") else ""
        if not asin and not sku:
            continue
        cid = bulkfmt.idstr(_str(r.get(cols["campaign_id"]))) if cols.get("campaign_id") else None
        agid = bulkfmt.idstr(_str(r.get(cols["ad_group_id"]))) if cols.get("ad_group_id") else None
        rows.append(dict(
            ad_id=bulkfmt.idstr(_str(r.get(cols["ad_id"]))) if cols.get("ad_id") else None,
            asin=asin or None, sku=sku or None,
            campaign_id=cid,
            campaign_name=_pick_name(r, cols, "campaign_name", camp_names, cid),
            campaign_type=ctypes.get(cid) if cid else None,
            ad_group_id=agid,
            ad_group_name=_pick_name(r, cols, "ad_group_name", ag_names, agid),
            state=(_str(r.get(cols["state"])) if cols.get("state") else "") or None,
            impressions=int(_num(r.get(cols["impressions"])) if cols.get("impressions") else 0),
            clicks=int(_num(r.get(cols["clicks"])) if cols.get("clicks") else 0),
            spend=round(_num(r.get(cols["spend"])) if cols.get("spend") else 0.0, 2),
            sales=round(_num(r.get(cols["sales"])) if cols.get("sales") else 0.0, 2),
            orders=int(_num(r.get(cols["orders"])) if cols.get("orders") else 0),
            units=int(_num(r.get(cols["units"])) if cols.get("units") else 0),
        ))
    if not rows:
        raise ValueError("No Product Ad rows found in that bulk (Entity='Product Ad'). Make sure you "
                         "exported the Sponsored Products campaigns with their Product Ads.")
    return rows


def parse_product_ads(path: str) -> list[dict]:
    """Read the bulk's Product Ad rows into normalized dicts (IDs kept as exact
    strings), each stamped with its campaign's targeting kind. Raises ValueError
    with a friendly message when none are found."""
    df = _read_df(path)
    return _rows_from_df(df, classify_campaigns(df))


def ingest(db: Session, path: str) -> dict:
    """Parse + store the Product Ad rows as the snapshot (replaces any prior one)."""
    _ensure_schema(db)
    rows = parse_product_ads(path)
    db.query(md.ProductAdFact).delete()
    db.bulk_insert_mappings(md.ProductAdFact, rows)
    db.commit()
    return {"product_ads": len(rows), "asins": len({r["asin"] for r in rows if r["asin"]})}


def has_data(db: Session) -> bool:
    _ensure_schema(db)
    return db.query(md.ProductAdFact.id).first() is not None


def delete_all(db: Session) -> int:
    """Wipe the Product Ads snapshot. Own table only — Product Ads never feeds the
    audit star schema, so nothing else clears with it."""
    _ensure_schema(db)
    n = db.query(md.ProductAdFact).delete()
    db.commit()
    return n


# ---- consolidated summary ---------------------------------------------------
def _status(g: dict) -> str:
    """no_data — ads but no traffic/spend; no_orders — traffic but 0 orders; ok — converting."""
    if g["impressions"] == 0 and g["clicks"] == 0 and g["spend"] == 0:
        return "no_data"
    if g["orders"] == 0:
        return "no_orders"
    return "ok"


def _type_counts(cmap: dict[str, str]) -> dict:
    """campaign_id->type map → counts per targeting kind + total distinct campaigns."""
    c = {"auto": 0, "keyword": 0, "product": 0, "manual": 0}
    for typ in cmap.values():
        c[typ if typ in c else "manual"] += 1
    c["total"] = len(cmap)
    return c


def summary(db: Session, model=None) -> dict:
    """One row per (ASIN, SKU): underlying Product Ads summed, rates recomputed from
    raw counts; plus one account total. Also traces each SKU's campaigns by targeting
    kind (Automatic / Keyword-target / Product-target).

    `model` defaults to `ProductAdFact`; Ads Studio passes its own `AdsStudioAdFact`
    so the two panels render an identical table off independent uploads."""
    model = model or md.ProductAdFact
    if model is md.ProductAdFact:
        _ensure_schema(db)
    groups: dict[tuple, dict] = {}
    acct_types: dict[str, str] = {}     # every distinct campaign_id -> type, account-wide
    t_im = t_cl = t_od = t_un = 0
    t_sp = t_sa = 0.0
    for r in db.query(model).all():
        key = (r.asin, r.sku)
        g = groups.get(key)
        if g is None:
            g = groups[key] = {"asin": r.asin, "sku": r.sku, "ads": 0, "states": set(),
                               "camp_types": {},   # campaign_id -> type (distinct per SKU)
                               "impressions": 0, "clicks": 0, "orders": 0, "units": 0,
                               "spend": 0.0, "sales": 0.0}
        g["ads"] += 1
        if r.state:
            g["states"].add(r.state)
        if r.campaign_id:
            g["camp_types"][r.campaign_id] = r.campaign_type or "manual"
            acct_types[r.campaign_id] = r.campaign_type or "manual"
        g["impressions"] += r.impressions or 0
        g["clicks"] += r.clicks or 0
        g["orders"] += r.orders or 0
        g["units"] += r.units or 0
        g["spend"] = round(g["spend"] + (r.spend or 0), 2)
        g["sales"] = round(g["sales"] + (r.sales or 0), 2)
        t_im += r.impressions or 0; t_cl += r.clicks or 0; t_od += r.orders or 0; t_un += r.units or 0
        t_sp += r.spend or 0; t_sa += r.sales or 0

    # per-product break-even ACoS: benchmark upload wins, store catalog (price +
    # per-SKU COGS, default 40%) fills the rest — same map the PPC Audit uses.
    # Match by ASIN first; when the ASIN doesn't line up across reports, match
    # the catalog LISTING by normalized SKU.
    from .. import database as dbmod
    from . import benchmark as bn
    from . import catalog as cat
    be_map = bn.break_even_map(db)
    sku_be = cat.be_by_sku(db.info.get("store"),
                           dbmod.get_project_econ(db.info.get("store"), db.info.get("project")))

    rows = []
    for g in groups.values():
        d = M.all_metrics(g["impressions"], g["clicks"], g["spend"], g["sales"], g["orders"], g["units"])
        # Unit price = PPC sales / units (an order can carry several units). No units
        # but spend -> surface the bleed as NEGATIVE; else None.
        if g["units"]:
            unit_price = round(g["sales"] / g["units"], 2)
        elif g["orders"]:
            unit_price = round(g["sales"] / g["orders"], 2)
        elif g["spend"]:
            unit_price = -round(g["spend"], 2)
        else:
            unit_price = None
        aov = round(g["sales"] / g["orders"], 2) if g["orders"] else None
        state = next(iter(g["states"])) if len(g["states"]) == 1 else ("mixed" if g["states"] else None)
        tc = _type_counts(g["camp_types"])
        be = be_map.get(g["asin"])
        if be is None and g["sku"]:
            be = sku_be.get(cat.norm_sku(g["sku"]))
        rows.append({"asin": g["asin"], "sku": g["sku"], "state": state, "ads": g["ads"],
                     "status": _status(g), "unit_price": unit_price, "aov": aov,
                     "break_even_acos": be,
                     "campaigns": tc["total"], "auto_campaigns": tc["auto"],
                     "keyword_campaigns": tc["keyword"], "product_campaigns": tc["product"],
                     "manual_campaigns": tc["manual"], **d})

    rows.sort(key=lambda r: (r["status"] == "ok", -r["spend"]))
    total = M.all_metrics(t_im, t_cl, t_sp, t_sa, t_od, t_un)
    by_status: dict[str, int] = {}
    for r in rows:
        by_status[r["status"]] = by_status.get(r["status"], 0) + 1
    # Average Product ACoS = Σ ad spend ÷ Σ ad sales over ALL products' campaigns
    # (the aggregate ratio, per the reporting spec — spend on zero-sale products
    # is included in the numerator).
    avg_acos = round(t_sp / t_sa, 4) if t_sa else None
    return {"snapshot": None, "count": len(rows), "total": total,
            "avg_acos": avg_acos, "avg_acos_products": len(rows),
            "by_status": by_status, "campaign_types": _type_counts(acct_types), "rows": rows}


def by_asin(db: Session) -> dict:
    """Per-ASIN join view for OTHER features (Product Benchmark catalog): every
    advertised ASIN → its Product-Ad rollup (ads, distinct campaigns by targeting
    kind, raw metrics + rates). Returns {"asins": {asin: {...}}, "campaign_types":
    account-wide distinct-campaign counts}. All from `ProductAdFact`."""
    _ensure_schema(db)
    per: dict[str, dict] = {}
    acct_types: dict[str, str] = {}
    for r in db.query(md.ProductAdFact).all():
        if not r.asin:
            continue
        g = per.get(r.asin)
        if g is None:
            g = per[r.asin] = {"ads": 0, "camp_types": {},
                               "impressions": 0, "clicks": 0, "orders": 0, "units": 0,
                               "spend": 0.0, "sales": 0.0}
        g["ads"] += 1
        if r.campaign_id:
            g["camp_types"][r.campaign_id] = r.campaign_type or "manual"
            acct_types[r.campaign_id] = r.campaign_type or "manual"
        g["impressions"] += r.impressions or 0
        g["clicks"] += r.clicks or 0
        g["orders"] += r.orders or 0
        g["units"] += r.units or 0
        g["spend"] += r.spend or 0
        g["sales"] += r.sales or 0
    asins = {}
    for asin, g in per.items():
        tc = _type_counts(g["camp_types"])
        m = M.all_metrics(g["impressions"], g["clicks"], g["spend"], g["sales"],
                          g["orders"], g["units"])
        asins[asin] = {"ads": g["ads"], "campaigns": tc["total"],
                       "auto_campaigns": tc["auto"], "keyword_campaigns": tc["keyword"],
                       "product_campaigns": tc["product"], "manual_campaigns": tc["manual"],
                       **m}
    return {"asins": asins, "campaign_types": _type_counts(acct_types)}


# ---- exec report (xlsx with native Excel charts) -----------------------------
def report_xlsx(db: Session) -> bytes:
    """Client-ready Product Ads workbook: Overview (account KPIs, status pie,
    campaign-type bar) + Products sheet (one row per ASIN+SKU, top-spend bar)."""
    import io
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill

    s = summary(db)
    if not s["rows"]:
        raise ValueError("No Product Ads data yet — upload a Sponsored Products bulk first.")

    TITLE = Font(bold=True, size=14)
    H = Font(bold=True, color="1F2329")
    HFILL = PatternFill("solid", fgColor="FCD535")
    SUB = Font(bold=True, size=11)

    def head(ws, row, labels, widths=None):
        for i, lab in enumerate(labels, start=1):
            c = ws.cell(row=row, column=i, value=lab)
            c.font = H; c.fill = HFILL
            if widths and i <= len(widths) and widths[i - 1]:
                ws.column_dimensions[c.column_letter].width = widths[i - 1]
        return row + 1

    wb = Workbook()

    # ---- Overview ---------------------------------------------------------------
    ws = wb.active
    ws.title = "Overview"
    ws.cell(row=1, column=1, value="Product Ads report").font = TITLE
    t = s["total"]
    ws.cell(row=2, column=1, value=f"{s['count']} products (ASIN+SKU) · "
            f"{s['campaign_types']['total']} distinct campaigns")
    r = 4
    for lab, val in [("Ad spend", t["spend"]), ("Ad sales", t["sales"]),
                     ("Orders (PPC)", t["orders"]), ("Units", t["units"]),
                     ("Clicks", t["clicks"]), ("Impressions", t["impressions"]),
                     ("ACoS", t["acos"]), ("Avg product ACoS", s.get("avg_acos")),
                     ("ROAS", t["roas"]), ("CPC", t["cpc"]),
                     ("CTR", t["ctr"]), ("CVR", t["cvr"])]:
        ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14

    # status breakdown + pie
    r += 1
    ws.cell(row=r, column=1, value="Product status").font = SUB
    r += 1
    r = head(ws, r, ["Status", "Products"], widths=[16, 10])
    first = r
    label = {"ok": "converting", "no_orders": "no orders", "no_data": "no traffic"}
    for k, n in sorted(s["by_status"].items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=1, value=label.get(k, k)); ws.cell(row=r, column=2, value=n); r += 1
    pie = PieChart(); pie.title = "Products by status"; pie.height = 7; pie.width = 11
    pie.add_data(Reference(ws, min_col=2, min_row=first, max_row=r - 1))
    pie.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
    ws.add_chart(pie, "D4")

    # campaign types + bar
    r += 1
    ws.cell(row=r, column=1, value="Campaigns by targeting kind").font = SUB
    r += 1
    hdr = r
    r = head(ws, r, ["Kind", "Campaigns"], widths=[16, 11])
    first = r
    ct = s["campaign_types"]
    for k, lab in (("auto", "Automatic"), ("keyword", "Keyword target"),
                   ("product", "Product target"), ("manual", "Manual (unknown)")):
        if ct.get(k):
            ws.cell(row=r, column=1, value=lab); ws.cell(row=r, column=2, value=ct[k]); r += 1
    if r > first:
        ch = BarChart(); ch.type = "col"; ch.title = "Campaigns by targeting kind"
        ch.height = 7; ch.width = 11
        ch.add_data(Reference(ws, min_col=2, min_row=hdr, max_row=r - 1), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
        ws.add_chart(ch, "D19")

    # ---- Products ------------------------------------------------------------------
    ws = wb.create_sheet("Products")
    hdr = 1
    r = head(ws, hdr, ["ASIN", "SKU", "Status", "State", "Ads", "Campaigns", "Auto",
                       "KW tgt", "PT tgt", "Unit price", "AOV", "Spend", "Sales",
                       "Orders", "Clicks", "Impressions", "ACoS", "ROAS", "CPC", "CTR", "CVR"],
             widths=[14, 18, 11, 9, 6, 10, 6, 7, 7, 10, 9, 10, 10, 8, 8, 12, 9, 8, 8, 8, 8])
    first = r
    rows = sorted(s["rows"], key=lambda x: -(x["spend"] or 0))
    for p in rows:
        for i, v in enumerate([p["asin"], p["sku"], label.get(p["status"], p["status"]),
                               p["state"], p["ads"], p["campaigns"], p["auto_campaigns"],
                               p["keyword_campaigns"], p["product_campaigns"], p["unit_price"],
                               p["aov"], p["spend"], p["sales"], p["orders"], p["clicks"],
                               p["impressions"], p["acos"], p["roas"], p["cpc"], p["ctr"],
                               p["cvr"]], start=1):
            ws.cell(row=r, column=i, value=v)
        r += 1
    top = min(len(rows), 15)
    if top:
        ch = BarChart(); ch.type = "col"; ch.title = f"Ad spend — top {top} products"
        ch.height = 9; ch.width = 22
        ch.add_data(Reference(ws, min_col=12, min_row=hdr, max_row=hdr + top), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=hdr + top))
        ws.add_chart(ch, "W2")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- drill-down -------------------------------------------------------------
def detail(db: Session, asins: list[str]) -> dict:
    """For the selected ASIN(s): their Product Ads + per-campaign rollups, all from
    `ProductAdFact`. (No audit flags / strategy — Product Ads is a separate model.)"""
    _ensure_schema(db)
    want = list(dict.fromkeys(a for a in asins if a))
    wset = set(want)
    by_asin: dict[str, list] = {a: [] for a in want}
    for r in db.query(md.ProductAdFact).filter(md.ProductAdFact.asin.in_(wset)).all():
        by_asin.setdefault(r.asin, []).append(r)

    rows = []
    grand = {"impressions": 0, "clicks": 0, "orders": 0, "spend": 0.0, "sales": 0.0}
    for asin in want:
        recs = by_asin.get(asin, [])
        ads, camps = [], {}
        agg = {"impressions": 0, "clicks": 0, "orders": 0, "spend": 0.0, "sales": 0.0}
        for r in recs:
            m = M.all_metrics(r.impressions, r.clicks, r.spend, r.sales, r.orders, r.units)
            ads.append({"ad_id": r.ad_id, "sku": r.sku, "state": r.state,
                        "campaign": r.campaign_name or r.campaign_id,
                        "ad_group": r.ad_group_name or r.ad_group_id, "metrics": m})
            for k in agg:
                agg[k] += (getattr(r, k) or 0)
            c = camps.get(r.campaign_id)
            if c is None:
                c = camps[r.campaign_id] = {"campaign_id": r.campaign_id,
                                            "name": r.campaign_name or r.campaign_id, "state": r.state,
                                            "type": r.campaign_type or "manual",
                                            "impressions": 0, "clicks": 0, "orders": 0, "spend": 0.0, "sales": 0.0}
            for k in ("impressions", "clicks", "orders", "spend", "sales"):
                c[k] += (getattr(r, k) or 0)
        for k in grand:
            grand[k] += agg[k]
        campaigns = [{"campaign_id": c["campaign_id"], "name": c["name"], "state": c["state"],
                      "type": c["type"],
                      "metrics": M.all_metrics(c["impressions"], c["clicks"], c["spend"], c["sales"], c["orders"], 0)}
                     for c in sorted(camps.values(), key=lambda x: -x["spend"])]
        type_counts = _type_counts({c["campaign_id"]: c["type"] for c in camps.values()})
        rows.append({"asin": asin,
                     "total": M.all_metrics(agg["impressions"], agg["clicks"], agg["spend"],
                                            agg["sales"], agg["orders"], 0),
                     "campaign_types": type_counts,
                     "ads": ads, "campaigns": campaigns})

    combined = M.all_metrics(grand["impressions"], grand["clicks"], grand["spend"],
                             grand["sales"], grand["orders"], 0)
    return {"asins": want, "count": len(rows), "total": combined, "rows": rows}

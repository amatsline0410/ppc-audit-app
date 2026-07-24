"""Reporting summary = the one-screen exec report + Excel export.

Assembles, from the deterministic engine, what a PPC manager hands a client/boss:
account health, headline PPC KPIs, period-over-period delta, the flag/action
breakdown and top movers. Pure aggregation over existing pipelines. (The profit
P&L / Sales-report features were removed, so this is PPC-only.)
"""
from __future__ import annotations
import io
from collections import Counter
from sqlalchemy.orm import Session
from ..config import Thresholds
from . import audit as audit_stage, sales as sales_stage, trends as trends_stage


def _health(acos, target_acos: float, bleeding: int, flag_total: int) -> dict:
    if not flag_total:
        return {"status": "healthy", "label": "Healthy", "note": "No open flags."}
    if bleeding > 0:
        return {"status": "unprofitable", "label": "Bleeding",
                "note": f"{bleeding} target(s) above break-even — losing money on those."}
    if acos is not None and target_acos and acos > target_acos:
        return {"status": "ad_heavy", "label": "Over goal ACoS",
                "note": f"ACoS {acos:.0%} over goal {target_acos:.0%}."}
    return {"status": "leaking", "label": "Needs attention",
            "note": "Open flags to action."}


def _seo_block(db: Session) -> dict | None:
    """SEO section of the exec report. Two sources, both optional:
    catalog = store-level listing-quality issues (Product Benchmark seo_check
    over every non-parent product); projects = keyword-based recommendation
    counts per Product Optimization project (base db — tracker is
    cadence-agnostic). None when neither has data."""
    from .. import database as dbmod
    from . import catalog as cat
    from . import tracker as tk

    out = {"catalog": None, "projects": []}
    store = db.info.get("store")
    if not store:
        return None

    prods = (cat.read_catalog(store) or {}).get("products", {})
    live = [p for p in prods.values() if p.get("parentage") != "parent"]
    if live:
        checked = [(p, cat.seo_check(p)) for p in live]
        with_issues = [(p, s) for p, s in checked if s and s["issues"]]
        areas = Counter()
        for _, s in with_issues:
            for r in s["recommendations"]:
                if r["severity"] in ("high", "medium"):
                    areas[r["area"]] += 1
        worst = sorted(with_issues, key=lambda x: (-x[1]["issues"], -x[1]["counts"]["high"]))[:5]
        out["catalog"] = {
            "products": len(live), "with_issues": len(with_issues),
            "by_area": dict(areas),
            "worst": [{"sku": p.get("sku"), "asin": p.get("asin"), "title": p.get("title"),
                       "issues": s["issues"], "high": s["counts"]["high"]} for p, s in worst],
        }

    base = dbmod.get_session(store, db.info.get("project") or dbmod.DEFAULT_PROJECT)
    try:
        for pr in tk.list_projects(base):
            if not pr["keywords"]:
                continue
            r = tk.seo_recommend(base, pr["id"])
            st = r["search_terms"]
            top = [k["keyword"] for rec in r["recommendations"]
                   if rec["area"] in ("title", "coverage") for k in rec["keywords"]][:5]
            out["projects"].append({
                "id": pr["id"], "name": pr["name"], "keywords": pr["keywords"],
                "high": r["counts"]["high"], "medium": r["counts"]["medium"],
                "st_current_bytes": st["current_bytes"],
                "st_over": st["current_bytes"] > st["max_bytes"],
                "top_keywords": top,
            })
    finally:
        base.close()

    if not out["catalog"] and not out["projects"]:
        return None
    return out


def summary(db: Session, th: Thresholds) -> dict:
    flags = audit_stage.audit(db, th)
    acct = sales_stage.total_ad(db)
    tr = trends_stage.compare(db)

    by_flag = Counter(f.flag for f in flags)
    ladder = Counter(f.stage for f in flags if f.flag == "HIGH_ACOS" and f.stage)
    wasted = round(sum(f.observed or 0 for f in flags if f.flag == "WASTED_SPEND"), 2)
    acos = round(acct["spend"] / acct["sales"], 4) if acct["sales"] else None
    snap = audit_stage.active_snapshot(db)
    # Avg Product ACoS (reporting spec): Σ ad spend ÷ Σ ad revenue across the
    # per-PRODUCT campaign rollups (build_tree, cached)
    p_sp = p_sa = 0.0
    for node in audit_stage.build_tree(db).values():
        for c in node["campaigns"]:
            m = c["metrics"]
            p_sp += m["spend"]; p_sa += m["sales"]
    avg_product_acos = round(p_sp / p_sa, 4) if p_sa else None

    actions = {
        "reduce_bid": ladder.get("REDUCE", 0),
        "monitor": ladder.get("MONITOR", 0),
        "pause": ladder.get("PAUSE", 0),
        "negate_wasted": by_flag.get("WASTED_SPEND", 0),
        "cut_bleeding": by_flag.get("BLEEDING", 0),
        "scale_winners": by_flag.get("SCALE_WINNER", 0),
    }
    period = None
    if tr.get("have_history"):
        period = {"previous": tr["previous"], "current": tr["current"], "delta": tr["totals"]["delta"]}

    # per-campaign state classifier (the Strategy methodology map): counts +
    # the top spenders for the exec view; the Excel export carries the full list
    from . import strategy as strategy_stage
    st = strategy_stage.account_states(db, th)

    return {
        "snapshot": snap.isoformat() if snap else None,
        "target_acos": th.target_acos,
        "target_roas": round(1 / th.target_acos, 2) if th.target_acos else None,
        "health": _health(acos, th.target_acos, by_flag.get("BLEEDING", 0), len(flags)),
        "kpis": {"ad_spend": acct["spend"], "ad_sales": acct["sales"], "acos": acos,
                 "avg_product_acos": avg_product_acos},
        "flags": {"total": len(flags), "by_type": dict(by_flag)},
        "actions": actions,
        "wasted_spend": wasted,
        "period": period,
        "movers": tr.get("movers", [])[:5],
        "states": {"counts": st["counts"], "top": st["campaigns"][:10]},
        "seo": _seo_block(db),
    }


def export_xlsx(db: Session, th: Thresholds) -> bytes:
    """Client-ready exec workbook with native Excel charts: Summary (KPIs +
    flag-breakdown bar + action pie), Flags (full table), Top movers (spend-Δ
    bar) and SEO (issue-by-area bar + worst offenders + per-project recs)."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill

    s = summary(db, th)
    flags = audit_stage.audit(db, th)
    k = s["kpis"]

    TITLE = Font(bold=True, size=14)
    H = Font(bold=True, color="1F2329")
    HFILL = PatternFill("solid", fgColor="FCD535")
    SUB = Font(bold=True, size=11)

    def head(ws, row, labels, widths=None):
        for i, lab in enumerate(labels, start=1):
            c = ws.cell(row=row, column=i, value=lab)
            c.font = H; c.fill = HFILL
            if widths and i <= len(widths) and widths[i - 1]:
                ws.column_dimensions[c.column_letter].width = widths[i - 1]
        return row + 1

    wb = Workbook()

    # ---- Summary ---------------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="PPC exec report").font = TITLE
    ws.cell(row=2, column=1, value=f"Snapshot {s['snapshot'] or '—'} · account health: "
            f"{s['health']['label']} — {s['health']['note']}")
    r = 4
    for lab, val in [("Goal ACoS", th.target_acos), ("Goal ROAS", s["target_roas"]),
                     ("Ad spend", k.get("ad_spend")), ("Ad sales", k.get("ad_sales")),
                     ("ACoS", k.get("acos")), ("Avg product ACoS", k.get("avg_product_acos")),
                     ("Wasted spend", s["wasted_spend"]),
                     ("Flags total", s["flags"]["total"]),
                     ("Targets over break-even (BLEEDING)", s["actions"]["cut_bleeding"])]:
        ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16

    # flag breakdown table + bar chart
    r += 1
    ws.cell(row=r, column=1, value="Flag breakdown").font = SUB
    r += 1
    hdr = r
    r = head(ws, r, ["Flag", "Count"], widths=[24, 10])
    first = r
    for flag, n in sorted(s["flags"]["by_type"].items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=1, value=flag); ws.cell(row=r, column=2, value=n); r += 1
    if r > first:
        ch = BarChart(); ch.type = "col"; ch.title = "Open flags by type"; ch.height = 8; ch.width = 15
        ch.add_data(Reference(ws, min_col=2, min_row=hdr, max_row=r - 1), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
        ws.add_chart(ch, "D4")

    # action checklist + pie
    r += 1
    ws.cell(row=r, column=1, value="Action checklist").font = SUB
    r += 1
    r = head(ws, r, ["Action", "Targets"], widths=[24, 10])
    first = r
    for a, n in sorted(s["actions"].items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=1, value=a.replace("_", " ")); ws.cell(row=r, column=2, value=n); r += 1
    acted = [n for n in s["actions"].values() if n]
    if acted:
        pie = PieChart(); pie.title = "Actions"; pie.height = 8; pie.width = 12
        pie.add_data(Reference(ws, min_col=2, min_row=first, max_row=r - 1))
        pie.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
        ws.add_chart(pie, "D22")

    # ---- Flags ------------------------------------------------------------------
    ws = wb.create_sheet("Flags")
    r = head(ws, 1, ["Flag", "Stage", "Severity", "Target", "ASIN", "Observed",
                     "BE ACoS", "Threshold", "New bid", "Action"],
             widths=[16, 10, 9, 34, 13, 10, 10, 10, 9, 44])
    for f in flags:
        for i, v in enumerate([f.flag, f.stage, f.severity, f.label or f.entity_id, f.asin,
                               f.observed, f.break_even, f.threshold, f.new_bid,
                               f.suggested_action], start=1):
            ws.cell(row=r, column=i, value=v)
        r += 1

    # ---- Account states (the methodology map's classifier) -------------------------
    from . import strategy as strategy_stage
    st = strategy_stage.account_states(db, th)
    if st["campaigns"]:
        ws = wb.create_sheet("Account states")
        ws.cell(row=1, column=1, value="Account states — ACoS vs break-even, per campaign").font = TITLE
        LABEL = {"below_target": "Below target (grow)", "at_target": "At target (balance)",
                 "above_target": "Above target (cut)",
                 "over_break_even": "Over break-even (cut hard)", "no_data": "No data (rank)"}
        r = head(ws, 3, ["State", "Campaigns"], widths=[30, 11])
        first = r
        for k, n in st["counts"].items():
            ws.cell(row=r, column=1, value=LABEL.get(k, k)); ws.cell(row=r, column=2, value=n); r += 1
        pie = PieChart(); pie.title = "Campaigns by state"; pie.height = 8; pie.width = 12
        pie.add_data(Reference(ws, min_col=2, min_row=first, max_row=r - 1))
        pie.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
        ws.add_chart(pie, "D3")

        r += 1
        ws.cell(row=r, column=1, value="Per campaign").font = SUB
        r += 1
        r = head(ws, r, ["State", "Campaign", "ASIN", "Spend", "Sales", "Orders",
                         "ACoS", "BE ACoS", "Goal", "Lever", "Why"],
                 widths=[18, 44, 13, 10, 10, 8, 8, 9, 8, 10, 40])
        for x in st["campaigns"]:
            ws.append([LABEL.get(x["state"], x["state"]).split(" (")[0], x["campaign"], x["asin"],
                       x["spend"], x["sales"], x["orders"], x["acos"], x["break_even"],
                       x["goal"], x["lever"], x["why"]])
            r += 1

    # ---- Top movers ---------------------------------------------------------------
    if s["movers"]:
        ws = wb.create_sheet("Top movers")
        ws.cell(row=1, column=1, value=f"Top movers · {s['period']['previous']} → {s['period']['current']}"
                if s.get("period") else "Top movers").font = SUB
        hdr = 3
        r = head(ws, hdr, ["Campaign", "Spend Δ", "Spend", "Sales", "Orders Δ", "ACoS prev", "ACoS cur"],
                 widths=[36, 10, 10, 10, 10, 10, 10])
        first = r
        for m in s["movers"]:
            ws.cell(row=r, column=1, value=m.get("name") or m.get("campaign_id"))
            ws.cell(row=r, column=2, value=m.get("spend_delta"))
            ws.cell(row=r, column=3, value=m.get("spend_cur"))
            ws.cell(row=r, column=4, value=m.get("sales_cur"))
            ws.cell(row=r, column=5, value=m.get("orders_delta"))
            ws.cell(row=r, column=6, value=m.get("acos_prev"))
            ws.cell(row=r, column=7, value=m.get("acos_cur"))
            r += 1
        ch = BarChart(); ch.type = "col"; ch.title = "Spend Δ by campaign"; ch.height = 8; ch.width = 16
        ch.add_data(Reference(ws, min_col=2, min_row=hdr, max_row=r - 1), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
        ws.add_chart(ch, "I3")

    # ---- SEO ------------------------------------------------------------------------
    seo = s.get("seo")
    if seo:
        ws = wb.create_sheet("SEO")
        ws.cell(row=1, column=1, value="SEO · listing quality & keyword coverage").font = TITLE
        r = 3
        if seo.get("catalog"):
            c = seo["catalog"]
            ws.cell(row=r, column=1,
                    value=f"{c['with_issues']} of {c['products']} catalog products carry high/medium listing issues").font = SUB
            r += 2
            hdr = r
            r = head(ws, r, ["Area", "Issues"], widths=[18, 10])
            first = r
            for area, n in sorted(c["by_area"].items(), key=lambda x: -x[1]):
                ws.cell(row=r, column=1, value=area.replace("_", " ")); ws.cell(row=r, column=2, value=n); r += 1
            ch = BarChart(); ch.type = "col"; ch.title = "Listing issues by area"; ch.height = 8; ch.width = 14
            ch.add_data(Reference(ws, min_col=2, min_row=hdr, max_row=r - 1), titles_from_data=True)
            ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
            ws.add_chart(ch, "E3")
            r += 1
            ws.cell(row=r, column=1, value="Worst offenders").font = SUB
            r += 1
            r = head(ws, r, ["SKU", "ASIN", "Title", "Issues (high+med)", "High"],
                     widths=[18, 14, 52, 17, 8])
            for x in c["worst"]:
                for i, v in enumerate([x["sku"], x["asin"], x["title"], x["issues"], x["high"]], start=1):
                    ws.cell(row=r, column=i, value=v)
                r += 1
        if seo.get("projects"):
            r += 1
            ws.cell(row=r, column=1, value="Product Optimization projects · keyword-based recs").font = SUB
            r += 1
            r = head(ws, r, ["Project", "Tracked keywords", "High recs", "Medium recs",
                             "Backend bytes", "Over 249B", "Top keywords to work in"],
                     widths=[34, 16, 10, 12, 13, 10, 50])
            for p in seo["projects"]:
                for i, v in enumerate([p["name"], p["keywords"], p["high"], p["medium"],
                                       p["st_current_bytes"], "YES" if p["st_over"] else "",
                                       ", ".join(p["top_keywords"])], start=1):
                    ws.cell(row=r, column=i, value=v)
                r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

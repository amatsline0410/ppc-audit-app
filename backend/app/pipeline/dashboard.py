"""Main Dashboard analytics hub — cross-feature rollup + charted Excel export.

`analytics(db, bdb)` assembles one payload from every data source (Product Ads,
Product Benchmark catalog, the store transaction ledger, the Monitoring daily
tracker, snapshot movers, Change Log, mined keywords); each block is None when
its source has no upload yet. `export_xlsx(db, bdb, th)` mirrors the Dashboard
into a client-ready workbook with native Excel charts.

db  = the audit/cadence-scoped session (get_db) — PPC facts, Product Ads, logs.
bdb = the base-file session (get_base_db)     — Monitoring's FactDaily.
"""
from __future__ import annotations

from datetime import timedelta

from sqlalchemy.orm import Session

from .. import database as dbmod
from .. import models as md
from ..config import Thresholds


def analytics(db: Session, bdb: Session, th: Thresholds | None = None) -> dict:
    """Cross-feature analytics blocks for the Main Dashboard (one fast call).
    With `th`, also runs the per-campaign account-state classifier (the Strategy
    methodology map's four ACoS-vs-break-even states)."""
    from . import benchmark as bench_stage
    from . import catalog as cat
    from . import monitoring as mon
    from . import productads as pa
    from . import strategy as strategy_stage
    from . import transactions as txn
    from . import trends as trends_stage

    out = {"product_ads": None, "catalog": None, "transactions": None, "monitoring": None,
           "movers": None, "period": None, "changelog": None, "keywords": None,
           "states": None}
    if th is not None:
        st = strategy_stage.account_states(db, th)
        if st["campaigns"]:
            out["states"] = {"counts": st["counts"], "top": st["campaigns"][:8]}
    # store-level files key off the SCOPED store id (per-user namespaced), not the
    # raw query param — same as the catalog router's _store(db).
    sid = db.info.get("store")

    ads_ok = pa.has_data(db)
    if ads_ok:
        s = pa.summary(db)
        out["product_ads"] = {"count": s["count"], "by_status": s["by_status"],
                              "campaigns": s["campaign_types"]["total"],
                              "total": s["total"], "avg_acos": s.get("avg_acos")}

    data = cat.read_catalog(sid)
    if data:
        ads = pa.by_asin(db) if ads_ok else None
        econ = dbmod.get_project_econ(sid, db.info.get("project"))
        bench = bench_stage._read_store(sid)
        out["catalog"] = cat.enrich(cat.overview(data), ads, bench, econ,
                                    cat.read_cogs(sid), cat.fees_by_sku(sid))["stats"]

    ledger = txn.read_txn(sid)
    if ledger.get("rows"):
        s = txn.summary(ledger)
        out["transactions"] = {"range": s["range"], "totals": s["totals"],
                               "days": s["days"], "top_skus": s["skus"][:5]}

    last = bdb.query(md.FactDaily.date).order_by(md.FactDaily.date.desc()).first()
    if last:
        m = mon.summary(bdb, last[0] - timedelta(days=13), last[0])
        ov = m["overview"]
        out["monitoring"] = {"latest_date": m["latest_date"], "health": m["health"],
                             "alerts": m["alerts"][:4], "data_days": m["data_days"],
                             "overview": {k: ov.get(k) for k in
                                          ("total_sales", "ad_spend", "ad_sales", "acos", "tacos")}}

    tr = trends_stage.compare(db)
    if tr.get("have_history"):
        out["movers"] = tr.get("movers", [])[:5]
        out["period"] = {"previous": tr["previous"], "current": tr["current"],
                         "delta": tr["totals"]["delta"]}

    n_log = db.query(md.ChangeLog).count()
    if n_log:
        last_log = db.query(md.ChangeLog).order_by(md.ChangeLog.id.desc()).first()
        out["changelog"] = {"count": n_log,
                            "last": {"ts": last_log.ts.isoformat() if last_log.ts else None,
                                     "field": last_log.field, "label": last_log.label}}

    n_kw = db.query(md.MinedKeyword).count()
    if n_kw:
        out["keywords"] = {"count": n_kw}
    return out


def export_xlsx(db: Session, bdb: Session, th: Thresholds) -> bytes:
    """Client-ready Dashboard workbook: Overview (PPC KPIs + flag-breakdown bar),
    Features (Product Ads status pie + catalog + monitoring blocks), Transactions
    (daily sales/net line + top-SKU net bar), Top movers (spend-Δ bar)."""
    import io

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill

    from . import report as report_stage

    s = report_stage.summary(db, th)          # PPC kpis / flags / actions / movers
    a = analytics(db, bdb, th)

    TITLE = Font(bold=True, size=14)
    H = Font(bold=True, color="1F2329")
    HFILL = PatternFill("solid", fgColor="FCD535")
    SUB = Font(bold=True, size=11)

    def head(ws, row, labels, widths=None):
        for i, lab in enumerate(labels, start=1):
            c = ws.cell(row=row, column=i, value=lab)
            c.font = H; c.fill = HFILL
            if widths and i <= len(widths) and widths[i - 1]:
                ws.column_dimensions[c.column_letter].width = widths[i - 1]
        return row + 1

    wb = Workbook()

    # ---- Overview: PPC KPIs + flag breakdown ---------------------------------------
    ws = wb.active
    ws.title = "Overview"
    ws.cell(row=1, column=1, value="Dashboard report").font = TITLE
    sub = f"snapshot {s['snapshot'] or '—'}"
    if th.target_acos:
        sub += f" · goal ACoS {th.target_acos:.0%}"
    ws.cell(row=2, column=1, value=sub)
    k = s["kpis"]
    r = 4
    for lab, val in [("Ad spend", k["ad_spend"]), ("Ad sales", k["ad_sales"]),
                     ("ACoS", k["acos"]), ("Avg product ACoS", k["avg_product_acos"]),
                     ("Open flags", s["flags"]["total"]), ("Wasted spend", s["wasted_spend"]),
                     ("Reduce bids", s["actions"]["reduce_bid"]), ("Monitor", s["actions"]["monitor"]),
                     ("Pause", s["actions"]["pause"]), ("Scale winners", s["actions"]["scale_winners"])]:
        ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14

    r += 1
    ws.cell(row=r, column=1, value="Flag breakdown").font = SUB
    r += 1
    r = head(ws, r, ["Flag", "Count"], widths=None)
    first = r
    for flag, n in sorted(s["flags"]["by_type"].items(), key=lambda x: -x[1]):
        ws.cell(row=r, column=1, value=flag); ws.cell(row=r, column=2, value=n); r += 1
    if r > first:
        bar = BarChart(); bar.type = "col"; bar.title = "Flags by type"
        bar.height = 8; bar.width = 16; bar.legend = None
        bar.add_data(Reference(ws, min_col=2, min_row=first - 1, max_row=r - 1), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
        ws.add_chart(bar, "D4")

    # account states (the Strategy methodology classifier) + pie
    if a["states"]:
        LABEL = {"below_target": "Below target (grow)", "at_target": "At target (balance)",
                 "above_target": "Above target (cut)",
                 "over_break_even": "Over break-even (cut hard)", "no_data": "No data (rank)"}
        r += 1
        ws.cell(row=r, column=1, value="Account states").font = SUB
        r += 1
        r = head(ws, r, ["State", "Campaigns"], widths=None)
        first = r
        for k, n in a["states"]["counts"].items():
            ws.cell(row=r, column=1, value=LABEL.get(k, k)); ws.cell(row=r, column=2, value=n); r += 1
        pie = PieChart(); pie.title = "Campaigns by state"; pie.height = 8; pie.width = 12
        pie.add_data(Reference(ws, min_col=2, min_row=first, max_row=r - 1))
        pie.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
        ws.add_chart(pie, "D22")

    # ---- Features: Product Ads / Catalog / Monitoring --------------------------------
    ws = wb.create_sheet("Features")
    ws.cell(row=1, column=1, value="Feature analytics").font = TITLE
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    r = 3
    pa = a["product_ads"]
    ws.cell(row=r, column=1, value="Product Ads (this audit)").font = SUB
    r += 1
    if pa:
        acos = pa["total"].get("acos")
        for lab, val in [("Products", pa["count"]), ("Campaigns", pa["campaigns"]),
                         ("Ad spend", pa["total"].get("spend")), ("Ad sales", pa["total"].get("sales")),
                         ("ACoS", round(acos, 4) if acos is not None else None),
                         ("Avg product ACoS", pa["avg_acos"])]:
            ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
            ws.cell(row=r, column=2, value=val); r += 1
        r += 1
        r = head(ws, r, ["Status", "Products"], widths=None)
        first = r
        label = {"ok": "converting", "no_orders": "no orders", "no_data": "no traffic"}
        for kk, n in sorted(pa["by_status"].items(), key=lambda x: -x[1]):
            ws.cell(row=r, column=1, value=label.get(kk, kk)); ws.cell(row=r, column=2, value=n); r += 1
        pie = PieChart(); pie.title = "Products by status"; pie.height = 7; pie.width = 11
        pie.add_data(Reference(ws, min_col=2, min_row=first, max_row=r - 1))
        pie.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
        ws.add_chart(pie, "D3")
    else:
        ws.cell(row=r, column=1, value="no Product Ads upload in this audit"); r += 1

    r += 2
    ct = a["catalog"]
    ws.cell(row=r, column=1, value="Product Benchmark (store catalog)").font = SUB
    r += 1
    if ct:
        for lab, val in [("Products", ct.get("total")), ("Listing issues", ct.get("listing_issues")),
                         ("Advertised", ct.get("advertised")), ("Avg ACoS", ct.get("avg_acos")),
                         ("Over break-even", ct.get("over_be")), ("Under break-even", ct.get("under_be"))]:
            ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
            ws.cell(row=r, column=2, value=val); r += 1
    else:
        ws.cell(row=r, column=1, value="no catalog uploaded"); r += 1

    r += 2
    mo = a["monitoring"]
    ws.cell(row=r, column=1, value="Monitoring (last 14 days)").font = SUB
    r += 1
    if mo:
        ov = mo["overview"]
        for lab, val in [("Latest day", mo["latest_date"]),
                         ("Health score (0-100)", (mo["health"] or {}).get("score")),
                         ("Total sales", ov.get("total_sales")), ("Ad spend", ov.get("ad_spend")),
                         ("Ad sales", ov.get("ad_sales")), ("ACoS %", ov.get("acos")),
                         ("TACoS %", ov.get("tacos"))]:
            ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
            ws.cell(row=r, column=2, value=val); r += 1
        for al in mo["alerts"]:
            ws.cell(row=r, column=1, value=f"⚠ {al['severity']}: {al['msg']}"); r += 1
    else:
        ws.cell(row=r, column=1, value="no daily tracker data"); r += 1

    r += 2
    for blk, lab in ((a["changelog"], "Actions logged"), (a["keywords"], "Keywords mined")):
        if blk:
            ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
            ws.cell(row=r, column=2, value=blk["count"]); r += 1

    # ---- Transactions: daily trend + top SKUs ----------------------------------------
    tx = a["transactions"]
    if tx:
        ws = wb.create_sheet("Transactions")
        ws.cell(row=1, column=1, value=f"SKU ledger · {tx['range']['min']} → {tx['range']['max']}").font = TITLE
        t = tx["totals"]
        r = 3
        for lab, val in [("Orders", t["orders"]), ("Refunds", t["refunds"]), ("Units", t["units"]),
                         ("Product sales", t["product_sales"]),
                         ("Selling fees", t["selling_fees"]), ("FBA fees", t["fba_fees"]),
                         ("Net proceeds", t["net"])]:
            ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
            ws.cell(row=r, column=2, value=val); r += 1
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 14

        r += 1
        ws.cell(row=r, column=1, value="Top SKUs by net proceeds").font = SUB
        r += 1
        r = head(ws, r, ["SKU", "Net proceeds"], widths=[30, 14])
        first = r
        for p in tx["top_skus"]:
            ws.cell(row=r, column=1, value=p["sku"]); ws.cell(row=r, column=2, value=p["net"]); r += 1
        if r > first:
            bar = BarChart(); bar.type = "col"; bar.title = "Top SKUs by net proceeds"
            bar.height = 7; bar.width = 14; bar.legend = None
            bar.add_data(Reference(ws, min_col=2, min_row=first - 1, max_row=r - 1), titles_from_data=True)
            bar.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
            ws.add_chart(bar, "H14")

        r += 1
        ws.cell(row=r, column=1, value="Daily trend").font = SUB
        r += 1
        r = head(ws, r, ["Date", "Orders", "Refunds", "Units", "Product sales", "Net proceeds"],
                 widths=None)
        first = r
        for d in tx["days"]:
            ws.cell(row=r, column=1, value=d["date"]); ws.cell(row=r, column=2, value=d["orders"])
            ws.cell(row=r, column=3, value=d["refunds"]); ws.cell(row=r, column=4, value=d["units"])
            ws.cell(row=r, column=5, value=d["product_sales"]); ws.cell(row=r, column=6, value=d["net"])
            r += 1
        if r > first:
            ln = LineChart(); ln.title = "Product sales vs net proceeds"; ln.height = 9; ln.width = 24
            ln.add_data(Reference(ws, min_col=5, max_col=6, min_row=first - 1, max_row=r - 1),
                        titles_from_data=True)
            ln.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
            ws.add_chart(ln, "H3")

    # ---- Top movers -------------------------------------------------------------------
    if a["movers"]:
        ws = wb.create_sheet("Top movers")
        p = a["period"]
        ws.cell(row=1, column=1, value=f"Top movers · {p['previous']} → {p['current']}"
                if p else "Top movers").font = TITLE
        r = head(ws, 3, ["Campaign", "Spend Δ", "Spend", "Sales", "ACoS prev", "ACoS cur"],
                 widths=[44, 10, 10, 10, 10, 10])
        first = r
        for x in a["movers"]:
            ws.append([x.get("name") or x.get("campaign_id"), x.get("spend_delta"),
                       x.get("spend_cur"), x.get("sales_cur"), x.get("acos_prev"), x.get("acos_cur")])
            r += 1
        bar = BarChart(); bar.type = "col"; bar.title = "Spend Δ vs previous snapshot"
        bar.height = 8; bar.width = 18; bar.legend = None
        bar.add_data(Reference(ws, min_col=2, min_row=first - 1, max_row=r - 1), titles_from_data=True)
        bar.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
        ws.add_chart(bar, "H3")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

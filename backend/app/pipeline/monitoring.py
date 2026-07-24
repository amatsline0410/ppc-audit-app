"""Daily SALES & PPC Tracker (Monitoring panel).

Consolidates two daily Amazon exports by Date — the Business Report (Sales &
Traffic by Date) and the Sponsored Products campaign report (summed to a daily
total) — into one `FactDaily` row per day, then computes the monitoring features
(deltas, rolling averages, alerts, divergence, weekday/weekend, B2B split, health
score) for a chosen date range.

Plain-language note on the cleaning: Amazon exports numbers as *text* with $ , %
in them ("$16,255.85", "24.03%"). Computers can't do math on text, so every value
is stripped to a clean number first. Percents are kept as the human number 24.03
(NOT 0.2403) — consistently, everywhere.
"""
from __future__ import annotations
import io
import re
import calendar
from datetime import date, datetime, timedelta

import pandas as pd
from sqlalchemy.orm import Session

from .. import models as md

# ---- tunable alert thresholds (one place to change them) --------------------
CVR_FALL_DAYS = 3        # CVR falling this many days in a row -> alert
BUYBOX_MIN = 90.0        # buy box below this % on the latest day -> alert
REFUND_SPIKE_MULT = 2.0  # latest refund rate above this x the 7d avg ...
REFUND_FLOOR = 3.0       # ... AND above this % -> spike alert
ASP_SWING = 0.15         # ASP moves more than ±15% vs 7d avg -> flag
ROLL_WINDOW = 7          # rolling-average window (days)
DIVERGE_WINDOW = 7       # divergence look-back window (days)
TARGET_TACOS = 12.0      # default TACOS target (%) — below it = good
ACOS_HIGH = 30.0         # overall ACOS above this -> PPC efficiency action
ROAS_LOW = 3.0           # overall ROAS below this -> PPC action
CTR_LOW = 0.3            # overall CTR (%) below this -> creative/targeting action
PPC_CVR_LOW = 8.0        # PPC CVR (orders/clicks %) below this -> relevance/listing action


# ---- string cleaning --------------------------------------------------------
def _money(v) -> float | None:
    """"$16,255.85" -> 16255.85"""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = re.sub(r"[$,\s]", "", str(v)).strip()
    if s in ("", "-", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _pct(v) -> float | None:
    """"24.03%" -> 24.03 (kept as the plain number, not 0.2403)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = re.sub(r"[%,\s]", "", str(v)).strip()
    if s in ("", "-", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _int(v) -> int | None:
    """"1,234" -> 1234"""
    f = _money(v)
    return int(round(f)) if f is not None else None


def _date(v) -> date | None:
    """"6/1/26" or "2026-06-01" -> date."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d", "%-m/%-d/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return pd.to_datetime(s).date()
    except Exception:
        return None


def _read(path: str) -> pd.DataFrame:
    from . import workbook
    return pd.read_csv(path) if path.lower().endswith(".csv") \
        else pd.read_excel(path, engine=workbook.excel_engine())


def _find(cols, *needles) -> str | None:
    """First column whose lowercased name contains every needle of any group."""
    low = {c: str(c).strip().lower() for c in cols}
    for group in needles:
        terms = group if isinstance(group, (list, tuple)) else [group]
        for c, l in low.items():
            if all(t in l for t in terms):
                return c
    return None


# ---- report type detection --------------------------------------------------
def detect_kind(path: str) -> str:
    """'business' | 'ppc' | 'unknown' — so one upload endpoint takes either."""
    if path.lower().endswith(".csv"):
        cols = pd.read_csv(path, nrows=0).columns
    else:
        from . import workbook
        cols = pd.read_excel(path, nrows=0, engine=workbook.excel_engine()).columns
    low = " | ".join(str(c).lower() for c in cols)
    if "unit session percentage" in low or "ordered product sales" in low:
        return "business"
    if "campaign name" in low and "impressions" in low:
        return "ppc"
    return "unknown"


# ---- parsers ----------------------------------------------------------------
def parse_business_daily(path: str) -> dict[date, dict]:
    """Business Report -> {date: {clean business fields}}. One row per day."""
    df = _read(path)
    c = df.columns
    col = {
        "date": _find(c, "date"),
        "ordered_sales": _find(c, ["ordered product sales"]),
        "units_ordered": _find(c, ["units ordered"]),
        "total_order_items": _find(c, ["total order items"]),
        "asp": _find(c, ["average selling price"]),
        "page_views": _find(c, ["page views", "total"]),
        "sessions": _find(c, ["sessions", "total"]),
        "buy_box_pct": _find(c, ["buy box", "percentage"], ["featured offer", "percentage"]),
        "unit_session_pct": _find(c, ["unit session percentage"]),
        "units_refunded": _find(c, ["units refunded"]),
        "refund_rate": _find(c, ["refund rate"]),
        "shipped_sales": _find(c, ["shipped product sales"]),
        "units_shipped": _find(c, ["units shipped"]),
        "orders_shipped": _find(c, ["orders shipped"]),
        "ordered_sales_b2b": _find(c, ["ordered product sales - b2b"]),
        "units_ordered_b2b": _find(c, ["units ordered - b2b"]),
        "unit_session_pct_b2b": _find(c, ["unit session percentage - b2b"]),
    }
    money = {"ordered_sales", "asp", "shipped_sales", "ordered_sales_b2b"}
    pct = {"buy_box_pct", "unit_session_pct", "refund_rate", "unit_session_pct_b2b"}

    out: dict[date, dict] = {}
    for _, r in df.iterrows():
        d = _date(r.get(col["date"]))
        if not d:
            continue
        rec = {}
        for field, src in col.items():
            if field == "date" or not src:
                continue
            raw = r.get(src)
            rec[field] = _money(raw) if field in money else _pct(raw) if field in pct else _int(raw)
        out[d] = rec
    return out


def parse_ppc_daily(path: str) -> dict[date, dict]:
    """SP campaign report (per campaign per day) -> {date: daily PPC totals}."""
    df = _read(path)
    c = df.columns
    cd = _find(c, "date")
    ci = _find(c, ["impressions"])
    cc = _find(c, ["clicks"])
    cs = _find(c, ["spend", "converted"], ["spend"])
    csa = _find(c, ["total sales", "converted"], ["7 day total sales"], ["sales", "converted"])
    co = _find(c, ["total orders"], ["orders"])
    if not cd:
        return {}

    agg: dict[date, dict] = {}
    for _, r in df.iterrows():
        d = _date(r.get(cd))
        if not d:
            continue
        a = agg.setdefault(d, {"ppc_impressions": 0, "ppc_clicks": 0,
                               "ppc_spend": 0.0, "ppc_sales": 0.0, "ppc_orders": 0})
        a["ppc_impressions"] += _int(r.get(ci)) or 0 if ci else 0
        a["ppc_clicks"] += _int(r.get(cc)) or 0 if cc else 0
        a["ppc_spend"] += _money(r.get(cs)) or 0.0 if cs else 0.0
        a["ppc_sales"] += _money(r.get(csa)) or 0.0 if csa else 0.0
        a["ppc_orders"] += _int(r.get(co)) or 0 if co else 0
    return agg


# ---- ingest (idempotent upsert by date) -------------------------------------
def ingest(db: Session, path: str) -> dict:
    """Detect the report kind, upsert its columns by date. Re-ingesting the same
    file never duplicates — we update the existing row for each date."""
    kind = detect_kind(path)
    if kind == "business":
        data = parse_business_daily(path)
    elif kind == "ppc":
        data = parse_ppc_daily(path)
    else:
        raise ValueError("unrecognized report — upload a Business Report or a "
                         "Sponsored Products campaign report")

    existing = {row.date: row for row in db.query(md.FactDaily).all()}
    for d, fields in data.items():
        row = existing.get(d)
        if row is None:
            row = md.FactDaily(date=d)
            db.add(row); existing[d] = row
        for k, v in fields.items():
            if v is not None:
                setattr(row, k, v)
    db.commit()
    return {"kind": kind, "days": len(data),
            "range": [min(data).isoformat(), max(data).isoformat()] if data else None}


def delete_all(db: Session) -> int:
    """Wipe every uploaded tracker day (FactDaily). Manual month-sales overrides
    (MonthSalesOverride) are hand-typed, not upload-derived — they survive."""
    n = db.query(md.FactDaily).delete()
    db.commit()
    return n


# ---- feature helpers --------------------------------------------------------
def _delta(cur, prev):
    """Absolute + % change vs prior day (None-safe)."""
    if cur is None or prev is None:
        return {"abs": None, "pct": None}
    ab = round(cur - prev, 2)
    pc = round((cur - prev) / prev * 100, 1) if prev else None
    return {"abs": ab, "pct": pc}


def _avg(vals):
    xs = [v for v in vals if v is not None]
    return round(sum(xs) / len(xs), 2) if xs else None


def _slope_up(vals) -> bool | None:
    """Trend direction by comparing the 2nd half's mean to the 1st half's."""
    xs = [v for v in vals if v is not None]
    if len(xs) < 4:
        return None
    h = len(xs) // 2
    a, b = _avg(xs[:h]), _avg(xs[h:])
    if a is None or b is None:
        return None
    return b > a


def _rate(num, den, scale=1.0, nd=2):
    """num/den (None-safe), optionally ×100 for a percent. None if no denominator."""
    if num is None or not den:
        return None
    return round(num / den * scale, nd)


def _row_dict(r: "md.FactDaily | None", d: date) -> dict:
    """Per-date row for the table; missing day -> all-None (UI shows '-').

    The PPC ratios (ACOS / ROAS / CPC / CTR / PPC-CVR / TACOS) are DERIVED here
    from the raw daily counts — never stored. Percents are plain numbers (27.44).
      ACOS  = ad spend / ad sales        ROAS = ad sales / ad spend
      CPC   = ad spend / clicks          CTR  = clicks / impressions
      CVR   = ad orders / clicks         TACOS = ad spend / TOTAL sales
    """
    base = {"date": d.isoformat(), "weekday": d.strftime("%a"),
            "weekend": d.weekday() >= 5}
    fields = ["ordered_sales", "units_ordered", "total_order_items", "asp",
              "page_views", "sessions", "buy_box_pct", "unit_session_pct",
              "units_refunded", "refund_rate", "shipped_sales", "units_shipped",
              "orders_shipped", "ordered_sales_b2b", "units_ordered_b2b",
              "unit_session_pct_b2b", "ppc_impressions", "ppc_clicks",
              "ppc_spend", "ppc_sales", "ppc_orders"]
    for f in fields:
        base[f] = getattr(r, f) if r else None
    # derived PPC ratios
    base["acos"] = _rate(base["ppc_spend"], base["ppc_sales"], 100)
    base["roas"] = _rate(base["ppc_sales"], base["ppc_spend"], 1.0)
    base["cpc"] = _rate(base["ppc_spend"], base["ppc_clicks"], 1.0)
    base["ctr"] = _rate(base["ppc_clicks"], base["ppc_impressions"], 100)
    base["ppc_cvr"] = _rate(base["ppc_orders"], base["ppc_clicks"], 100)
    base["tacos"] = _rate(base["ppc_spend"], base["ordered_sales"], 100)
    return base


# ---- main read API ----------------------------------------------------------
def _month_sales(db: Session, year: int, month: int) -> tuple[float | None, str]:
    """Total ordered sales for a calendar month + its source.

    Prefer the daily uploads; fall back to a manual override; else (None, 'none').
    """
    last = calendar.monthrange(year, month)[1]
    xs = [r.ordered_sales for r in db.query(md.FactDaily).filter(
            md.FactDaily.date >= date(year, month, 1),
            md.FactDaily.date <= date(year, month, last)).all()
          if r.ordered_sales is not None]
    if xs:
        return round(sum(xs), 2), "daily"
    ov = db.get(md.MonthSalesOverride, (year, month))
    if ov:
        return round(ov.sales, 2), "manual"
    return None, "none"


def set_month_sales(db: Session, year: int, month: int, sales) -> dict:
    """Set or clear (sales is None) the manual month-sales override."""
    ov = db.get(md.MonthSalesOverride, (year, month))
    if sales is None:
        if ov:
            db.delete(ov)
    elif ov:
        ov.sales = float(sales)
    else:
        db.add(md.MonthSalesOverride(year=year, month=month, sales=float(sales)))
    db.commit()
    return {"year": year, "month": month, "sales": None if sales is None else float(sales)}


def _growth(this, prior):
    """% growth vs a prior period. base 0 -> 100% when we have sales, else None."""
    if this is None:
        return None
    if not prior:
        return 100.0 if this else None
    return round((this - prior) / prior * 100, 2)


def _overview(db, rows, start: date, end: date, target_tacos: float) -> dict:
    """The 'Overall Performance' header: range totals + overall ratios, vs last
    month / last year, and a run-rate estimate for the range's month."""
    def s(attr):
        xs = [r[attr] for r in rows if r.get(attr) is not None]
        return round(sum(xs), 2) if xs else None
    total_sales = s("ordered_sales"); units = s("units_ordered")
    spend = s("ppc_spend"); sales = s("ppc_sales"); orders = s("ppc_orders")
    clicks = s("ppc_clicks"); impr = s("ppc_impressions")
    tot = {
        "total_sales": total_sales, "units_ordered": units, "ad_spend": spend,
        "ad_sales": sales, "ppc_orders": orders, "clicks": clicks, "impressions": impr,
        "acos": _rate(spend, sales, 100), "roas": _rate(sales, spend, 1.0),
        "cpc": _rate(spend, clicks, 1.0), "ctr": _rate(clicks, impr, 100),
        "cvr": _rate(orders, clicks, 100), "tacos": _rate(spend, total_sales, 100),
    }
    # anchor comparisons + estimate on the range's END month
    y, mo = end.year, end.month
    pm_y, pm_mo = (y - 1, 12) if mo == 1 else (y, mo - 1)
    lm_sales, lm_src = _month_sales(db, pm_y, pm_mo)
    ly_sales, ly_src = _month_sales(db, y - 1, mo)
    # run-rate: month-to-date sales / days with data * days in the month
    data_days = len([r for r in rows if r.get("ordered_sales") is not None])
    days_in_month = calendar.monthrange(y, mo)[1]
    est = round(total_sales / data_days * days_in_month, 2) if (total_sales and data_days) else None
    return {
        **tot,
        "target_tacos": target_tacos,
        "tacos_below_target": (tot["tacos"] is not None and tot["tacos"] <= target_tacos),
        "last_month": {"label": f"{pm_y} {calendar.month_name[pm_mo]}", "sales": lm_sales,
                       "growth": _growth(total_sales, lm_sales), "source": lm_src,
                       "year": pm_y, "month": pm_mo},
        "last_year": {"label": f"{y-1} {calendar.month_name[mo]}", "sales": ly_sales,
                      "growth": _growth(total_sales, ly_sales), "source": ly_src,
                      "year": y - 1, "month": mo},
        "estimate": {"label": f"{calendar.month_name[mo]} {y}", "sales": est,
                     "day": data_days, "days_in_month": days_in_month},
    }


def _recommendations(ov, latest, prev, streak, divergence, avg_cvr_7, avg_refund_7,
                     avg_asp_7, sess_up) -> list[dict]:
    """Turn the data into concrete next steps, split into PPC vs Listing buckets.

    PPC = ad efficiency / targeting (ACOS, TACOS, ROAS, CTR, PPC CVR, wasted spend).
    Listing = conversion / catalog health (buy box, CVR, refunds, price, traffic).
    Each: {category, severity, title, action}.
    """
    recs = []
    def add(cat, sev, title, action):
        recs.append({"category": cat, "severity": sev, "title": title, "action": action})

    # ---- PPC ----
    if ov["tacos"] is not None and ov["tacos"] > ov["target_tacos"]:
        add("ppc", "high", f"TACOS {ov['tacos']}% over target {ov['target_tacos']}%",
            "Ad spend is eating too much of total sales — trim non-converting campaigns / lower bids, or grow organic.")
    if ov["acos"] is not None and ov["acos"] > ACOS_HIGH:
        add("ppc", "high", f"ACOS {ov['acos']}% is high",
            "Cut bids on high-spend low-converting targets and add negatives; pause keywords above break-even.")
    if ov["roas"] is not None and ov["roas"] < ROAS_LOW:
        add("ppc", "medium", f"ROAS {ov['roas']}× is low",
            "Reallocate budget to your best ROAS campaigns; pause the worst.")
    if ov["ctr"] is not None and ov["ctr"] < CTR_LOW:
        add("ppc", "medium", f"CTR {ov['ctr']}% is low",
            "Refresh ad creative / main image, tighten match types, and raise bids on the most relevant terms.")
    if ov["cvr"] is not None and ov["cvr"] < PPC_CVR_LOW:
        add("ppc", "medium", f"Ad CVR {ov['cvr']}% is low",
            "Clicks aren't converting — review search-term relevance, add negatives, and check the landing listing.")
    if divergence:
        add("ppc", "high", "Traffic up but conversion down",
            "The extra sessions aren't qualified — audit search terms, add negatives, and pause broad/irrelevant targets.")

    # ---- Listing ----
    if latest and latest.buy_box_pct is not None and latest.buy_box_pct < BUYBOX_MIN:
        add("listing", "high", f"Buy Box {latest.buy_box_pct}% (< {BUYBOX_MIN}%)",
            "You don't own the box, so ads can't convert — check price vs competitors, suppressed/inactive offers, stock.")
    if streak >= CVR_FALL_DAYS:
        add("listing", "high", f"CVR fell {streak} days in a row",
            "Conversion is sliding — review price, main image, title/bullets, reviews/rating, and recent competitor moves.")
    if latest and latest.refund_rate is not None and avg_refund_7 and \
            latest.refund_rate > REFUND_SPIKE_MULT * avg_refund_7 and latest.refund_rate > REFUND_FLOOR:
        add("listing", "high", f"Refund spike {latest.refund_rate}%",
            "Likely a product/expectation-mismatch issue — check defects, listing accuracy (size/specs/images), recent batch.")
    if latest and latest.asp is not None and avg_asp_7:
        swing = (latest.asp - avg_asp_7) / avg_asp_7
        if abs(swing) > ASP_SWING:
            add("listing", "medium", f"ASP {swing*100:+.0f}% vs 7d avg",
                "Confirm this price/coupon change is intended; silent promos quietly cut margin.")
    if sess_up is False:
        add("listing", "medium", "Sessions trending down",
            "Traffic is shrinking — check organic rank/SEO, ad impression share, and run a deal/coupon to recover.")
    if latest and prev and latest.unit_session_pct is not None and avg_cvr_7 and \
            latest.unit_session_pct < avg_cvr_7 and streak < CVR_FALL_DAYS:
        add("listing", "low", f"CVR {latest.unit_session_pct}% below its 7d avg {avg_cvr_7}%",
            "Conversion dipped — keep an eye on it; optimize price/images/reviews if it persists.")

    sev_rank = {"high": 0, "medium": 1, "low": 2}
    recs.sort(key=lambda r: sev_rank.get(r["severity"], 3))
    return recs


def summary(db: Session, start: date, end: date, target_tacos: float = TARGET_TACOS) -> dict:
    """Consolidated tracker for [start, end]: per-day rows (gaps -> '-'),
    day-over-day deltas, 7d rolling lines, alerts, divergence, weekday/weekend,
    B2B split, and the latest-day health score."""
    by_date = {r.date: r for r in db.query(md.FactDaily)
               .filter(md.FactDaily.date >= start, md.FactDaily.date <= end)
               .order_by(md.FactDaily.date).all()}

    # every calendar day in range (so gaps render as "-")
    days = []
    d = start
    while d <= end:
        days.append(d); d += timedelta(days=1)
    rows = [_row_dict(by_date.get(d), d) for d in days]

    # the days with data, in order. Business features key off days that actually
    # carry Business-Report metrics (a PPC-only day has no sessions/CVR/buy box).
    data_days = [d for d in days if d in by_date]
    R = [by_date[d] for d in data_days]                       # any data (incl. ppc-only)
    Rb = [r for r in R if r.ordered_sales is not None or r.unit_session_pct is not None]
    b_days = [r.date for r in Rb]

    def series(attr):
        return [getattr(r, attr) for r in Rb]

    # 7-day rolling averages over business days, attached onto the calendar rows
    def rolling(attr):
        vals = series(attr); out = []
        for i in range(len(vals)):
            out.append(_avg(vals[max(0, i - ROLL_WINDOW + 1): i + 1]))
        return out
    roll = {a: rolling(a) for a in ("ordered_sales", "sessions", "unit_session_pct")}
    roll_by_date = {b_days[i]: {a: roll[a][i] for a in roll} for i in range(len(b_days))}
    for row in rows:
        rd = roll_by_date.get(date.fromisoformat(row["date"]))
        row["sales_roll7"] = rd["ordered_sales"] if rd else None
        row["sessions_roll7"] = rd["sessions"] if rd else None
        row["cvr_roll7"] = rd["unit_session_pct"] if rd else None

    latest = Rb[-1] if Rb else None       # latest day WITH business data
    prev = Rb[-2] if len(Rb) >= 2 else None

    # 1. day-over-day deltas (latest vs prior data day)
    deltas = {}
    if latest:
        for f in ("ordered_sales", "units_ordered", "sessions", "unit_session_pct",
                  "asp", "buy_box_pct"):
            deltas[f] = _delta(getattr(latest, f), getattr(prev, f) if prev else None)

    # trailing 7d averages (excluding the latest day) for spike/swing tests
    tail = Rb[-(ROLL_WINDOW + 1):-1] if len(Rb) > 1 else []
    avg_refund_7 = _avg([r.refund_rate for r in tail])
    avg_asp_7 = _avg([r.asp for r in tail])
    avg_cvr_7 = _avg([r.unit_session_pct for r in tail])

    alerts = []
    # 3. CVR falling N days in a row
    cvr = [r.unit_session_pct for r in Rb]
    streak = 0
    for i in range(len(cvr) - 1, 0, -1):
        if cvr[i] is not None and cvr[i - 1] is not None and cvr[i] < cvr[i - 1]:
            streak += 1
        else:
            break
    if streak >= CVR_FALL_DAYS:
        alerts.append({"type": "cvr_fall", "severity": "high",
                       "msg": f"CVR fell {streak} days in a row (latest {cvr[-1]}%)"})
    # 4. buy box < 90 on the latest day
    if latest and latest.buy_box_pct is not None and latest.buy_box_pct < BUYBOX_MIN:
        alerts.append({"type": "buy_box", "severity": "high",
                       "msg": f"Buy Box {latest.buy_box_pct}% < {BUYBOX_MIN}% on {latest.date.isoformat()}"})
    # 5. refund spike: latest > 2x trailing-7d avg AND latest > 3%
    if latest and latest.refund_rate is not None and avg_refund_7:
        if latest.refund_rate > REFUND_SPIKE_MULT * avg_refund_7 and latest.refund_rate > REFUND_FLOOR:
            alerts.append({"type": "refund_spike", "severity": "high",
                           "msg": f"Refund rate {latest.refund_rate}% vs 7d avg {avg_refund_7}% "
                                  f"(>{REFUND_SPIKE_MULT}x and >{REFUND_FLOOR}%)"})
    # 7. ASP swing ±15% vs trailing 7d avg
    if latest and latest.asp is not None and avg_asp_7:
        swing = (latest.asp - avg_asp_7) / avg_asp_7
        if abs(swing) > ASP_SWING:
            alerts.append({"type": "asp_swing", "severity": "medium",
                           "msg": f"ASP ${latest.asp} is {swing*100:+.0f}% vs 7d avg ${avg_asp_7} "
                                  "(silent promo / coupon?)"})

    # 6. traffic-vs-conversion divergence over the last DIVERGE_WINDOW data days
    win = Rb[-DIVERGE_WINDOW:]
    sess_up = _slope_up([r.sessions for r in win])
    cvr_up = _slope_up([r.unit_session_pct for r in win])
    divergence = bool(sess_up is True and cvr_up is False)
    if divergence:
        alerts.append({"type": "divergence", "severity": "high",
                       "msg": "Sessions trending UP while CVR trends DOWN — traffic isn't converting"})

    # 8. weekday vs weekend averages over the range
    wk = {"weekday": {"sales": [], "cvr": []}, "weekend": {"sales": [], "cvr": []}}
    for r in Rb:
        b = wk["weekend" if r.date.weekday() >= 5 else "weekday"]
        b["sales"].append(r.ordered_sales); b["cvr"].append(r.unit_session_pct)
    weekday_weekend = {k: {"avg_sales": _avg(v["sales"]), "avg_cvr": _avg(v["cvr"]),
                           "days": len([x for x in v["sales"] if x is not None])}
                       for k, v in wk.items()}

    # 9. B2B vs B2C split (totals over the range; CVR as an average)
    def _sum(attr):
        xs = [getattr(r, attr) for r in Rb if getattr(r, attr) is not None]
        return round(sum(xs), 2) if xs else None
    b2b = {
        "sales": {"total": _sum("ordered_sales"), "b2b": _sum("ordered_sales_b2b")},
        "units": {"total": _sum("units_ordered"), "b2b": _sum("units_ordered_b2b")},
        "cvr": {"total": _avg(series("unit_session_pct")), "b2b": _avg(series("unit_session_pct_b2b"))},
    }

    # 10. health score (0-100) for the latest day — see weighting in comments
    health = _health(latest, avg_cvr_7, prev, avg_refund_7)

    ov = _overview(db, rows, start, end, target_tacos)
    recommendations = _recommendations(ov, latest, prev, streak, divergence,
                                       avg_cvr_7, avg_refund_7, avg_asp_7, sess_up)

    return {
        "range": {"start": start.isoformat(), "end": end.isoformat()},
        "data_days": len(R), "missing_days": len(days) - len(R),
        "latest_date": latest.date.isoformat() if latest else None,
        "overview": ov, "recommendations": recommendations,
        "rows": rows, "deltas": deltas, "alerts": alerts,
        "divergence": divergence, "weekday_weekend": weekday_weekend,
        "b2b": b2b, "health": health,
        "thresholds": {"cvr_fall_days": CVR_FALL_DAYS, "buy_box_min": BUYBOX_MIN,
                       "refund_mult": REFUND_SPIKE_MULT, "refund_floor": REFUND_FLOOR,
                       "asp_swing_pct": ASP_SWING * 100, "roll_window": ROLL_WINDOW},
    }


EXPORT_COLS = ["Date", "Total Sales", "Units Ordered", "Ad Spend", "Ad Sales",
               "Orders (PPC)", "Clicks", "Impressions"]


def export_xlsx(db: Session, start: date, end: date,
                target_tacos: float = TARGET_TACOS) -> bytes:
    """Client-ready tracker workbook with native Excel charts: Overview (range
    KPIs, health, alerts, recommendations, weekday/weekend bar, B2B pie) + the
    Daily Tracker sheet (one row per calendar day, sales/spend line chart)."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill

    s = summary(db, start, end, target_tacos=target_tacos)
    ov = s["overview"] or {}

    TITLE = Font(bold=True, size=14)
    H = Font(bold=True, color="1F2329")
    HFILL = PatternFill("solid", fgColor="FCD535")
    SUB = Font(bold=True, size=11)

    def head(ws, row, labels, widths=None):
        for i, lab in enumerate(labels, start=1):
            c = ws.cell(row=row, column=i, value=lab)
            c.font = H; c.fill = HFILL
            if widths and i <= len(widths) and widths[i - 1]:
                ws.column_dimensions[c.column_letter].width = widths[i - 1]
        return row + 1

    wb = Workbook()

    # ---- Overview -------------------------------------------------------------
    ws = wb.active
    ws.title = "Overview"
    ws.cell(row=1, column=1, value="Daily SALES & PPC Tracker").font = TITLE
    ws.cell(row=2, column=1,
            value=f"{s['range']['start']} → {s['range']['end']} · {s['data_days']} days of data"
                  + (f" · {s['missing_days']} missing" if s["missing_days"] else "")
                  + (f" · health {s['health']['score']}/100" if s.get("health", {}).get("score") is not None else ""))
    r = 4
    for lab, val in [("Total sales", ov.get("total_sales")), ("Units ordered", ov.get("units_ordered")),
                     ("Ad spend", ov.get("ad_spend")), ("Ad sales", ov.get("ad_sales")),
                     ("Orders (PPC)", ov.get("ppc_orders")), ("ACoS %", ov.get("acos")),
                     ("ROAS", ov.get("roas")), ("CPC", ov.get("cpc")), ("CTR %", ov.get("ctr")),
                     ("CVR %", ov.get("cvr")),
                     ("TACOS %", ov.get("tacos")), ("TACOS target %", ov.get("target_tacos")),
                     ("TACOS at/below target", "yes" if ov.get("tacos_below_target") else "no")]:
        ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16

    # weekday vs weekend + bar chart
    wwk = s.get("weekday_weekend") or {}
    if any((v or {}).get("days") for v in wwk.values()):
        r += 1
        ws.cell(row=r, column=1, value="Weekday vs weekend").font = SUB
        r += 1
        hdr = r
        r = head(ws, r, ["", "Avg sales", "Avg CVR %", "Days"], widths=[12, 11, 11, 8])
        first = r
        for k in ("weekday", "weekend"):
            v = wwk.get(k) or {}
            ws.cell(row=r, column=1, value=k)
            ws.cell(row=r, column=2, value=v.get("avg_sales"))
            ws.cell(row=r, column=3, value=v.get("avg_cvr"))
            ws.cell(row=r, column=4, value=v.get("days"))
            r += 1
        ch = BarChart(); ch.type = "col"; ch.title = "Avg daily sales — weekday vs weekend"
        ch.height = 7; ch.width = 11
        ch.add_data(Reference(ws, min_col=2, min_row=hdr, max_row=r - 1), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
        ws.add_chart(ch, "F4")

    # B2B split pie
    b2b = (s.get("b2b") or {}).get("sales") or {}
    if b2b.get("total") and b2b.get("b2b") is not None:
        base = r + 2
        ws.cell(row=base - 1, column=1, value="B2B vs B2C sales").font = SUB
        ws.cell(row=base, column=1, value="B2B"); ws.cell(row=base, column=2, value=b2b["b2b"])
        ws.cell(row=base + 1, column=1, value="B2C")
        ws.cell(row=base + 1, column=2, value=round((b2b["total"] or 0) - (b2b["b2b"] or 0), 2))
        pie = PieChart(); pie.title = "B2B share of sales"; pie.height = 7; pie.width = 10
        pie.add_data(Reference(ws, min_col=2, min_row=base, max_row=base + 1))
        pie.set_categories(Reference(ws, min_col=1, min_row=base, max_row=base + 1))
        ws.add_chart(pie, "F19")
        r = base + 2

    # alerts + recommendations
    if s.get("alerts"):
        r += 1
        ws.cell(row=r, column=1, value=f"Alerts · {len(s['alerts'])}").font = SUB
        r += 1
        r = head(ws, r, ["Severity", "Alert"], widths=[10, 90])
        for a in s["alerts"]:
            ws.cell(row=r, column=1, value=a["severity"]); ws.cell(row=r, column=2, value=a["msg"]); r += 1
    if s.get("recommendations"):
        r += 1
        ws.cell(row=r, column=1, value=f"Recommendations · {len(s['recommendations'])}").font = SUB
        r += 1
        r = head(ws, r, ["Category", "Severity", "Title", "Action"], widths=[10, 10, 44, 70])
        for x in s["recommendations"]:
            for i, v in enumerate([x["category"], x["severity"], x["title"], x["action"]], start=1):
                ws.cell(row=r, column=i, value=v)
            r += 1

    # ---- Daily Tracker ----------------------------------------------------------
    by_date = {r_.date: r_ for r_ in db.query(md.FactDaily)
               .filter(md.FactDaily.date >= start, md.FactDaily.date <= end)
               .order_by(md.FactDaily.date).all()}
    ws = wb.create_sheet("Daily Tracker")
    hdr = 1
    r = head(ws, hdr, EXPORT_COLS, widths=[14, 11, 12, 10, 10, 11, 9, 12])
    first = r
    d = start
    while d <= end:
        rec = by_date.get(d)
        ws.cell(row=r, column=1, value=f"{d.strftime('%B')} {d.day}")
        ws.cell(row=r, column=2, value=round(rec.ordered_sales, 2) if rec and rec.ordered_sales is not None else 0.0)
        ws.cell(row=r, column=3, value=(rec.units_ordered or 0) if rec else 0)
        ws.cell(row=r, column=4, value=round(rec.ppc_spend, 2) if rec and rec.ppc_spend is not None else 0.0)
        ws.cell(row=r, column=5, value=round(rec.ppc_sales, 2) if rec and rec.ppc_sales is not None else 0.0)
        ws.cell(row=r, column=6, value=(rec.ppc_orders or 0) if rec else 0)
        ws.cell(row=r, column=7, value=(rec.ppc_clicks or 0) if rec else 0)
        ws.cell(row=r, column=8, value=(rec.ppc_impressions or 0) if rec else 0)
        d += timedelta(days=1)
        r += 1
    ln = LineChart(); ln.title = "Total sales · ad spend · ad sales by day"
    ln.height = 9; ln.width = 24
    ln.add_data(Reference(ws, min_col=2, max_col=5, min_row=hdr, max_row=r - 1), titles_from_data=True)
    # drop the Units series (col 3) — keep sales/spend/ad-sales money lines together
    del ln.series[1]
    ln.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
    ws.add_chart(ln, "J2")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _health(latest, avg_cvr_7, prev, avg_refund_7) -> dict:
    """Composite 0-100 for the latest day. Baseline 50, then add/subtract signals:

        CVR vs 7d avg      ±20   (conversion is the headline metric)
        sessions vs prior  ±10   (traffic momentum)
        buy box ≥ 90%      ±15   (you can't sell what you don't own the box on)
        refund rate vs 7d  ±15   (quality / returns)

    Max swing ±60 around 50 -> clamped to [0, 100]. Green ≥70, amber 40-69, red <40.
    """
    if not latest:
        return {"score": None, "band": "na", "parts": {}}
    score = 50.0
    parts = {}
    if latest.unit_session_pct is not None and avg_cvr_7:
        up = latest.unit_session_pct >= avg_cvr_7
        score += 20 if up else -20; parts["cvr"] = 20 if up else -20
    if prev and latest.sessions is not None and prev.sessions is not None:
        up = latest.sessions >= prev.sessions
        score += 10 if up else -10; parts["sessions"] = 10 if up else -10
    if latest.buy_box_pct is not None:
        ok = latest.buy_box_pct >= BUYBOX_MIN
        score += 15 if ok else -15; parts["buy_box"] = 15 if ok else -15
    if latest.refund_rate is not None and avg_refund_7 is not None:
        down = latest.refund_rate <= avg_refund_7
        score += 15 if down else -15; parts["refunds"] = 15 if down else -15
    score = max(0, min(100, round(score)))
    band = "green" if score >= 70 else "amber" if score >= 40 else "red"
    return {"score": score, "band": band, "parts": parts}

"""Product Benchmark catalog — parse Amazon Category Listings Reports into a
store-level product catalog (title / description / bullets / price / images /
variations), used as the product base for mapping features.

Scope: STORE-level (like Product Costs) — one catalog per store, shared by every
audit/cadence in it, never visible from another store (store dirs are per-user
namespaced). Persisted as `<store>/_catalog.json` via database.get_store_catalog /
set_store_catalog; deleted with the store.

Amazon splits the Category Listings Report per category (one .xlsm per product
type), so uploads MERGE by SKU — each file upserts its products into the catalog
instead of replacing it.

File shape (Seller Central > Reports > Category Listings Report): sheet
"Template"; a header row of display names, then a row of attribute keys like
`item_name[marketplace_id=...][language_tag=en_US]#1.value`, then one example
row (SKU "ABC123"), then real data. Column POSITIONS drift per category (800+
columns), so parsing keys off the attribute-key row is the only stable approach.
"""
from __future__ import annotations

import csv
import json
import os
import re
from datetime import datetime

from .. import database as dbmod
from . import benchmark as bn


# ---- store-scoped catalog file ----------------------------------------------
# One _catalog.json per STORE (store dirs are per-user namespaced, so two users'
# "zvalves" never share). Removed with the store dir on store delete.
def _store_path(store_id: str) -> str:
    return os.path.join(dbmod._store_dir(store_id), "_catalog.json")


def read_catalog(store_id: str | None) -> dict:
    if not store_id:
        return {}
    p = _store_path(store_id)
    if os.path.exists(p):
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return {}


def write_catalog(store_id: str, data: dict) -> None:
    with open(_store_path(store_id), "w", encoding="utf-8") as f:
        json.dump(data, f, indent=1, ensure_ascii=False)


# ---- per-SKU COGS overrides (store-level, % of selling price) -----------------
# Break-even ACoS / profit-per-unit derive from the selling price minus referral
# fee and COGS. COGS defaults to 40% of the selling price; each product can
# override it (entered as a decimal `0.35` or a percent `35` / `35%`).
DEFAULT_COGS_PCT = 0.40


def _cogs_path(store_id: str) -> str:
    return os.path.join(dbmod._store_dir(store_id), "_cogs.json")


def read_cogs(store_id: str | None) -> dict:
    """sku -> COGS fraction override (0.35 = 35% of selling price)."""
    if not store_id:
        return {}
    p = _cogs_path(store_id)
    if os.path.exists(p):
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    return {}


def write_cogs(store_id: str, data: dict) -> None:
    with open(_cogs_path(store_id), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)


def norm_sku(s) -> str:
    """SKU normalized for cross-report matching — Amazon exports drift on case,
    spaces and dashes ('PI 100' vs 'pi-100')."""
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def fees_by_sku(store_id: str | None) -> dict:
    """normalized SKU -> real per-unit fees from the store's Transactions ledger
    (Payments Date Range report): referral_pct = |selling fees| ÷ product sales,
    fba_fee = |FBA fees| ÷ units — actual observed fees, per SKU. These make the
    break-even ACoS price-DEPENDENT (a fixed $ FBA fee weighs more on a cheap
    product), instead of the constant 45% the percent-only defaults produce.

    The fee-report upload (Selling economics / FBA Fee Preview) WINS over the
    ledger for BOTH fees: its base fulfillment fee per unit is the current
    published per-unit fee (the ledger average smears fee changes and multi-unit
    orders across a period), and its referral fee per unit ÷ price is the
    marginal rate an incremental ad sale pays (the ledger's |selling fees| ÷
    product sales is refund- and promo-netted, so it reads low). SKUs the report
    covers but the ledger doesn't (no sales yet) still get fees.
    `fba_source` = "report" | "ledger"."""
    from . import fbafees as fba
    from . import transactions as txn
    agg: dict[str, dict] = {}
    for r in (txn.read_txn(store_id) or {}).get("rows", []):
        if (r.get("type") or "").lower() != "order" or not r.get("sku"):
            continue
        a = agg.setdefault(norm_sku(r["sku"]), {"sales": 0.0, "sell": 0.0, "fba": 0.0, "units": 0})
        a["sales"] += r.get("product_sales") or 0
        a["sell"] += r.get("selling_fees") or 0
        a["fba"] += r.get("fba_fees") or 0
        a["units"] += r.get("quantity") or 0
    out = {}
    for k, a in agg.items():
        ref = round(abs(a["sell"]) / a["sales"], 4) if a["sales"] > 0 else None
        fee = round(abs(a["fba"]) / a["units"], 2) if a["units"] > 0 else None
        if ref is not None or fee is not None:
            out[k] = {"referral_pct": ref, "fba_fee": fee,
                      "fba_source": "ledger" if fee is not None else None}
    def _apply(key: str, r: dict) -> None:
        row = out.setdefault(key, {"referral_pct": None, "fba_fee": None, "fba_source": None})
        if r.get("fulfillment_fee") is not None:
            row["fba_fee"] = r["fulfillment_fee"]
            row["fba_source"] = "report"
            row["size_tier"] = r.get("size_tier")
        if r.get("total_fee") is not None:
            # exact per-unit total (fulfillment + referral) from the report —
            # charged as-is by break-even
            row["total_fee"] = r["total_fee"]
            row["referral_fee"] = r.get("referral_fee")
            # the price that total was quoted at, so break-even can tell whether
            # it still applies to the price it's dividing by
            row["fee_price"] = r.get("price")
        if r.get("referral_source"):
            row["referral_source"] = r["referral_source"]   # "referral_preview"
        if r.get("referral_pct") is not None:
            # the report's referral fee per unit ÷ price is the published MARGINAL
            # rate; the ledger's |selling fees| ÷ product sales is refund- and
            # promo-netted, so the report wins where it has a value
            row["referral_pct"] = r["referral_pct"]
            row["referral_source"] = "report"

    for k, r in fba.by_sku(store_id).items():           # SKU-keyed rows (Fee Preview)
        _apply(k, r)
    # ASIN-keyed rows (Selling economics leaves MSKU empty) reach a SKU through
    # the catalog's SKU -> ASIN map; parents are already excluded by by_asin()
    fee_asins = fba.by_asin(store_id)
    if fee_asins:
        for sku, p in (read_catalog(store_id) or {}).get("products", {}).items():
            r = fee_asins.get((p.get("asin") or "").upper())
            if r is not None and not r.get("sku"):
                _apply(norm_sku(sku), r)
    return out


def be_by_sku(store_id: str | None, econ: dict) -> dict:
    """normalized SKU -> break-even ACoS, computed per catalog listing from its
    selling price + per-SKU COGS override (default 40%) + the SKU's real fees
    from the Transactions ledger. Lets Product Ads / the audit match a product
    by SKU when its ASIN doesn't line up across reports."""
    data = read_catalog(store_id)
    if not data:
        return {}
    cogs = read_cogs(store_id)
    fees = fees_by_sku(store_id)
    out = {}
    for sku, p in data.get("products", {}).items():
        if p.get("parentage") == "parent":
            continue
        b = be_metrics(p, None, econ, cogs.get(sku), fees.get(norm_sku(sku)))
        if b and b.get("break_even_acos") is not None:
            out[norm_sku(sku)] = b["break_even_acos"]
    return out


def parse_cogs(v) -> float | None:
    """'0.35' -> 0.35 · '35' / '35%' -> 0.35 · ''/None -> None (clear override).
    Values >= 1 read as percent. Clamped to (0, 0.95]."""
    s = str(v if v is not None else "").strip().rstrip("%").strip()
    if not s:
        return None
    x = float(s)
    if x >= 1:
        x = x / 100.0
    if x <= 0:
        return None
    return round(min(x, 0.95), 4)

# attribute-key row marker: present in every Category Listings Report
_SKU_KEY = "contribution_sku#1.value"
_EXAMPLE_SKU = "ABC123"          # Amazon's built-in example row

_num_re = re.compile(r"-?\d+(?:\.\d+)?")


def _base(key: str) -> str:
    """`item_name[marketplace_id=..][language_tag=..]#1.value` -> `item_name`."""
    return key.split("[")[0].split("#")[0]


def _inst(key: str) -> int:
    """Instance number: `bullet_point[...]#3.value` -> 3 (default 1)."""
    m = re.search(r"#(\d+)\.", key)
    return int(m.group(1)) if m else 1


def _s(v) -> str:
    return str(v).strip() if v is not None else ""


def _price(v):
    """'320.0' / '$19.95' / 19.95 -> float, else None."""
    if v is None or v == "":
        return None
    m = _num_re.search(str(v).replace(",", ""))
    return float(m.group()) if m else None


def _rows_from_file(path: str):
    """Yield raw rows (lists of cell values) from the Template sheet (or CSV)."""
    if path.lower().endswith(".csv"):
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
            yield from csv.reader(f)
        return
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = None
        for name in wb.sheetnames:
            if name.strip().lower() == "template":
                ws = wb[name]
                break
        sheets = [ws] if ws is not None else [wb[n] for n in wb.sheetnames]
        for cand in sheets:
            probe = []
            for i, row in enumerate(cand.iter_rows(values_only=True)):
                probe.append(row)
                if i >= 9:
                    break
            if any(any(_s(c) == _SKU_KEY for c in r) for r in probe):
                yield from cand.iter_rows(values_only=True)
                return
        raise ValueError(
            "This doesn't look like an Amazon Category Listings Report — no sheet "
            "with listing attribute keys (contribution_sku…) found. Download it from "
            "Seller Central > Reports > Category Listings Report.")
    finally:
        wb.close()


def parse_clr(path: str) -> list[dict]:
    """Parse one Category Listings Report into product dicts (one per SKU row)."""
    rows = _rows_from_file(path)
    keys = None
    for row in rows:
        if any(_s(c) == _SKU_KEY for c in row):
            keys = [_s(c) for c in row]
            break
    if keys is None:
        raise ValueError(
            "This doesn't look like an Amazon Category Listings Report — the "
            "attribute-key header row (contribution_sku…) is missing.")

    # map every column we care about once; positions drift per category
    col = {}                      # simple single-value fields: name -> index
    bullets, keywords, other_imgs = {}, {}, {}
    for i, k in enumerate(keys):
        if not k:
            continue
        b = _base(k)
        if k == _SKU_KEY:
            col["sku"] = i
        elif b == "item_name":
            col.setdefault("title", i)
        elif k == "::title":
            col.setdefault("fallback_title", i)
        elif k == "::listing_status":
            col["status"] = i
        elif b == "brand":
            col.setdefault("brand", i)
        elif b == "product_type":
            col.setdefault("product_type", i)
        elif k == "amzn1.volt.ca.product_id_type":
            col["id_type"] = i
        elif k == "amzn1.volt.ca.product_id_value":
            col["id_value"] = i
        elif b == "parentage_level":
            col.setdefault("parentage", i)
        elif b == "child_parent_sku_relationship" and ".parent_sku" in k:
            col.setdefault("parent_sku", i)
        elif b == "variation_theme" and k.endswith(".name"):
            col.setdefault("variation_theme", i)
        elif b == "color":
            # prefer the free-text .value over .standardized_value
            if k.endswith("#1.value") or "color" not in col:
                col["color"] = i
        elif b == "size":
            col.setdefault("size", i)
        elif b == "product_description":
            col.setdefault("description", i)
        elif b == "main_product_image_locator":
            col.setdefault("main_image", i)
        elif b == "swatch_product_image_locator":
            col.setdefault("swatch_image", i)
        elif b == "list_price":
            col.setdefault("list_price", i)
        elif b == "purchasable_offer" and "audience=ALL" in k:
            if ".our_price" in k:
                col.setdefault("price", i)
            elif ".discounted_price" in k and ".value" in k:
                col.setdefault("sale_price", i)
        elif b == "bullet_point":
            bullets.setdefault(_inst(k), i)
        elif b == "generic_keyword":
            keywords.setdefault(_inst(k), i)
        else:
            m = re.match(r"other_product_image_locator_(\d+)$", b)
            if m:
                other_imgs.setdefault(int(m.group(1)), i)

    def cell(r, name):
        i = col.get(name)
        return _s(r[i]) if i is not None and i < len(r) else ""

    products = []
    for r in rows:                      # continues AFTER the key row
        sku = cell(r, "sku")
        title = cell(r, "title") or cell(r, "fallback_title")
        if not sku and not title:
            continue                    # blank filler row
        if sku.upper() == _EXAMPLE_SKU:
            continue                    # Amazon's example row
        id_type = cell(r, "id_type").upper()
        id_value = cell(r, "id_value")
        parentage = cell(r, "parentage").lower()
        products.append({
            "sku": sku,
            "asin": id_value if id_type == "ASIN" else "",
            "id_type": id_type,
            "id_value": id_value,
            "title": title,
            "brand": cell(r, "brand"),
            "product_type": cell(r, "product_type"),
            "status": cell(r, "status"),
            "parentage": parentage if parentage in ("parent", "child") else "",
            "parent_sku": cell(r, "parent_sku"),
            "variation_theme": cell(r, "variation_theme"),
            "color": cell(r, "color"),
            "size": cell(r, "size"),
            "price": _price(cell(r, "price")),
            "sale_price": _price(cell(r, "sale_price")),
            "list_price": _price(cell(r, "list_price")),
            "main_image": cell(r, "main_image"),
            "swatch_image": cell(r, "swatch_image"),
            "images": [v for n in sorted(other_imgs)
                       if (v := (_s(r[other_imgs[n]]) if other_imgs[n] < len(r) else ""))],
            "description": cell(r, "description"),
            "bullets": [v for n in sorted(bullets)
                        if (v := (_s(r[bullets[n]]) if bullets[n] < len(r) else ""))],
            "search_terms": [v for n in sorted(keywords)
                             if (v := (_s(r[keywords[n]]) if keywords[n] < len(r) else ""))],
        })
    if not products:
        raise ValueError("No product rows found in that Category Listings Report.")
    return products


def merge(data: dict, products: list[dict], filename: str) -> tuple[dict, int, int]:
    """Upsert parsed products into the store catalog blob (keyed by SKU)."""
    cat = data if isinstance(data, dict) else {}
    items = cat.setdefault("products", {})
    added = updated = 0
    for p in products:
        key = p["sku"] or p["id_value"] or p["title"]
        if key in items:
            updated += 1
        else:
            added += 1
        items[key] = p
    now = datetime.now().isoformat(timespec="seconds")
    files = [f for f in cat.get("files", []) if f.get("name") != filename]
    files.append({"name": filename, "rows": len(products), "uploaded": now})
    cat["files"] = files
    cat["updated"] = now
    return cat, added, updated


def _light(p: dict) -> dict:
    """Table-row projection: everything but the heavy copy fields."""
    return {k: p.get(k) for k in (
        "sku", "asin", "title", "brand", "product_type", "status", "parentage",
        "parent_sku", "variation_theme", "color", "size",
        "price", "sale_price", "list_price", "main_image")} | {
        "image_count": (1 if p.get("main_image") else 0) + len(p.get("images") or []),
        "bullet_count": len(p.get("bullets") or []),
        "desc_chars": len(p.get("description") or ""),
        "search_terms": bool(p.get("search_terms")),
        # high+medium SEO issue count (None for parents — not live listings)
        "seo_issues": (seo_check(p) or {}).get("issues"),
    }


def overview(data: dict) -> dict:
    """Catalog summary: light product rows + account stats + upload history."""
    items = (data or {}).get("products", {})
    prods = sorted(items.values(), key=lambda p: ((p.get("title") or "").lower(), p.get("sku") or ""))
    parents = [p for p in prods if p.get("parentage") == "parent"]
    children = [p for p in prods if p.get("parentage") == "child"]
    priced = [p for p in prods if p.get("price") is not None]
    return {
        "products": [_light(p) for p in prods],
        "stats": {
            "total": len(prods),
            "parents": len(parents),
            "children": len(children),
            "standalone": len(prods) - len(parents) - len(children),
            "brands": len({p.get("brand") for p in prods if p.get("brand")}),
            "priced": len(priced),
            "missing_image": sum(1 for p in prods
                                 if p.get("parentage") != "parent" and not p.get("main_image")),
            "missing_desc": sum(1 for p in prods
                                if p.get("parentage") != "parent" and not p.get("description")),
            "listing_issues": sum(1 for p in prods
                                  if ((seo_check(p) or {}).get("issues") or 0) > 0),
        },
        "files": (data or {}).get("files", []),
        "updated": (data or {}).get("updated"),
    }


# ---- catalog-level SEO check --------------------------------------------------
# Listing-quality recommendations computed from the Category Listings Report
# fields alone — no Product Optimization project needed. Same rules the tracker's
# SEO engine applies to pasted copy (lengths, 249-byte backend field, wasted
# words), plus catalog-only checks (images). The "Perform Listing Audit" hand-off
# then adds the keyword-based recommendations on top.
_ST_MAX_BYTES = 249
_TITLE_MAX = 200
_TITLE_THIN = 80
_DESC_THIN = 400
_SEV_ORDER = {"high": 0, "medium": 1, "low": 2}
_word_re = re.compile(r"[^a-z0-9]+")


def _words(s: str) -> list[str]:
    return [w for w in _word_re.sub(" ", (s or "").lower()).split() if w]


def seo_check(p: dict) -> dict | None:
    """SEO recommendations for one catalog product. None for parent rows —
    they're variation containers, not live listings."""
    if p.get("parentage") == "parent":
        return None
    recs = []

    def rec(severity, area, title, action, words=None):
        recs.append({"severity": severity, "area": area, "title": title,
                     "action": action, "words": words or []})

    title = (p.get("title") or "").strip()
    if not title:
        rec("high", "title", "No title", "Every listing needs one — add it in Seller Central.")
    elif len(title) > _TITLE_MAX:
        rec("medium", "title", f"Title is {len(title)} chars (over {_TITLE_MAX})",
            "Most categories cap the title at 200 chars — Amazon may suppress it. Trim it.")
    elif len(title) < _TITLE_THIN:
        rec("low", "title", f"Title is only {len(title)} chars",
            f"There's room up to ~{_TITLE_MAX} — add your top keyword phrases.")

    bullets = [b for b in (p.get("bullets") or []) if (b or "").strip()]
    if len(bullets) < 5:
        rec("medium", "bullets", f"Only {len(bullets)} bullet point{'s' if len(bullets) != 1 else ''}",
            "Amazon gives you 5 — each is indexed. Fill the missing ones.")

    desc = (p.get("description") or "").strip()
    if not desc:
        rec("medium", "description", "No product description",
            "The description is indexed and converts — write one.")
    elif len(desc) < _DESC_THIN:
        rec("low", "description", f"Description is only {len(desc)} chars",
            "Thin descriptions rank and convert worse — expand it.")

    st = " ".join(s for s in (p.get("search_terms") or []) if (s or "").strip()).strip()
    if not st:
        rec("high", "search_terms", "No backend search terms",
            "The generic-keywords field is free indexing space (249 bytes) — fill it.")
    else:
        st_bytes = len(st.encode("utf-8"))
        if st_bytes > _ST_MAX_BYTES:
            rec("high", "search_terms", f"Backend search terms are {st_bytes} bytes (over {_ST_MAX_BYTES})",
                "Amazon ignores the whole field beyond 249 bytes — cut it below the limit.")
        visible = set(_words(title)) | set(_words(" ".join(bullets))) | set(_words(desc))
        wasted = [w for w in dict.fromkeys(_words(st)) if w in visible]
        if wasted:
            rec("medium", "search_terms",
                f"{len(wasted)} backend word{'s' if len(wasted) != 1 else ''} already indexed by the visible copy",
                "Words in the title/bullets/description are already indexed — spend the field's bytes on words the copy doesn't have.",
                wasted[:10])

    n_img = (1 if p.get("main_image") else 0) + len(p.get("images") or [])
    if not p.get("main_image"):
        rec("high", "images", "No main image", "Amazon suppresses listings without one.")
    elif n_img < 5:
        rec("medium", "images", f"Only {n_img} image{'s' if n_img != 1 else ''}",
            "Aim for 6+ (lifestyle, dimensions, infographic) — images drive CVR.")

    recs.sort(key=lambda r: _SEV_ORDER.get(r["severity"], 9))
    counts = {s: sum(1 for r in recs if r["severity"] == s) for s in ("high", "medium", "low")}
    return {"recommendations": recs, "counts": counts,
            "issues": counts["high"] + counts["medium"]}


# ---- Product Ads + break-even join ------------------------------------------
def be_metrics(p: dict, bench_row: dict | None, econ: dict,
               cogs_pct: float | None = None, fees: dict | None = None) -> dict | None:
    """Break-even block for one catalog product. An uploaded Product Benchmark
    break-even ACoS wins; otherwise derived from the catalog selling price
    (sale price when set, else price, else the benchmark file's price) + fees +
    COGS. Fees: the SKU's REAL referral % and FBA $/unit observed in the
    Transactions ledger when available (fees arg, fees_by_sku), else the econ
    referral default and no FBA fee. COGS = the product's per-SKU override when
    set, else the audit econ default, else 40% of the selling price. None when
    no price and no uploaded value — nothing to compute on."""
    from types import SimpleNamespace
    ref = (fees or {}).get("referral_pct")
    if ref is None:
        ref = econ.get("default_referral_pct", 0.15) or 0.15
    fba = (fees or {}).get("fba_fee") or 0.0
    total = (fees or {}).get("total_fee")     # exact per-unit fulfillment + referral
    pct = cogs_pct if cogs_pct is not None else \
        (econ.get("default_cogs_pct") or DEFAULT_COGS_PCT)
    price = p.get("sale_price") or p.get("price") or (bench_row or {}).get("sale_price")
    # The report quoted that total at ITS price. If this listing sells at a
    # different price, the fixed $ referral no longer applies — scale it by the
    # report's RATE instead (referral is a percentage of the actual sale price;
    # only the fulfillment fee is a fixed $).
    fee_price = (fees or {}).get("fee_price")
    if total is not None and ref and price and fee_price and abs(fee_price - price) > 0.01:
        total = None
    uploaded = (bench_row or {}).get("break_even_acos")
    cost = SimpleNamespace(referral_pct=ref, fba_fee=fba, misc_fee=0.0, unit_cost=0.0,
                           total_fee=total)
    cogs, amazon_fee, profit = bn._unit_economics(price, cost, ref, pct)
    be = uploaded if uploaded is not None else bn._derive_be(price, cost, ref, pct)
    if be is None and profit is None:
        return None
    return {"price": price, "cogs": cogs, "amazon_fee": amazon_fee,
            "profit_per_unit": profit,
            "cogs_pct": round(pct, 4), "cogs_custom": cogs_pct is not None,
            # amazon_fee broken out, so the UI can show what the fixed $ per-unit
            # fulfillment fee costs vs the referral fee. total_fee = the report's
            # exact fulfillment + referral per unit (None -> derived from the %).
            "referral_pct": round(ref, 4), "fba_fee": fba,
            # total_fee = what break-even actually charges (== amazon_fee).
            # total_source: "report"  the report's exact quoted total
            #               "scaled"  report fees, referral re-priced to this listing
            #               "derived" no fee report — referral % + ledger/no FBA fee
            "total_fee": amazon_fee,
            "total_source": ("report" if total is not None
                             else "scaled" if ((fees or {}).get("fba_source") == "report"
                                               or (fees or {}).get("referral_source") == "report")
                             else "derived"),
            "referral_fee": (round(price * ref, 2) if price else None) if total is None
                            else (fees or {}).get("referral_fee"),
            "fba_source": (fees or {}).get("fba_source"),
            "size_tier": (fees or {}).get("size_tier"),
            "fees_real": fees is not None,     # fees observed in the ledger / Fee Preview
            "break_even_acos": be, "break_even_roas": round(1 / be, 2) if be else None,
            "source": "uploaded" if uploaded is not None else "derived"}


def _be_status(ads: dict | None, be: dict | None):
    """profitable / bleeding vs break-even. Spend with zero sales = bleeding even
    without a break-even (ACoS is undefined-worst). None when nothing to judge."""
    if not ads or not ads.get("spend"):
        return None
    if not ads.get("sales"):
        return "bleeding"
    b = (be or {}).get("break_even_acos")
    if b is None or ads.get("acos") is None:
        return None
    return "bleeding" if ads["acos"] > b else "profitable"


def enrich(view: dict, ads: dict | None, bench: dict, econ: dict,
           cogs: dict | None = None, fees: dict | None = None) -> dict:
    """Join the catalog overview to Product Ads + break-even economics.

    ads = productads.by_asin() output ({"asins": {...}, "campaign_types": {...}})
    or None when no Product Ads upload exists in the selected audit; bench = the
    store's _benchmark.json rows; econ = project econ defaults; cogs = per-SKU
    COGS-fraction overrides (read_cogs). Each product row gains `ads` (campaign
    counts by targeting kind + spend/sales/ACoS/orders), `be` (break-even block
    incl. COGS) and `be_status`; stats gain advertised / campaign /
    over-break-even counts."""
    per = (ads or {}).get("asins", {})
    cogs = cogs or {}
    fees = fees or {}
    advertised = over = under = 0
    sum_spend = sum_sales = 0.0
    for p in view.get("products", []):
        a = per.get(p.get("asin"))
        b = be_metrics(p, bench.get(p.get("asin")), econ, cogs.get(p.get("sku")),
                       fees.get(norm_sku(p.get("sku"))))
        p["ads"] = a
        p["be"] = b
        p["be_status"] = _be_status(a, b)
        if a:
            advertised += 1
            sum_spend += a.get("spend") or 0
            sum_sales += a.get("sales") or 0
        if p["be_status"] == "bleeding":
            over += 1
        elif p["be_status"] == "profitable":
            under += 1
    view["ads_connected"] = ads is not None
    view["stats"].update({
        "advertised": advertised,
        "campaigns": (ads or {}).get("campaign_types", {}).get("total", 0),
        "over_be": over,
        "under_be": under,
        # Average ACoS = Σ ad spend ÷ Σ ad sales over the advertised products'
        # campaigns (aggregate ratio, per the reporting spec)
        "avg_acos": round(sum_spend / sum_sales, 4) if sum_sales else None,
    })
    return view


# ---- exec report (xlsx with native Excel charts) -----------------------------
def report_xlsx(data: dict, view: dict) -> bytes:
    """Client-ready Product Benchmark workbook: Overview (catalog stats,
    composition pie, listing-issues-by-area bar) + Products sheet (one row per
    product with SEO issues + the Product Ads / break-even join, ad-spend or
    issue bar chart). `data` = raw catalog blob (full products, for the SEO
    area aggregate); `view` = the enriched overview the tab renders."""
    import io
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill

    prods = view.get("products") or []
    if not prods:
        raise ValueError("No products in this store's catalog yet — upload a Category "
                         "Listings Report first.")
    stats = view["stats"]

    TITLE = Font(bold=True, size=14)
    H = Font(bold=True, color="1F2329")
    HFILL = PatternFill("solid", fgColor="FCD535")
    SUB = Font(bold=True, size=11)

    def head(ws, row, labels, widths=None):
        for i, lab in enumerate(labels, start=1):
            c = ws.cell(row=row, column=i, value=lab)
            c.font = H; c.fill = HFILL
            if widths and i <= len(widths) and widths[i - 1]:
                ws.column_dimensions[c.column_letter].width = widths[i - 1]
        return row + 1

    wb = Workbook()

    # ---- Overview ---------------------------------------------------------------
    ws = wb.active
    ws.title = "Overview"
    ws.cell(row=1, column=1, value="Product Benchmark report").font = TITLE
    ws.cell(row=2, column=1, value=f"{stats['total']} catalog products · "
            + ("Product Ads joined" if view.get("ads_connected") else "no Product Ads upload in this audit"))
    r = 4
    for lab, val in [("Products", stats["total"]), ("Variation families", stats["parents"]),
                     ("Child variations", stats["children"]), ("Standalone", stats["standalone"]),
                     ("Priced", stats["priced"]), ("Brands", stats["brands"]),
                     ("Missing image", stats["missing_image"]), ("Missing description", stats["missing_desc"]),
                     ("Listing issues (high+med)", stats.get("listing_issues", 0)),
                     ("Advertised (Product Ads)", stats.get("advertised", 0)),
                     ("Avg ACoS (advertised products)", stats.get("avg_acos")),
                     ("Distinct campaigns", stats.get("campaigns", 0)),
                     ("Over break-even", stats.get("over_be", 0)),
                     ("Under break-even", stats.get("under_be", 0))]:
        ws.cell(row=r, column=1, value=lab).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12

    # catalog composition pie
    base = r + 1
    ws.cell(row=base - 1, column=1, value="Catalog composition").font = SUB
    for i, (lab, n) in enumerate([("Parents", stats["parents"]), ("Children", stats["children"]),
                                  ("Standalone", stats["standalone"])]):
        ws.cell(row=base + i, column=1, value=lab); ws.cell(row=base + i, column=2, value=n)
    pie = PieChart(); pie.title = "Catalog composition"; pie.height = 7; pie.width = 11
    pie.add_data(Reference(ws, min_col=2, min_row=base, max_row=base + 2))
    pie.set_categories(Reference(ws, min_col=1, min_row=base, max_row=base + 2))
    ws.add_chart(pie, "D4")
    r = base + 4

    # listing-issues-by-area bar (aggregate seo_check over the full products)
    areas: dict[str, int] = {}
    for p in (data or {}).get("products", {}).values():
        s = seo_check(p)
        if not s:
            continue
        for x in s["recommendations"]:
            if x["severity"] in ("high", "medium"):
                areas[x["area"]] = areas.get(x["area"], 0) + 1
    if areas:
        ws.cell(row=r, column=1, value="Listing issues by area").font = SUB
        r += 1
        hdr = r
        r = head(ws, r, ["Area", "Issues"], widths=[18, 10])
        first = r
        for area, n in sorted(areas.items(), key=lambda x: -x[1]):
            ws.cell(row=r, column=1, value=area.replace("_", " ")); ws.cell(row=r, column=2, value=n); r += 1
        ch = BarChart(); ch.type = "col"; ch.title = "Listing issues by area"; ch.height = 7; ch.width = 12
        ch.add_data(Reference(ws, min_col=2, min_row=hdr, max_row=r - 1), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=r - 1))
        ws.add_chart(ch, "D19")

    # ---- Products ------------------------------------------------------------------
    ws = wb.create_sheet("Products")
    hdr = 1
    r = head(ws, hdr, ["SKU", "ASIN", "Title", "Type", "Parentage", "Price", "Sale price",
                       "Status", "Images", "Bullets", "Desc chars", "SEO issues",
                       "COGS / unit", "Fulfillment / unit", "Referral / unit",
                       "Total fees / unit", "Profit / unit",
                       "Campaigns", "Ad spend", "Ad sales", "ACoS", "Break-even ACoS", "Verdict"],
             widths=[18, 14, 44, 16, 10, 9, 10, 10, 8, 8, 10, 10, 12, 15, 13, 14, 12,
                     10, 10, 10, 9, 15, 11])
    first = r
    def spend(p): return (p.get("ads") or {}).get("spend") or 0
    rows = sorted(prods, key=lambda p: (-spend(p), -(p.get("seo_issues") or 0)))
    for p in rows:
        a, b = p.get("ads") or {}, p.get("be") or {}
        for i, v in enumerate([p.get("sku"), p.get("asin"), p.get("title"), p.get("product_type"),
                               p.get("parentage"), p.get("price"), p.get("sale_price"),
                               p.get("status"), p.get("image_count"), p.get("bullet_count"),
                               p.get("desc_chars"), p.get("seo_issues"),
                               b.get("cogs"), b.get("fba_fee"), b.get("referral_fee"),
                               b.get("amazon_fee"), b.get("profit_per_unit"),
                               a.get("campaigns"), a.get("spend"), a.get("sales"), a.get("acos"),
                               b.get("break_even_acos"), p.get("be_status")], start=1):
            ws.cell(row=r, column=i, value=v)
        r += 1
    top = min(len(rows), 15)
    if view.get("ads_connected") and any(spend(p) for p in rows[:top]):
        ch = BarChart(); ch.type = "col"; ch.title = f"Ad spend — top {top} products"
        ch.height = 9; ch.width = 22
        ch.add_data(Reference(ws, min_col=19, min_row=hdr, max_row=hdr + top), titles_from_data=True)
        ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=hdr + top))
        ws.add_chart(ch, "Y2")
    else:
        worst = sorted(rows, key=lambda p: -(p.get("seo_issues") or 0))[:top]
        if worst and (worst[0].get("seo_issues") or 0) > 0:
            base = r + 2
            ws.cell(row=base - 1, column=1, value="Most listing issues").font = SUB
            for i, p in enumerate(worst):
                ws.cell(row=base + i, column=1, value=p.get("sku"))
                ws.cell(row=base + i, column=2, value=p.get("seo_issues"))
            ch = BarChart(); ch.type = "col"; ch.title = f"SEO issues — top {len(worst)} products"
            ch.height = 9; ch.width = 18
            ch.add_data(Reference(ws, min_col=2, min_row=base, max_row=base + len(worst) - 1))
            ch.set_categories(Reference(ws, min_col=1, min_row=base, max_row=base + len(worst) - 1))
            ws.add_chart(ch, "Y2")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def item(data: dict, sku: str) -> dict:
    """Full detail for one SKU + its variation family (children / siblings)."""
    items = (data or {}).get("products", {})
    p = items.get(sku)
    if p is None:
        raise ValueError(f"SKU {sku!r} is not in the catalog.")
    fam = []
    if p.get("parentage") == "parent":
        fam = [c for c in items.values() if c.get("parent_sku") == p.get("sku")
               and c.get("sku") != p.get("sku")]
    elif p.get("parent_sku"):
        fam = [c for c in items.values()
               if (c.get("parent_sku") == p["parent_sku"] or c.get("sku") == p["parent_sku"])
               and c.get("sku") != p.get("sku")]
    fam.sort(key=lambda c: ((c.get("parentage") != "parent"),
                            (c.get("title") or "").lower()))
    return {**p, "family": [_light(c) for c in fam]}

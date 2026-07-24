"""Per-cadence Clear: each cadence's delete_all wipes its own term table AND the
star schema its uploads fed (via weekly.clear_star_schema), so the generic
optimizer sub-panels clear with it. Daily Watch is watch-only — table only.
Also covers the per-cadence upload-meta helpers (file · rows · uploaded)."""
from datetime import date

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app import models as md
from app.pipeline import dailywatch as dw
from app.pipeline import midmonth as mm
from app.pipeline import weekly as wk


@pytest.fixture()
def db():
    eng = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(eng)
    s = sessionmaker(bind=eng)()
    yield s
    s.close()


def _seed_star(db):
    db.add(md.DimCampaign(campaign_id="c1", name="Camp"))
    db.add(md.DimProduct(asin="B0X"))
    db.commit()


def test_midmonth_delete_all_clears_terms_and_star(db):
    db.add(md.MidMonthTermFact(period=0, campaign_id="c1", ad_group_id="ag1",
                               search_term="widget", clicks=3, spend=1.5))
    _seed_star(db)
    db.commit()
    assert mm.has_data(db)

    out = mm.delete_all(db)
    assert out["terms"] == 1 and out["star_rows"] == 2
    assert not mm.has_data(db)
    assert db.query(md.DimCampaign).count() == 0
    assert db.query(md.DimProduct).count() == 0


def test_weekly_clear_star_schema_counts(db):
    _seed_star(db)
    assert wk.clear_star_schema(db) == 2
    assert wk.clear_star_schema(db) == 0        # idempotent


def test_dailywatch_delete_all(db):
    db.add(md.DailyWatchFact(date=date(2026, 7, 1), campaign_id="c1"))
    db.add(md.DailyWatchFact(date=date(2026, 7, 2), campaign_id="c1"))
    _seed_star(db)          # daily watch never fed the star schema — must survive
    db.commit()
    assert dw.delete_all(db) == 2
    assert db.query(md.DailyWatchFact).count() == 0
    assert db.query(md.DimCampaign).count() == 1


def test_productads_delete_all(db):
    from app.pipeline import productads as pa
    db.add(md.ProductAdFact(asin="B0X", sku="S1"))
    _seed_star(db)          # Product Ads never fed the star schema — must survive
    db.commit()
    assert pa.delete_all(db) == 1
    assert db.query(md.ProductAdFact).count() == 0
    assert db.query(md.DimCampaign).count() == 1


def test_monitoring_delete_all(db):
    from app.pipeline import monitoring as mon
    db.add(md.FactDaily(date=date(2026, 7, 1)))
    db.add(md.FactDaily(date=date(2026, 7, 2)))
    db.add(md.MonthSalesOverride(year=2026, month=6, sales=1000.0))
    db.commit()
    assert mon.delete_all(db) == 2
    assert db.query(md.FactDaily).count() == 0
    # manual month-sales overrides are hand-typed, not upload-derived — kept
    assert db.query(md.MonthSalesOverride).count() == 1


def test_upload_meta_roundtrip(tmp_path, monkeypatch):
    """set/get/clear_upload_meta persist per (base project, cadence) in _meta.json."""
    from app import database as dbmod
    from app.pipeline import cadence as cad
    monkeypatch.setattr(dbmod, "STORES_DIR", str(tmp_path))

    class FakeSession:
        info = {"store": "s1", "project": "p1", "cadence": "mid_month"}

    db = FakeSession()
    meta = cad.set_upload_meta(db, "bulk.xlsx", 42)
    assert meta["file"] == "bulk.xlsx" and meta["rows"] == 42 and meta["uploaded"]
    assert cad.get_upload_meta(db) == meta
    # another cadence of the same project is independent
    db2 = FakeSession(); db2.info = {**db.info, "cadence": "pause_scale"}
    assert cad.get_upload_meta(db2) is None
    # feature namespace (Product Ads) never collides with the cadence's own meta
    pa_meta = cad.set_upload_meta(db, "ads.xlsx", 7, feature="product_ads")
    assert cad.get_upload_meta(db, feature="product_ads") == pa_meta
    assert cad.get_upload_meta(db) == meta
    cad.clear_upload_meta(db, feature="product_ads")
    assert cad.get_upload_meta(db, feature="product_ads") is None
    assert cad.get_upload_meta(db) == meta
    cad.clear_upload_meta(db)
    assert cad.get_upload_meta(db) is None


def test_changelog_clear(db):
    from app.pipeline import changelog as cl
    cl.log(db, [{"action": "bid", "field": "bid", "old_value": "1", "new_value": "2"}], "harvest")
    cl.log(db, [{"action": "negate", "field": "state"}], "weekly")
    assert cl.clear(db, source="harvest") == 1      # scoped clear
    assert cl.clear(db) == 1                        # rest wiped
    assert cl.recent(db) == []


def test_monitoring_export_xlsx_charts(db):
    """Tracker workbook: Overview + Daily Tracker sheets, line chart attached."""
    import io as _io
    import openpyxl
    from app.pipeline import monitoring as mon
    db.add(md.FactDaily(date=date(2026, 7, 1), ordered_sales=100.0, units_ordered=4,
                        ppc_spend=10.0, ppc_sales=40.0, ppc_orders=2, ppc_clicks=20,
                        ppc_impressions=500))
    db.add(md.FactDaily(date=date(2026, 7, 2), ordered_sales=150.0, units_ordered=6,
                        ppc_spend=12.0, ppc_sales=50.0, ppc_orders=3, ppc_clicks=25,
                        ppc_impressions=600))
    db.commit()
    data = mon.export_xlsx(db, date(2026, 7, 1), date(2026, 7, 3), target_tacos=12)
    wb = openpyxl.load_workbook(_io.BytesIO(data))
    assert wb.sheetnames == ["Overview", "Daily Tracker"]
    ov = wb["Overview"]
    assert ov.cell(row=1, column=1).value == "Daily SALES & PPC Tracker"
    dt = wb["Daily Tracker"]
    assert dt.cell(row=1, column=1).value == "Date"
    assert dt.cell(row=2, column=2).value == 100.0     # July 1 total sales
    assert dt.cell(row=4, column=2).value == 0.0       # missing day -> 0
    assert len(dt._charts) == 1


def test_productads_report_xlsx(db):
    """Product Ads workbook: Overview + Products sheets, charts attached."""
    import io as _io
    import openpyxl
    from app.pipeline import productads as pa
    db.add(md.ProductAdFact(asin="B0X", sku="S1", campaign_id="c1", campaign_type="auto",
                            impressions=100, clicks=10, spend=5.0, sales=50.0, orders=2, units=2))
    db.add(md.ProductAdFact(asin="B0Y", sku="S2", campaign_id="c2", campaign_type="keyword",
                            impressions=50, clicks=5, spend=3.0, sales=0.0, orders=0, units=0))
    db.commit()
    data = pa.report_xlsx(db)
    wb = openpyxl.load_workbook(_io.BytesIO(data))
    assert wb.sheetnames == ["Overview", "Products"]
    assert len(wb["Overview"]._charts) == 2 and len(wb["Products"]._charts) == 1
    pr = wb["Products"]
    assert pr.cell(row=2, column=1).value == "B0X"        # top spend first
    assert pr.cell(row=2, column=12).value == 5.0
    # empty table -> friendly error
    from app.pipeline import productads as pa2
    db.query(md.ProductAdFact).delete(); db.commit()
    with pytest.raises(ValueError):
        pa2.report_xlsx(db)

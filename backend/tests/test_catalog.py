"""Product Benchmark catalog: Category Listings Report parse + merge + views."""
import openpyxl
import pytest

from app.pipeline import catalog as cat


def _clr_file(tmp_path, rows, name="clr.xlsx"):
    """Build a minimal Category Listings Report: display row, attribute-key row,
    Amazon's example row, then data rows. Columns deliberately reordered vs the
    real export to prove parsing is key-driven, not position-driven."""
    keys = [
        "::listing_status", "::title", "contribution_sku#1.value", "product_type#1.value",
        "parentage_level[marketplace_id=X]#1.value",
        "child_parent_sku_relationship[marketplace_id=X]#1.parent_sku",
        "variation_theme#1.name",
        "item_name[marketplace_id=X][language_tag=en_US]#1.value",
        "brand[marketplace_id=X][language_tag=en_US]#1.value",
        "amzn1.volt.ca.product_id_type", "amzn1.volt.ca.product_id_value",
        "main_product_image_locator[marketplace_id=X]#1.media_location",
        "other_product_image_locator_1[marketplace_id=X]#1.media_location",
        "other_product_image_locator_2[marketplace_id=X]#1.media_location",
        "product_description[marketplace_id=X][language_tag=en_US]#1.value",
        "bullet_point[marketplace_id=X][language_tag=en_US]#1.value",
        "bullet_point[marketplace_id=X][language_tag=en_US]#2.value",
        "generic_keyword[marketplace_id=X][language_tag=en_US]#1.value",
        "color[marketplace_id=X][language_tag=en_US]#1.value",
        "size[marketplace_id=X][language_tag=en_US]#1.value",
        "list_price[marketplace_id=X]#1.value",
        "purchasable_offer[marketplace_id=X][audience=ALL]#1.our_price#1.schedule#1.value_with_tax",
        "purchasable_offer[marketplace_id=X][audience=ALL]#1.discounted_price#1.schedule#1.value_with_tax",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(["Status", "Title", "SKU"] + [""] * (len(keys) - 3))   # display names
    ws.append(keys)                                                  # attribute keys
    ws.append(["Active", "Example product", "ABC123"] + [""] * (len(keys) - 3))  # Amazon example
    for r in rows:
        ws.append([r.get(k, "") for k in keys])
    p = tmp_path / name
    wb.save(p)
    return str(p)


def _row(sku, title, **kw):
    base = {"contribution_sku#1.value": sku,
            "item_name[marketplace_id=X][language_tag=en_US]#1.value": title,
            "::listing_status": "Active"}
    base.update(kw)
    return base


K = {
    "type": "amzn1.volt.ca.product_id_type", "id": "amzn1.volt.ca.product_id_value",
    "parentage": "parentage_level[marketplace_id=X]#1.value",
    "parent": "child_parent_sku_relationship[marketplace_id=X]#1.parent_sku",
    "theme": "variation_theme#1.name",
    "img": "main_product_image_locator[marketplace_id=X]#1.media_location",
    "img1": "other_product_image_locator_1[marketplace_id=X]#1.media_location",
    "img2": "other_product_image_locator_2[marketplace_id=X]#1.media_location",
    "desc": "product_description[marketplace_id=X][language_tag=en_US]#1.value",
    "b1": "bullet_point[marketplace_id=X][language_tag=en_US]#1.value",
    "b2": "bullet_point[marketplace_id=X][language_tag=en_US]#2.value",
    "kw1": "generic_keyword[marketplace_id=X][language_tag=en_US]#1.value",
    "color": "color[marketplace_id=X][language_tag=en_US]#1.value",
    "price": "purchasable_offer[marketplace_id=X][audience=ALL]#1.our_price#1.schedule#1.value_with_tax",
    "sale": "purchasable_offer[marketplace_id=X][audience=ALL]#1.discounted_price#1.schedule#1.value_with_tax",
    "list": "list_price[marketplace_id=X]#1.value",
}


def test_catalog_parse_merge_and_views(tmp_path):
    f1 = _clr_file(tmp_path, [
        _row("PARENT-1", "Widget (parent)", **{
            K["parentage"]: "Parent", K["theme"]: "COLOR"}),
        _row("KID-BLK", "Widget Black", **{
            K["parentage"]: "Child", K["parent"]: "PARENT-1", K["theme"]: "COLOR",
            K["type"]: "ASIN", K["id"]: "B0TEST0001", K["color"]: "Black",
            K["img"]: "https://img/main.jpg", K["img1"]: "https://img/1.jpg",
            K["desc"]: "A very nice widget.", K["b1"]: "First bullet", K["b2"]: "Second bullet",
            K["kw1"]: "widget black gadget", K["price"]: "19.99", K["sale"]: "15.50",
            K["list"]: "25.0"}),
        _row("SOLO-1", "Standalone thing", **{
            K["type"]: "ASIN", K["id"]: "B0TEST0002", K["price"]: "$9.95"}),
    ], name="a.xlsx")

    prods = cat.parse_clr(f1)
    # example row skipped, 3 real rows kept
    assert [p["sku"] for p in prods] == ["PARENT-1", "KID-BLK", "SOLO-1"]

    kid = prods[1]
    assert kid["asin"] == "B0TEST0001"
    assert kid["parentage"] == "child" and kid["parent_sku"] == "PARENT-1"
    assert kid["price"] == 19.99 and kid["sale_price"] == 15.5 and kid["list_price"] == 25.0
    assert kid["main_image"] == "https://img/main.jpg" and kid["images"] == ["https://img/1.jpg"]
    assert kid["bullets"] == ["First bullet", "Second bullet"]
    assert kid["description"] == "A very nice widget."
    assert kid["search_terms"] == ["widget black gadget"]
    assert prods[2]["price"] == 9.95           # $-stripped

    data, a, u = cat.merge({}, prods, "a.xlsx")
    assert (a, u) == (3, 0)
    light = {p["sku"]: p for p in cat.overview(data)["products"]}
    assert "description" not in light["KID-BLK"]           # heavy fields projected out
    assert light["KID-BLK"]["bullet_count"] == 2 and light["KID-BLK"]["image_count"] == 2

    # merge: second (category) file upserts by SKU — KID-BLK replaced (latest
    # report wins wholly), KID-GRN added
    f2 = _clr_file(tmp_path, [
        _row("KID-BLK", "Widget Black v2", **{K["parentage"]: "Child", K["parent"]: "PARENT-1"}),
        _row("KID-GRN", "Widget Green", **{K["parentage"]: "Child", K["parent"]: "PARENT-1"}),
    ], name="b.xlsx")
    data, a, u = cat.merge(data, cat.parse_clr(f2), "b.xlsx")
    assert (a, u) == (1, 1)
    assert data["products"]["KID-BLK"]["title"] == "Widget Black v2"
    assert data["products"]["KID-BLK"]["price"] is None    # full replace, no stale price
    assert [f["name"] for f in data["files"]] == ["a.xlsx", "b.xlsx"]

    ov = cat.overview(data)
    assert ov["stats"] == {"total": 4, "parents": 1, "children": 2, "standalone": 1,
                           "brands": 0, "priced": 1,
                           # parents don't count toward missing image/desc
                           "missing_image": 3, "missing_desc": 3,
                           # every non-parent has high/medium SEO issues in this fixture
                           "listing_issues": 3}

    # detail: parent lists children; child lists parent + sibling
    fam = cat.item(data, "PARENT-1")["family"]
    assert {c["sku"] for c in fam} == {"KID-BLK", "KID-GRN"}
    fam = cat.item(data, "KID-GRN")["family"]
    assert {c["sku"] for c in fam} == {"PARENT-1", "KID-BLK"}
    with pytest.raises(ValueError):
        cat.item(data, "NOPE")


def test_catalog_ads_breakeven_join():
    """enrich(): Product-Ads join (campaigns per ASIN) + break-even economics."""
    econ = {"default_referral_pct": 0.15, "default_cogs_pct": 0.30}

    # be_metrics: derived from price — BE = 1 - ref - cogs% = 0.55
    be = cat.be_metrics({"price": 20.0}, None, econ)
    assert be["break_even_acos"] == 0.55 and be["source"] == "derived"
    assert be["profit_per_unit"] == 11.0 and be["cogs"] == 6.0 and be["amazon_fee"] == 3.0
    # sale price wins over price; uploaded benchmark BE wins over derived
    assert cat.be_metrics({"price": 20, "sale_price": 10}, None, econ)["price"] == 10
    up = cat.be_metrics({"price": 20.0}, {"break_even_acos": 0.25}, econ)
    assert up["break_even_acos"] == 0.25 and up["source"] == "uploaded"
    # no price anywhere -> nothing to compute
    assert cat.be_metrics({}, None, econ) is None
    # benchmark file's own sale_price fills a price-less catalog row
    assert cat.be_metrics({}, {"sale_price": 20.0}, econ)["break_even_acos"] == 0.55

    view = {"products": [
        {"sku": "A", "asin": "B0AAA", "price": 20.0},    # ACoS 0.30 < BE 0.55 -> profitable
        {"sku": "B", "asin": "B0BBB", "price": 20.0},    # spend, 0 sales -> bleeding
        {"sku": "C", "asin": "B0CCC", "price": 20.0},    # not advertised
    ], "stats": {}}
    ads = {"asins": {
        "B0AAA": {"ads": 2, "campaigns": 3, "spend": 30.0, "sales": 100.0, "acos": 0.30, "orders": 5},
        "B0BBB": {"ads": 1, "campaigns": 1, "spend": 12.0, "sales": 0.0, "acos": None, "orders": 0},
    }, "campaign_types": {"total": 4, "auto": 1, "keyword": 2, "product": 1, "manual": 0}}
    out = cat.enrich(view, ads, {}, econ)
    rows = {p["sku"]: p for p in out["products"]}
    assert rows["A"]["ads"]["campaigns"] == 3 and rows["A"]["be_status"] == "profitable"
    assert rows["B"]["be_status"] == "bleeding"
    assert rows["C"]["ads"] is None and rows["C"]["be_status"] is None
    assert rows["C"]["be"]["break_even_acos"] == 0.55      # BE needs no ads
    assert out["ads_connected"] is True
    # Avg ACoS = aggregate Σ spend / Σ sales over the advertised products
    # (30 + 12) / (100 + 0) = 0.42 — zero-sale spend counts in the numerator
    assert out["stats"] == {"advertised": 2, "campaigns": 4, "over_be": 1, "under_be": 1,
                            "avg_acos": 0.42}

    # no Product Ads upload: BE still computed, ads_connected False
    out = cat.enrich({"products": [{"sku": "A", "asin": "B0AAA", "price": 20.0}], "stats": {}},
                     None, {}, econ)
    assert out["ads_connected"] is False
    assert out["products"][0]["be"]["break_even_acos"] == 0.55
    assert out["stats"]["campaigns"] == 0 and out["stats"]["avg_acos"] is None


def test_cogs_override_and_default():
    """Per-SKU COGS (% of selling price): 40% default, override wins, parse both
    decimal and percent notation."""
    econ = {"default_referral_pct": 0.15, "default_cogs_pct": 0.0}

    # no econ COGS + no override -> 40% default: BE = 1 - 0.15 - 0.40 = 0.45
    be = cat.be_metrics({"price": 20.0}, None, econ)
    assert be["break_even_acos"] == 0.45 and be["cogs_pct"] == 0.40
    assert be["cogs"] == 8.0 and be["profit_per_unit"] == 9.0
    assert be["cogs_custom"] is False

    # per-SKU override wins: 25% -> BE = 0.60
    be = cat.be_metrics({"price": 20.0}, None, econ, 0.25)
    assert be["break_even_acos"] == 0.60 and be["cogs"] == 5.0
    assert be["profit_per_unit"] == 12.0 and be["cogs_custom"] is True

    # parse: decimal, bare percent, % suffix, clear, clamp
    assert cat.parse_cogs("0.35") == 0.35
    assert cat.parse_cogs("35") == 0.35
    assert cat.parse_cogs("35%") == 0.35
    assert cat.parse_cogs(" 40 % ") == 0.40
    assert cat.parse_cogs("") is None and cat.parse_cogs(None) is None
    assert cat.parse_cogs("250") == 0.95        # clamped
    with pytest.raises(ValueError):
        cat.parse_cogs("abc")

    # enrich threads the per-SKU map through
    view = {"products": [{"sku": "A", "asin": "B0AAA", "price": 20.0},
                         {"sku": "B", "asin": "B0BBB", "price": 20.0}], "stats": {}}
    out = cat.enrich(view, None, {}, econ, {"A": 0.25})
    rows = {p["sku"]: p for p in out["products"]}
    assert rows["A"]["be"]["break_even_acos"] == 0.60
    assert rows["B"]["be"]["break_even_acos"] == 0.45


def test_cogs_store_roundtrip_and_audit_be_map(tmp_path, monkeypatch):
    """Catalog products feed the PPC Audit's break-even map (BLEEDING flags /
    BidOptimizer caps) using price + per-SKU COGS; benchmark upload wins."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from app import database as dbmod
    from app.pipeline import benchmark as bn

    monkeypatch.setattr(dbmod, "_store_dir", lambda sid: str(tmp_path))
    cat.write_catalog("s1", {"products": {
        "A": {"sku": "A", "asin": "B0AAA", "price": 20.0},
        "B": {"sku": "B", "asin": "B0BBB", "price": 20.0},
        "C": {"sku": "C", "asin": "B0CCC"},                    # no price -> skipped
    }})
    cat.write_cogs("s1", {"B": 0.25})
    assert cat.read_cogs("s1") == {"B": 0.25}

    eng = create_engine("sqlite:///:memory:")
    dbmod.Base.metadata.create_all(eng)
    db = sessionmaker(bind=eng)()
    db.info.update(store="s1", project="p1")
    be = bn.break_even_map(db)
    assert be["B0AAA"] == 0.45      # 40% default COGS
    assert be["B0BBB"] == 0.60      # 25% override
    assert "B0CCC" not in be
    db.close()


def test_be_uses_real_ledger_fees(tmp_path, monkeypatch):
    """BE ACoS must be price-DEPENDENT: the SKU's real referral % and per-unit
    FBA fee come from the Transactions ledger, so two products with the same
    COGS % no longer share one constant 45% break-even."""
    from app import database as dbmod
    from app.pipeline import transactions as txn
    monkeypatch.setattr(dbmod, "_store_dir", lambda sid: str(tmp_path))

    econ = {"default_referral_pct": 0.15, "default_cogs_pct": 0.0}
    # no ledger fees -> percent-only default: BE = 1 - .15 - .40 = 45% for ANY price
    assert cat.be_metrics({"price": 10.0}, None, econ)["break_even_acos"] == 0.45
    assert cat.be_metrics({"price": 50.0}, None, econ)["break_even_acos"] == 0.45

    # real fees: referral 15%, FBA $3.54/unit — weighs 35.4% on a $10 product
    # but 7.1% on a $50 one -> different break-evens
    fees = {"referral_pct": 0.15, "fba_fee": 3.54}
    lo = cat.be_metrics({"price": 10.0}, None, econ, None, fees)
    hi = cat.be_metrics({"price": 50.0}, None, econ, None, fees)
    assert lo["break_even_acos"] == pytest.approx(0.096, abs=1e-3)   # 1-.15-.40-.354
    assert hi["break_even_acos"] == pytest.approx(0.3792, abs=1e-3)
    assert lo["fees_real"] and lo["amazon_fee"] == pytest.approx(5.04)

    # fees_by_sku aggregates real order rows from the ledger (normalized SKU)
    txn.write_txn("s1", {"rows": [
        {"type": "Order", "sku": "PI 100", "date": "2026-06-01", "quantity": 2,
         "product_sales": 70.0, "selling_fees": -10.5, "fba_fees": -7.08,
         "promo": 0, "shipping_credits": 0, "gift_wrap": 0, "withheld_tax": 0,
         "other_fees": 0, "other": 0, "total": 52.42, "datetime": "x", "settlement_id": "1",
         "order_id": "o1", "description": "", "marketplace": "", "fulfillment": "",
         "city": "", "state": "", "status": ""},
    ]})
    f = cat.fees_by_sku("s1")
    assert f == {"pi100": {"referral_pct": 0.15, "fba_fee": 3.54, "fba_source": "ledger"}}


def test_be_sku_listing_match(tmp_path, monkeypatch):
    """When the ad's ASIN isn't in the catalog, the break-even joins by
    NORMALIZED SKU against the catalog listing ('PI 100' == 'pi-100')."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from app import database as dbmod
    from app import models as md
    from app.pipeline import productads as pa

    monkeypatch.setattr(dbmod, "_store_dir", lambda sid: str(tmp_path))
    cat.write_catalog("s1", {"products": {
        "pi-100": {"sku": "pi-100", "asin": "B0CATALOG", "price": 20.0},
    }})
    cat.write_cogs("s1", {"pi-100": 0.25})       # BE = 1 - .15 - .25 = 0.60
    assert cat.be_by_sku("s1", {"default_referral_pct": 0.15}) == {"pi100": 0.60}

    eng = create_engine("sqlite:///:memory:")
    dbmod.Base.metadata.create_all(eng)
    db = sessionmaker(bind=eng)()
    db.info.update(store="s1", project="p1")
    # ad row with a DIFFERENT asin but a sku that matches the listing loosely
    db.add(md.ProductAdFact(asin="B0ADSONLY", sku="PI 100", campaign_id="c1",
                            campaign_type="manual", state="enabled",
                            impressions=100, clicks=10, spend=5.0, sales=10.0,
                            orders=1, units=1))
    db.commit()
    rows = pa.summary(db)["rows"]
    assert rows[0]["break_even_acos"] == 0.60    # matched via normalized SKU
    db.close()


def test_catalog_seo_check():
    """Catalog-level SEO recommendations from CLR fields alone."""
    assert cat.seo_check({"parentage": "parent"}) is None   # variation containers skipped

    # thin listing: short title, 2 bullets, no desc, no search terms, no main image
    p = {"title": "Short Widget Title", "bullets": ["a", "b"], "description": "",
         "search_terms": [], "main_image": "", "images": []}
    s = cat.seo_check(p)
    areas = {(r["area"], r["severity"]) for r in s["recommendations"]}
    assert ("search_terms", "high") in areas and ("images", "high") in areas
    assert ("bullets", "medium") in areas and ("description", "medium") in areas
    assert ("title", "low") in areas
    assert s["issues"] == s["counts"]["high"] + s["counts"]["medium"] == 4

    # wasted backend words: 'widget'/'blue' visible in title/bullets -> flagged;
    # over-249-byte field -> high
    p2 = {"title": "Premium Widget " + "x" * 70, "bullets": ["Blue and sturdy"] * 5,
          "description": "d" * 500, "main_image": "https://img/m.jpg",
          "images": ["1", "2", "3", "4"],
          "search_terms": ["widget blue gadget " + "longword" * 30]}
    s2 = cat.seo_check(p2)
    st_recs = [r for r in s2["recommendations"] if r["area"] == "search_terms"]
    assert any("over 249" in r["title"] for r in st_recs)
    wasted = next(r for r in st_recs if "already indexed" in r["title"])
    assert set(wasted["words"]) == {"widget", "blue"}
    # clean elsewhere: no title/bullets/description/images recs
    assert not {r["area"] for r in s2["recommendations"]} - {"search_terms"}


def test_productads_by_asin():
    """by_asin(): per-ASIN rollup with DISTINCT campaign counts by targeting kind."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from app.database import Base
    from app import models as md
    from app.pipeline import productads as pa

    eng = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(eng)
    db = sessionmaker(bind=eng)()
    try:
        # B0X: 3 ads across 2 campaigns (auto + keyword; c1 repeated — distinct count)
        db.add_all([
            md.ProductAdFact(asin="B0X", sku="S1", campaign_id="c1", campaign_type="auto",
                             impressions=100, clicks=10, spend=5.0, sales=50.0, orders=2, units=2),
            md.ProductAdFact(asin="B0X", sku="S2", campaign_id="c1", campaign_type="auto",
                             impressions=50, clicks=5, spend=2.5, sales=0.0, orders=0, units=0),
            md.ProductAdFact(asin="B0X", sku="S1", campaign_id="c2", campaign_type="keyword",
                             impressions=10, clicks=1, spend=1.0, sales=10.0, orders=1, units=1),
            md.ProductAdFact(asin="B0Y", sku="S3", campaign_id="c2", campaign_type="keyword",
                             impressions=10, clicks=2, spend=3.0, sales=0.0, orders=0, units=0),
        ])
        db.commit()
        out = pa.by_asin(db)
        x = out["asins"]["B0X"]
        assert x["ads"] == 3 and x["campaigns"] == 2
        assert x["auto_campaigns"] == 1 and x["keyword_campaigns"] == 1
        assert x["spend"] == 8.5 and x["sales"] == 60.0 and x["orders"] == 3
        assert x["acos"] == pytest.approx(8.5 / 60.0)
        assert out["asins"]["B0Y"]["campaigns"] == 1
        # account-wide distinct campaigns: c1 + c2 = 2 (c2 shared across ASINs)
        assert out["campaign_types"]["total"] == 2
    finally:
        db.close()


def test_catalog_rejects_non_clr(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.append(["Campaign", "Ad Group", "Spend"])
    p = tmp_path / "bulk.xlsx"
    wb.save(p)
    with pytest.raises(ValueError):
        cat.parse_clr(str(p))


def test_report_seo_block(tmp_path, monkeypatch):
    """Exec report SEO section: catalog listing issues + per-project keyword recs."""
    from app import database as dbmod
    from app import models as md
    from app.pipeline import report as rp
    from app.pipeline import tracker as tk
    monkeypatch.setattr(dbmod, "STORES_DIR", str(tmp_path))

    cat.write_catalog("s1", {"products": {
        "SKU1": {"sku": "SKU1", "asin": "B0A", "title": "Tiny", "bullets": [],
                 "description": "", "search_terms": [], "main_image": "", "images": []},
        "PAR1": {"sku": "PAR1", "parentage": "parent"},      # parents skipped
    }})
    db = dbmod.get_session("s1", "default")
    try:
        pid = tk.create_project(db, "p1")["project_id"]
        db.add(md.TrackedKeyword(project_id=pid, keyword="alpha beta", search_volume=100))
        db.commit()
        blk = rp._seo_block(db)
    finally:
        db.close()

    c = blk["catalog"]
    assert c["products"] == 1 and c["with_issues"] == 1
    assert c["worst"][0]["sku"] == "SKU1" and c["worst"][0]["issues"] >= 3
    assert c["by_area"].get("search_terms") == 1
    p = blk["projects"][0]
    assert p["name"] == "p1" and p["keywords"] == 1
    assert p["high"] >= 1       # nothing pasted -> missing-copy recs
    assert p["st_over"] is False


def test_report_export_xlsx_charts(tmp_path, monkeypatch):
    """Exec-report workbook: styled sheets + native charts when data exists."""
    import io as _io
    import openpyxl
    from app import database as dbmod
    from app import models as md
    from app.config import default_thresholds
    from app.pipeline import report as rp
    monkeypatch.setattr(dbmod, "STORES_DIR", str(tmp_path))
    db = dbmod.get_session("s2", "default")
    try:
        # one flag source: a campaign+ad+target with wasted spend
        db.add(md.DimCampaign(campaign_id="c1", name="Camp"))
        db.commit()
        th = default_thresholds.merged(target_acos=0.25)
        data = rp.export_xlsx(db, th)
    finally:
        db.close()
    wb = openpyxl.load_workbook(_io.BytesIO(data))
    assert "Summary" in wb.sheetnames and "Flags" in wb.sheetnames
    assert wb["Summary"].cell(row=1, column=1).value == "PPC exec report"
    # header style applied on the Flags sheet
    assert wb["Flags"].cell(row=1, column=1).value == "Flag"
    assert wb["Flags"].cell(row=1, column=1).font.bold


def test_catalog_report_xlsx():
    """Product Benchmark workbook: Overview + Products sheets, charts attached."""
    import io as _io
    import openpyxl
    data = {"products": {
        "SKU1": {"sku": "SKU1", "asin": "B0A", "title": "Thin thing", "bullets": [],
                 "description": "", "search_terms": [], "main_image": "", "images": [],
                 "price": 20.0},
        "PAR1": {"sku": "PAR1", "parentage": "parent", "title": "Parent"},
    }}
    view = cat.enrich(cat.overview(data),
                      {"asins": {"B0A": {"ads": 1, "campaigns": 2, "auto_campaigns": 1,
                                         "keyword_campaigns": 1, "product_campaigns": 0,
                                         "manual_campaigns": 0, "spend": 9.0, "sales": 30.0,
                                         "acos": 0.3, "orders": 1, "impressions": 10,
                                         "clicks": 2, "units": 1, "ctr": None, "cpc": None,
                                         "cvr": None, "roas": None}},
                       "campaign_types": {"total": 2, "auto": 1, "keyword": 1, "product": 0, "manual": 0}},
                      {}, {"default_referral_pct": 0.15, "default_cogs_pct": 0.0})
    out = cat.report_xlsx(data, view)
    wb = openpyxl.load_workbook(_io.BytesIO(out))
    assert wb.sheetnames == ["Overview", "Products"]
    assert len(wb["Overview"]._charts) == 2         # composition pie + issues bar
    assert len(wb["Products"]._charts) == 1         # ad-spend bar (ads connected)
    pr = wb["Products"]
    assert pr.cell(row=2, column=1).value == "SKU1"     # top spend first
    assert pr.cell(row=2, column=13).value == 8.0       # COGS / unit (40% of $20)
    assert pr.cell(row=2, column=14).value == 0.0       # fulfillment / unit (no fee data)
    assert pr.cell(row=2, column=15).value == 3.0       # referral / unit (15%)
    assert pr.cell(row=2, column=16).value == 3.0       # total fees / unit
    assert pr.cell(row=2, column=18).value == 2         # campaigns
    with pytest.raises(ValueError):
        cat.report_xlsx({}, {"products": [], "stats": {}})

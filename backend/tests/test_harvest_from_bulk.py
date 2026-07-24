"""Search-Term Harvest from the SP BULK file: the embedded 'SP Search Term
Report' sheet carries real entity IDs, so harvest promotes/negates act by
exact Campaign / Ad Group ID (Weekly engine reused; no name mapping)."""
from types import SimpleNamespace

import openpyxl
import pytest

from app.config import default_thresholds
from app.pipeline import weekly as wk

HEAD = ["Campaign ID", "Ad Group ID", "Keyword ID", "Product Targeting ID",
        "Campaign Name", "Ad Group Name", "Match Type", "Customer Search Term",
        "Impressions", "Clicks", "Spend", "7 Day Total Sales", "7 Day Total Orders"]


def _bulk(tmp_path, rows):
    wb = openpyxl.Workbook()
    wb.active.title = "Sponsored Products Campaigns"          # decoy sheet
    wb.active.append(["Entity"])
    ws = wb.create_sheet("SP Search Term Report")
    ws.append(HEAD)
    for r in rows:
        ws.append(r)
    p = tmp_path / "bulk.xlsx"
    wb.save(p)
    return str(p)


def test_harvest_from_bulk_keeps_exact_ids(tmp_path):
    # 18-digit IDs — past float64 exactness; must survive as exact strings
    path = _bulk(tmp_path, [
        ["502533615771891001", "523899586073950002", "111", "", "Camp A", "AG kw",
         "broad", "winner term", 1000, 40, 10.0, 100.0, 5],
        ["502533615771891001", "523899586073950002", "111", "", "Camp A", "AG kw",
         "broad", "loser term", 500, 30, 12.0, 0.0, 0],
    ])
    rows = [SimpleNamespace(**r) for r in wk.parse_str_sheet(path)]
    assert rows[0].campaign_id == "502533615771891001"        # exact string

    th = default_thresholds.merged(target_acos=0.25, min_spend=5.0)
    h = wk.compute_harvest(rows, th, min_orders=1)
    pro = {p["search_term"]: p for p in h["promotes"]}
    neg = {n["search_term"]: n for n in h["negates"]}
    assert pro["winner term"]["campaign_id"] == "502533615771891001"
    assert pro["winner term"]["as"] == "keyword"
    assert neg["loser term"]["ad_group_id"] == "523899586073950002"

    # chosen rows -> bulk by exact ID (the /harvest/from-bulk/file path)
    data = wk.to_bulk([], h["promotes"] + h["negates"])
    ws = openpyxl.load_workbook(__import__("io").BytesIO(data)).active
    out = [dict(zip([c.value for c in ws[1]], [c.value for c in r])) for r in ws.iter_rows(min_row=2)]
    kw = next(r for r in out if r["Entity"] == "Keyword")
    nk = next(r for r in out if r["Entity"] == "Negative Keyword")
    assert kw["Campaign ID"] == "502533615771891001" and kw["Keyword Text"] == "winner term"
    assert kw["Match Type"] == "Exact" and kw["Operation"] == "create"
    assert nk["Keyword Text"] == "loser term" and nk["Match Type"] == "Negative Exact"


def test_harvest_from_bulk_rejects_idless_str(tmp_path):
    """A standalone STR (no ID columns) must raise the friendly error — the
    router then falls back to the legacy name-mapping flow."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SP Search Term Report"
    ws.append(["Campaign Name", "Ad Group Name", "Customer Search Term",
               "Impressions", "Clicks", "Spend", "7 Day Total Sales", "7 Day Total Orders"])
    ws.append(["Camp A", "AG kw", "some term", 10, 1, 1.0, 0.0, 0])
    p = tmp_path / "str.xlsx"
    wb.save(p)
    with pytest.raises(ValueError, match="Campaign ID / Ad Group ID"):
        wk.parse_str_sheet(str(p))


def test_ngrams_mine_from_bulk_file(tmp_path):
    """The N-gram miner reads the BULK file too: harvest.parse_str skips the
    Campaigns sheet and picks the embedded 'SP Search Term Report' sheet."""
    from app.pipeline import harvest as hv
    from app.pipeline import ngrams as ng
    path = _bulk(tmp_path, [
        ["1", "2", "", "", "Camp A", "AG", "broad", "ice pack wrap", 100, 10, 5.0, 50.0, 3],
        ["1", "2", "", "", "Camp A", "AG", "broad", "ice pack sleeve", 80, 8, 20.0, 0.0, 0],
    ])
    df = hv.parse_str(path)
    assert list(df["search_term"]) == ["ice pack wrap", "ice pack sleeve"]
    th = default_thresholds.merged(target_acos=0.25, min_spend=5.0)
    grams = {g.gram: g for g in ng.mine(df, th, min_clicks=2)}
    assert grams["ice pack"].clicks == 18            # aggregated across both terms
    assert grams["wrap"].verdict == "winner"
    assert grams["sleeve"].verdict == "waster"

"""Main Dashboard analytics hub — cross-feature rollup + charted Excel export."""
import io
from datetime import date

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app import models as md
from app.config import default_thresholds
from app.pipeline import dashboard as dash
from app.pipeline import transactions as t

from .test_transactions import CSV


@pytest.fixture()
def db(tmp_path, monkeypatch):
    from app import database as dbmod
    monkeypatch.setattr(dbmod, "_store_dir", lambda sid: str(tmp_path))
    eng = create_engine("sqlite:///:memory:")
    dbmod.Base.metadata.create_all(eng)
    s = sessionmaker(bind=eng)()
    s.info.update(store="s1", project="p1")
    yield s
    s.close()


def _seed(db, tmp_path):
    # minimal star schema: one campaign -> ad group -> ad -> target + one fact
    db.add(md.DimProduct(asin="B0TEST"))
    db.add(md.DimCampaign(campaign_id="c1", name="Camp", targeting_type="manual", state="enabled"))
    db.add(md.DimAdGroup(ad_group_id="ag1", campaign_id="c1", name="AG", state="enabled"))
    db.add(md.DimAd(ad_id="ad1", ad_group_id="ag1", asin="B0TEST", sku="SKU-A", state="enabled"))
    db.add(md.DimTarget(target_id="t1", ad_group_id="ag1", target_type="keyword",
                        keyword_text="widget", match_type="exact", bid=1.0, state="enabled"))
    db.add(md.FactPerformance(entity_type="target", entity_id="t1", snapshot_date=date(2026, 6, 30),
                              impressions=1000, clicks=50, spend=40.0, sales=20.0, orders=1, units=1))
    db.add(md.FactPerformance(entity_type="campaign", entity_id="c1", snapshot_date=date(2026, 6, 30),
                              impressions=1000, clicks=50, spend=40.0, sales=20.0, orders=1, units=1))
    db.add(md.MinedKeyword(keyword="widget pro", display="widget pro"))
    db.commit()
    # store transaction ledger
    p = tmp_path / "txn.csv"
    p.write_text(CSV)
    data, _, _, _ = t.merge({}, t.parse_txn(str(p)), "june.csv")
    t.write_txn("s1", data)


def test_analytics_blocks(db, tmp_path):
    _seed(db, tmp_path)
    a = dash.analytics(db, db)      # base db == cadence db in tests
    assert a["product_ads"] is None and a["catalog"] is None and a["monitoring"] is None
    assert a["transactions"]["totals"]["orders"] == 3
    assert len(a["transactions"]["days"]) == 4
    assert a["keywords"]["count"] == 1
    assert a["movers"] is None      # single snapshot


def test_export_xlsx(db, tmp_path):
    from openpyxl import load_workbook
    _seed(db, tmp_path)
    th = default_thresholds.merged(target_acos=0.25)
    wb = load_workbook(io.BytesIO(dash.export_xlsx(db, db, th)))
    assert wb.sheetnames == ["Overview", "Features", "Transactions"]   # no movers sheet
    ov = wb["Overview"]
    kpis = {ov.cell(row=r, column=1).value: ov.cell(row=r, column=2).value for r in range(4, 14)}
    assert kpis["Ad spend"] == 40.0 and kpis["Ad sales"] == 20.0
    assert kpis["Open flags"] >= 1                     # 200% ACoS must flag
    assert len(ov._charts) == 2                        # flag bar + account-states pie
    assert len(wb["Transactions"]._charts) == 2        # daily line + top-SKU bar
    fx = wb["Features"]
    texts = [fx.cell(row=r, column=1).value for r in range(1, 30)]
    assert "no Product Ads upload in this audit" in texts
    assert "no catalog uploaded" in texts
    assert "no daily tracker data" in texts

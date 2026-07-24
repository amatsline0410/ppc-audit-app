"""Competitor/SEO tracker — Phase 0 mapping-driven importer + metrics.

Fixture is a small anonymized slice built in the SAME layout as the real sheet
("Project Snore x Competitor Research"): Main KPI/attr header + keyword grid with
'Organic Rank'/'Filter' chrome rows, X-ray tabs, Listing Audit markers,
Listing Copy Current/Proposed blocks, Search Terms keyword list.
"""
from __future__ import annotations
import io
from datetime import date
import pandas as pd
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from app.database import Base
from app import models as md
from app.pipeline import tracker as tk

YOU, C1, C2 = "B0TESTPRI1", "B0TESTCMP1", "B0TESTCMP2"


@pytest.fixture()
def db():
    eng = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(eng)
    s = sessionmaker(bind=eng)()
    yield s
    s.close()


def _sheet_path(tmp_path) -> str:
    """Build the fixture xlsx in the real sheet's cell layout."""
    def grid(rows, ncols):
        out = []
        for r in rows:
            r = list(r) + [None] * (ncols - len(r))
            out.append(r[:ncols])
        return pd.DataFrame(out)

    main = grid([
        [],
        [],
        [],
        [],
        [],
        [],
        [None, None, None, "Brand", "You", "CompOne", "CompTwo"],
        [None, None, None, "ASIN", YOU, C1, C2],
        [None, None, None, "ASIN", YOU, C1, C2],
        [None, None, None, "Price", "10", "20", "30"],
        [None, None, None, "Listing Health Score", "8.75", "10", "6.25"],
        [None, None, None, "PDP Images", "Yes", "Yes", None],
        [None, None, None, "CEREBRO DATA"],
        [],
        ["Search Terms", "Search Volume", "Relevancy", "Keyword Sales", YOU, C1, C2],
        [None, None, None, None, "Organic Rank"],
        [None, None, None, None, "Filter", "Filter", "Filter"],
        ["alpha keyword", "1000", "1", "5", "3", "12", "-"],
        ["beta keyword", "500", "2", "-", "-", "50", "100"],
        ["gamma keyword", "200", "0", None, "0", "-", "49"],   # 0 rank = unranked
    ], 8)

    xray_head = ["Product Details", "ASIN", "Brand", "Price $", "Sales", "Revenue", "BSR",
                 "Seller Country/Region", "FBA Fees $", "Active Sellers #", "Ratings",
                 "Review Count", "Images", "Review velocity", "Buy Box", "Category",
                 "Size Tier", "Fulfillment", "Dimensions", "Weight", "Creation Date",
                 "Image Source URL"]
    your_xray = pd.DataFrame([xray_head,
        ["Your Product", YOU, "You", "10", "5", "50.0", "999", "US", "3", "1", "4.5",
         "10", "7", "0", "You", "Cat", "Large", "FBA", "1x1x1", "0.1", "2025-01-01", "http://x"]])
    comp_xray = pd.DataFrame([xray_head,
        ["Comp One", C1, "CompOne", "20", "100", "2000.0", "50", "US", "5", "1", "4.0",
         "500", "8", "9", "CompOne", "Cat", "Large", "FBA", "1x1x1", "0.2", "2020-01-01", "http://y"],
        ["Comp Two", C2, "CompTwo", "30", "50", "1500.0", "80", "CN", "6", "2", "3.5",
         "300", "6", "4", "CompTwo", "Cat", "Small", "FBA", "1x1x1", "0.3", "2021-06-01", "http://z"]])

    la = grid([
        [],
        [None] * 15 + ["You", "CompOne", "CompTwo"],
        ["Search Terms", "Search Volume", "Relevancy", "Title", "Bullet Points",
         "A+ /\nBrand Story", "Description", "Search Terms", "Alt text",
         "SP Targeting Broad", "SP Targeting Phrase", "SP Targeting Exact",
         "SB Targeting Broad", "SB Targeting Phrase", "SB Targeting Exact", YOU, C1, C2],
        ["alpha keyword", "1000", "1", "exact", None, "broad", None, None, None,
         None, None, "exact", None, None, None, "3", "12", "-"],
        ["beta keyword", "500", "2", None, "broad", None, None, None, None,
         None, None, None, None, None, None, "-", "50", "100"],
    ], 18)

    lc = grid([
        [None, None, None, None, None, "Current", None, None, "Current",
         None, None, None, None, None, None, None, None, None, None, None, None, None, None, "Proposed"],
        [None, None, None, None, None, "Our current title alpha keyword", None, None,
         "Our current bullets", None, None, None, None, None, None, None, None, None,
         None, None, None, None, None, "Proposed title beta keyword"],
        [None, None, None, None, None, "Title", None, None, "Bullet Points",
         None, None, None, None, None, None, None, None, None, None, None, None, None, None, "Title"],
        ["alpha keyword", "1000", "1"],
    ], 24)

    bp = grid([
        [None, None, None, None, "Your BP", "CompOne", "CompTwo"],
        [None, None, None, None, "our bullets text", "comp1 bullets", "comp2 bullets"],
        [None, None, "No. of exact", None, "1", "2", "3"],
        [None, None, "Total exact SV", None, "100", "200", "300"],
        ["Search Terms", "Search Volume", "Relevancy", None, YOU, C1, C2],
    ], 7)

    st = grid([
        [None, None, None, None, None, "Proposed BP"],
        [],
        [],
        ["delta keyword", "750", "3"],
        ["alpha keyword", "1000", "1"],       # dup — must not duplicate
    ], 6)

    path = str(tmp_path / "fixture_sheet.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for name, df in [("Main", main), ("Your X-ray", your_xray),
                         ("Comp X-ray", comp_xray), ("Listing Audit", la),
                         ("Listing Copy", lc), ("BP Comparison", bp),
                         ("Search Terms", st)]:
            df.to_excel(xw, sheet_name=name, index=False, header=False)
    return path


# ---- pure metrics ------------------------------------------------------------

def test_page_boundaries():
    assert tk.page(None) is None
    assert tk.page(1) == 1
    assert tk.page(48) == 1
    assert tk.page(49) == 2
    assert tk.page(96) == 2
    assert tk.page(97) == 3


def test_index_rate_none_safe():
    assert tk.index_rate(5, 0) is None
    assert tk.index_rate(5, 10) == 0.5


def test_rank_delta_sign():
    assert tk.rank_delta(10, 30) == -20      # negative = improved
    assert tk.rank_delta(30, 10) == 20
    assert tk.rank_delta(None, 10) is None
    assert tk.rank_delta(10, None) is None


def test_match_kind():
    assert tk.match_kind("Best alpha keyword strips", "alpha keyword") == "exact"
    assert tk.match_kind("keyword for alpha people", "alpha keyword") == "broad"
    assert tk.match_kind("totally unrelated", "alpha keyword") is None
    assert tk.match_kind("", "alpha keyword") is None


def test_count_exact_and_sv(db):
    kws = [md.TrackedKeyword(project_id=1, keyword="alpha keyword", search_volume=100),
           md.TrackedKeyword(project_id=1, keyword="beta keyword", search_volume=50)]
    text = "contains alpha keyword only"
    assert tk.count_exact(text, kws) == 1
    assert tk.total_exact_sv(text, kws) == 100


# ---- migration e2e -----------------------------------------------------------

def test_migrate_e2e(db, tmp_path):
    res = tk.migrate(db, _sheet_path(tmp_path), "Fixture", date(2026, 7, 1))
    assert res["primary_asin"] == YOU
    assert res["competitors"] == 3
    assert res["keywords"] == 4                     # 3 grid + delta from Search Terms
    assert res["search_terms_added"] == 1
    # rank cells: alpha 3/12, beta 50/100, gamma 49 -> 5 ('-' and 0 dropped)
    assert res["rank_cells"] == 5
    assert res["usage_markers"] == 4                # exact,broad,exact + broad

    m = tk.matrix(db, res["project_id"])
    assert m["asins"][0] == YOU                     # primary pinned first
    grid = {r["keyword"]: r["ranks"] for r in m["rows"]}
    a = m["asins"]
    assert grid["alpha keyword"][a.index(YOU)] == 3
    assert grid["alpha keyword"][a.index(C1)] == 12
    assert grid["alpha keyword"][a.index(C2)] is None       # '-' -> unranked
    assert grid["gamma keyword"][a.index(YOU)] is None      # 0 -> unranked
    assert grid["gamma keyword"][a.index(C2)] == 49
    assert grid["delta keyword"] == [None, None, None]      # list-only keyword

    # attributes + manual health block + copy blocks
    comp = {c.asin: c for c in db.query(md.TrackedCompetitor)}
    assert comp[YOU].is_primary and comp[C1].revenue == 2000.0
    assert comp[YOU].listing_health_score == 8.75
    assert comp[YOU].pdp_images is True and comp[C2].pdp_images is None
    assert comp[C1].bullet_points == "comp1 bullets"
    lc = {(x.variant, x.element): x.text for x in db.query(md.ListingCopy)}
    assert lc[("current", "title")].startswith("Our current title")
    assert lc[("proposed", "title")].startswith("Proposed title")

    # re-migrate same name replaces, not duplicates
    tk.migrate(db, _sheet_path(tmp_path), "Fixture", date(2026, 7, 1))
    assert db.query(md.TrackerProject).count() == 1
    assert db.query(md.TrackedKeyword).count() == 4


def test_scorecard_and_kpis(db, tmp_path):
    res = tk.migrate(db, _sheet_path(tmp_path), "Fixture", date(2026, 7, 1))
    sc = tk.scorecard(db, res["project_id"])
    cards = {c["asin"]: c for c in sc["cards"]}
    you = cards[YOU]
    assert you["kw_tracked"] == 4 and you["kw_ranked"] == 1   # alpha only (gamma=0)
    assert you["page1_count"] == 1 and you["index_rate"] == 0.25
    # best competitor page1: C1 has ranks 12,50 -> page1=1 ; coverage 1/1
    assert you["coverage_vs_best"] == 1.0
    assert sc["kpis"]["total_revenue_top10"] == 3500.0
    assert sc["kpis"]["market_share"] == round(50.0 / 3500.0, 6)


# ---- snapshots ---------------------------------------------------------------

def _cerebro_csv(tmp_path, rows):
    p = tmp_path / "cerebro.csv"
    pd.DataFrame(rows).to_csv(p, index=False)
    return str(p)


def test_snapshot_import_and_idempotent_replace(db, tmp_path):
    res = tk.migrate(db, _sheet_path(tmp_path), "Fixture", date(2026, 7, 1))
    pid = res["project_id"]
    # week 2: alpha improved 3->2, beta now ranked, epsilon is NEW
    path = _cerebro_csv(tmp_path, [
        {"Keyword Phrase": "alpha keyword", "Search Volume": "1100", "Organic Rank": "2"},
        {"Keyword Phrase": "beta keyword", "Search Volume": "500", "Organic Rank": "40"},
        {"Keyword Phrase": "epsilon keyword", "Search Volume": "900", "Organic Rank": "7"},
    ])
    r2 = tk.import_cerebro(db, pid, path, date(2026, 7, 8), asin=YOU)
    assert r2["rank_rows"] == 3 and r2["keywords_added"] == 1

    # re-import SAME date with a different rank -> replaces, no duplicates
    path2 = _cerebro_csv(tmp_path, [
        {"Keyword Phrase": "alpha keyword", "Search Volume": "1100", "Organic Rank": "5"}])
    tk.import_cerebro(db, pid, path2, date(2026, 7, 8), asin=YOU)
    rows = (db.query(md.RankSnapshot)
            .join(md.TrackedKeyword, md.RankSnapshot.keyword_id == md.TrackedKeyword.id)
            .filter(md.TrackedKeyword.keyword == "alpha keyword",
                    md.RankSnapshot.asin == YOU,
                    md.RankSnapshot.checked_at == date(2026, 7, 8)).all())
    assert len(rows) == 1 and rows[0].organic_rank == 5
    # other dates untouched
    d1 = (db.query(md.RankSnapshot)
          .join(md.TrackedKeyword, md.RankSnapshot.keyword_id == md.TrackedKeyword.id)
          .filter(md.TrackedKeyword.keyword == "alpha keyword",
                  md.RankSnapshot.asin == YOU,
                  md.RankSnapshot.checked_at == date(2026, 7, 1)).one())
    assert d1.organic_rank == 3


def test_movers_new_vs_delta(db, tmp_path):
    res = tk.migrate(db, _sheet_path(tmp_path), "Fixture", date(2026, 7, 1))
    pid = res["project_id"]
    path = _cerebro_csv(tmp_path, [
        {"Keyword Phrase": "alpha keyword", "Organic Rank": "1"},     # 3 -> 1 climber
        {"Keyword Phrase": "epsilon keyword", "Organic Rank": "7"},   # new
    ])
    tk.import_cerebro(db, pid, path, date(2026, 7, 8), asin=YOU)
    mv = tk.movers(db, pid)
    assert [c["keyword"] for c in mv["climbers"]] == ["alpha keyword"]
    assert mv["climbers"][0]["delta"] == -2
    assert [n["keyword"] for n in mv["new"]] == ["epsilon keyword"]   # no fake delta
    assert all("delta" not in n for n in mv["new"])


def test_manual_cell_edit(db, tmp_path):
    res = tk.migrate(db, _sheet_path(tmp_path), "Fixture", date(2026, 7, 1))
    k = db.query(md.TrackedKeyword).filter_by(keyword="delta keyword").one()
    out = tk.set_cell(db, k.id, YOU, 30, date(2026, 7, 2))
    assert out["organic_rank"] == 30
    m = tk.matrix(db, res["project_id"])
    row = next(r for r in m["rows"] if r["keyword"] == "delta keyword")
    assert row["ranks"][m["asins"].index(YOU)] == 30


def test_ppc_suggest(db, tmp_path):
    res = tk.migrate(db, _sheet_path(tmp_path), "Fixture", date(2026, 7, 1))
    pid = res["project_id"]
    path = _cerebro_csv(tmp_path, [
        {"Keyword Phrase": "alpha keyword", "Search Volume": "1000", "Organic Rank": "60"},   # page 2
        {"Keyword Phrase": "beta keyword", "Search Volume": "500", "Organic Rank": "10"},     # page 1 -> out
        {"Keyword Phrase": "gamma keyword", "Search Volume": "200", "Organic Rank": "70"},    # sv < min
    ])
    tk.import_cerebro(db, pid, path, date(2026, 7, 8), asin=YOU)
    sg = tk.suggest(db, pid, min_sv=500)
    assert [s["keyword"] for s in sg["rank_support"]] == ["alpha keyword"]
    assert {p["asin"] for p in sg["product_targets"]} == {C1, C2}


def test_competitor_matrix_and_edit(db, tmp_path):
    res = tk.migrate(db, _sheet_path(tmp_path), "Fixture", date(2026, 7, 1))
    cm = tk.competitors(db, res["project_id"])
    rows = {c["asin"]: c for c in cm["competitors"]}
    assert cm["competitors"][0]["asin"] == YOU            # primary pinned first
    assert rows[C1]["revenue"] == 2000.0 and rows[C1]["image_url"] == "http://y"
    # health score COMPUTED: 1.25/Yes — YOU has only PDP Images = Yes in fixture
    assert rows[YOU]["listing_health_score"] == 1.25
    assert rows[YOU]["pdp_images"] is True
    assert rows[YOU]["listing_age"] is not None           # computed from creation_date
    assert cm["kpis"]["total_revenue_top10"] == 3500.0
    # manual audit edit: toggle Yes/None rows; score recomputes; num fields rejected
    out = tk.set_competitor_field(db, rows[C2]["id"], "pdp_videos", False)
    assert out["value"] is False and out["listing_health_score"] == 0.0
    out = tk.set_competitor_field(db, rows[C2]["id"], "brand_story", True)
    assert out["listing_health_score"] == 1.25
    with pytest.raises(ValueError):
        tk.set_competitor_field(db, rows[C2]["id"], "listing_health_score", 7.5)
    with pytest.raises(ValueError):
        tk.set_competitor_field(db, rows[C2]["id"], "revenue", 0)


def test_health_score_scale():
    c = md.TrackedCompetitor()
    assert tk.health_score(c) == 0.0
    for f in tk.AUDIT_FIELDS:
        setattr(c, f, True)
    assert tk.health_score(c) == 10.0                     # 8 x 1.25 = perfect 10
    c.alt_text = False
    c.amazon_badge = None
    assert tk.health_score(c) == 7.5


def test_listing_age():
    assert tk.listing_age_years(None) is None
    assert tk.listing_age_years("garbage") is None
    assert tk.listing_age_years("2025-07-11", date(2026, 7, 11)) == 1.0


def test_export_matrix(db, tmp_path):
    res = tk.migrate(db, _sheet_path(tmp_path), "Fixture", date(2026, 7, 1))
    data = tk.export_matrix(db, res["project_id"])
    df = pd.read_excel(io.BytesIO(data))
    assert len(df) == 4
    assert any("(YOU)" in c for c in df.columns)


# ---- raw-data flow (Listing Optimizer): no sheet needed ------------------------

def _xray_csv(tmp_path, rows):
    p = tmp_path / "xray.csv"
    pd.DataFrame(rows).to_csv(p, index=False)
    return str(p)


def test_create_project_raw(db):
    p = tk.create_project(db, "Raw Project", YOU)
    assert p["primary_asin"] == YOU and p["project_id"]
    with pytest.raises(ValueError):
        tk.create_project(db, "Raw Project")            # duplicate name
    with pytest.raises(ValueError):
        tk.create_project(db, "Other", "not-an-asin")   # bad ASIN
    with pytest.raises(ValueError):
        tk.create_project(db, "   ")                    # blank name


def test_raw_cerebro_defaults_to_primary_asin(db, tmp_path):
    pid = tk.create_project(db, "Raw", YOU)["project_id"]
    path = _cerebro_csv(tmp_path, [
        {"Keyword Phrase": "alpha keyword", "Search Volume": "1000", "Organic Rank": "3"},
    ])
    r = tk.import_cerebro(db, pid, path, date(2026, 7, 1))   # no asin arg
    assert r["asins"] == [YOU] and r["rank_rows"] == 1 and r["keywords_added"] == 1


def test_raw_xray_upsert(db, tmp_path):
    pid = tk.create_project(db, "Raw", YOU)["project_id"]
    # Cerebro tracks YOU (primary via rank col) + C1 (per-ASIN col); C2 NOT tracked
    tk.import_cerebro(db, pid, _cerebro_csv(tmp_path, [
        {"Keyword Phrase": "alpha keyword", "Search Volume": "1000",
         "Organic Rank": "3", C1: "9"},
    ]), date(2026, 7, 1))
    path = _xray_csv(tmp_path, [
        {"ASIN": YOU, "Brand": "You", "Price $": "10", "Revenue": "1500", "Review Count": "100"},
        {"ASIN": C1, "Brand": "CompOne", "Price $": "20", "Revenue": "2000", "Review Count": "50"},
        {"ASIN": C2, "Brand": "NotTracked", "Price $": "30", "Revenue": "9000"},
    ])
    r = tk.import_xray(db, pid, path)
    assert r["matched"] == 2 and r["skipped"] == 1 and r["added"] == 3
    assert r["active"] == 2                             # YOU + C1; C2 stored hidden
    rows = {c.asin: c for c in db.query(md.TrackedCompetitor).filter_by(project_id=pid)}
    assert rows[C2].active is False                     # kept, not displayed
    assert rows[YOU].is_primary is True and rows[C1].is_primary is False
    assert rows[C1].revenue == 2000.0
    cm = tk.competitors(db, pid)
    assert {c["asin"] for c in cm["competitors"]} == {YOU, C1}
    # manual audit survives re-import; imported attrs update
    tk.set_competitor_field(db, rows[C1].id, "pdp_images", True)
    path2 = _xray_csv(tmp_path, [{"ASIN": C1, "Brand": "CompOne", "Price $": "25"}])
    r2 = tk.import_xray(db, pid, path2)
    assert r2["added"] == 0 and r2["updated"] == 1
    c1 = db.get(md.TrackedCompetitor, rows[C1].id)
    assert c1.price == 25.0 and c1.pdp_images is True and c1.revenue == 2000.0
    with pytest.raises(ValueError):
        tk.import_xray(db, pid, _xray_csv(tmp_path, [{"Foo": "bar"}]))  # not an X-ray


def test_xray_first_then_cerebro_order_independent(db, tmp_path):
    """Bug fix: X-ray uploaded BEFORE Cerebro used to drop the competitor rows
    for good — now they're stored inactive and re-activate when Cerebro lands."""
    pid = tk.create_project(db, "Order", YOU)["project_id"]
    r = tk.import_xray(db, pid, _xray_csv(tmp_path, [
        {"ASIN": YOU, "Brand": "You", "Revenue": "1500"},
        {"ASIN": C1, "Brand": "CompOne", "Revenue": "2000"},
        {"ASIN": C2, "Brand": "CompTwo", "Revenue": "900"},
    ]))
    assert r["matched"] == 1 and r["skipped"] == 2      # only primary tracked yet
    assert {c["asin"] for c in tk.competitors(db, pid)["competitors"]} == {YOU}
    # Cerebro arrives after: tracks C1 (rank col) — its X-ray row activates
    tk.import_cerebro(db, pid, _cerebro_csv(tmp_path, [
        {"Keyword Phrase": "alpha keyword", "Search Volume": "1000",
         "Organic Rank": "3", C1: "9"},
    ]), date(2026, 7, 1))
    cm = tk.competitors(db, pid)
    assert {c["asin"] for c in cm["competitors"]} == {YOU, C1}
    got = next(c for c in cm["competitors"] if c["asin"] == C1)
    assert got["revenue"] == 2000.0                     # X-ray data displayed
    rows = {c.asin: c for c in db.query(md.TrackedCompetitor).filter_by(project_id=pid)}
    assert rows[C2].active is False                     # still untracked, still hidden


def test_xray_active_cap_top10_by_revenue(db, tmp_path):
    pid = tk.create_project(db, "Cap", YOU)["project_id"]
    comps = [f"B0CAP00{i:03d}" for i in range(12)]      # 12 tracked competitors
    tk.import_cerebro(db, pid, _cerebro_csv(tmp_path, [
        {"Keyword Phrase": "alpha keyword", "Search Volume": "1000",
         "Organic Rank": "3", **{a: "5" for a in comps}},
    ]), date(2026, 7, 1))
    r = tk.import_xray(db, pid, _xray_csv(tmp_path,
        [{"ASIN": YOU, "Brand": "You", "Revenue": "50"}] +
        [{"ASIN": a, "Brand": f"C{i}", "Revenue": str((i + 1) * 100)}
         for i, a in enumerate(comps)]))
    assert r["matched"] == 13
    rows = {c.asin: c for c in db.query(md.TrackedCompetitor).filter_by(project_id=pid)}
    assert rows[YOU].active is True                     # primary always active
    actives = [a for a in comps if rows[a].active]
    assert len(actives) == 10
    assert set(actives) == set(comps[2:])               # two lowest-revenue capped out
    cm = tk.competitors(db, pid)
    assert len(cm["competitors"]) == 11                 # primary + top 10 displayed
    m = tk.matrix(db, pid)
    assert len(m["asins"]) == 11                        # coverage matrix capped too


def test_listing_copy_computed_audit(db, tmp_path):
    pid = tk.create_project(db, "Raw", YOU)["project_id"]
    path = _cerebro_csv(tmp_path, [
        {"Keyword Phrase": "alpha keyword", "Search Volume": "1000", "Organic Rank": "3"},
        {"Keyword Phrase": "beta keyword", "Search Volume": "500", "Organic Rank": "9"},
        {"Keyword Phrase": "missing phrase", "Search Volume": "900", "Organic Rank": "20"},
    ])
    tk.import_cerebro(db, pid, path, date(2026, 7, 1))
    tk.set_listing_copy(db, pid, "title", "Alpha Keyword deluxe — now with beta")
    a = tk.listing_audit(db, pid)
    marks = {r["keyword"]: r["marks"] for r in a["rows"]}
    assert marks["alpha keyword"]["title"] == "exact"       # phrase appears verbatim
    assert marks["beta keyword"]["title"] == "broad"        # both words, not adjacent
    assert "title" not in marks["missing phrase"]
    title = next(e for e in a["elements"] if e["element"] == "title")
    assert title["has_copy"] and title["exact"] == 1 and title["broad"] == 1
    assert title["total_exact_sv"] == 1000
    assert a["coverage"]["covered"] == 2
    assert [u["keyword"] for u in a["coverage"]["uncovered_top"]] == ["missing phrase"]
    # empty text clears the block; markers recompute away
    tk.set_listing_copy(db, pid, "title", "  ")
    a2 = tk.listing_audit(db, pid)
    assert all(not r["marks"] for r in a2["rows"])
    with pytest.raises(ValueError):
        tk.set_listing_copy(db, pid, "sp_broad", "x")       # not a pasteable element
    with pytest.raises(ValueError):
        tk.set_listing_copy(db, pid, "title", "x", variant="draft")


def test_competitor_copy_comparison(db, tmp_path):
    """Competitor copy pasted per ASIN: own markers stay separate, competitor
    stats computed against the same keywords; search_terms blocked for comps."""
    pid = tk.migrate(db, _sheet_path(tmp_path), "Fixture", date(2026, 7, 1))["project_id"]
    with pytest.raises(ValueError):
        tk.set_listing_copy(db, pid, "search_terms", "x", asin=C1)   # no comp data source
    with pytest.raises(ValueError):
        tk.set_listing_copy(db, pid, "title", "x", asin="B0UNKNOWN1")  # not tracked
    kw = db.query(md.TrackedKeyword).filter_by(project_id=pid).first()
    own_before = tk.listing_audit(db, pid)["elements"]
    tk.set_listing_copy(db, pid, "title", f"Best {kw.keyword} ever made", asin=C1)
    a = tk.listing_audit(db, pid)
    assert a["elements"] == own_before                    # own audit untouched
    comp = next(c for c in a["competitors"] if c["asin"] == C1)
    t = next(e for e in comp["elements"] if e["element"] == "title")
    assert t["has_copy"] and t["exact"] >= 1
    assert comp["covered"] >= 1
    assert all(e["element"] != "search_terms" for e in comp["elements"])
    assert a["comp_copy"][C1]["title"].startswith("Best ")
    other = next(c for c in a["competitors"] if c["asin"] == C2)
    assert other["covered"] == 0                          # nothing pasted for C2
    # empty text clears the competitor block
    tk.set_listing_copy(db, pid, "title", "  ", asin=C1)
    a2 = tk.listing_audit(db, pid)
    comp2 = next(c for c in a2["competitors"] if c["asin"] == C1)
    assert not next(e for e in comp2["elements"] if e["element"] == "title")["has_copy"]


def test_sanitizer_banned_keywords(db, tmp_path):
    """Banned list flags OUR copy only: word-boundary matches, punctuation
    folded, competitors never scanned."""
    pid = tk.migrate(db, _sheet_path(tmp_path), "Fixture", date(2026, 7, 1))["project_id"]
    assert tk.set_banned(db, "cure\nFDA approved, anti-bacterial\n\ncure")["count"] == 3
    assert tk.get_banned(db) == ["FDA approved", "anti-bacterial", "cure"]
    tk.set_listing_copy(db, pid, "title", "Miracle CURE strips — FDA Approved!")
    tk.set_listing_copy(db, pid, "bullet_points", "Anti-Bacterial coating secures your device")
    tk.set_listing_copy(db, pid, "description", "no curevery here")   # no word-boundary hit
    # competitor copy with banned words must NOT be scanned
    tk.set_listing_copy(db, pid, "title", "the best cure ever", asin=C1)
    r = tk.sanitize(db, pid)
    by = {e["element"]: e for e in r["elements"]}
    assert by["title"]["flagged"] == 2 and set(by["title"]["phrases"]) == {"cure", "FDA approved"}
    assert by["bullet_points"]["phrases"] == ["anti-bacterial"]       # hyphen folded
    assert by["description"]["flagged"] == 0                          # "curevery" != "cure"; "secures" != "cure"
    assert by["search_terms"]["checked"] in (True, False)             # element always reported
    assert r["total_flagged"] == 3 and r["banned_count"] == 3
    # clearing the list clears the report
    tk.set_banned(db, "")
    r2 = tk.sanitize(db, pid)
    assert r2["banned_count"] == 0 and r2["total_flagged"] == 0


def test_sheet_markers_survive_as_fallback(db, tmp_path):
    """Migrated sheets keep their hand-marked audit until raw copy is pasted."""
    res = tk.migrate(db, _sheet_path(tmp_path), "Fixture", date(2026, 7, 1))
    a = tk.listing_audit(db, res["project_id"])
    assert any(r["marks"] for r in a["rows"])               # sheet markers visible
    # pasting copy for an element overrides its stored markers with computed ones
    tk.set_listing_copy(db, res["project_id"], "title", "zzz nothing matches")
    a2 = tk.listing_audit(db, res["project_id"])
    assert all("title" not in r["marks"] for r in a2["rows"])


def test_raw_xray_real_export_headers(db, tmp_path):
    """Real Helium10 X-ray CSV drift: BOM, 'Price  $' (double space), 'Fees  $',
    'ASIN Sales'/'ASIN Revenue', 'Active Sellers' (no #), 'Image URL',
    'Mar 22, 2011' creation dates."""
    p = tmp_path / "xray_real.csv"
    p.write_text(
        '﻿"Display Order","Product Details","ASIN","URL","Image URL","Brand",'
        '"Price  $","Parent Level Sales","ASIN Sales","ASIN Revenue","BSR",'
        '"Seller Country/Region","Fees  $","Active Sellers","Ratings","Review Count",'
        '"Creation Date"\n'
        f'4.,"Rescue Tube","{YOU}","http://u","http://img","Kemp USA",67.95,'
        '"1,180",568,"35,577.01","22,416","US",22.76,13,4.7,308,"Mar 22, 2011"\n',
        encoding="utf-8")
    pid = tk.create_project(db, "Real", YOU)["project_id"]
    tk.import_cerebro(db, pid, _cerebro_csv(tmp_path, [
        {"Keyword Phrase": "rescue tube", "Search Volume": "5000", "Organic Rank": "3"},
    ]), date(2026, 7, 12))                              # tracks YOU
    r = tk.import_xray(db, pid, str(p))
    assert r["added"] == 1 and r["matched"] == 1
    c = db.query(md.TrackedCompetitor).filter_by(project_id=pid).one()
    assert c.is_primary and c.price == 67.95 and c.sales == 568
    assert c.revenue == 35577.01 and c.bsr == 22416 and c.fba_fees == 22.76
    assert c.active_sellers == 13 and c.rating == 4.7 and c.review_count == 308
    assert c.creation_date == "2011-03-22" and c.image_url == "http://img"


def test_raw_cerebro_real_export_shape(db, tmp_path):
    """Real Cerebro CSV: BOM'd 'Keyword Phrase', 'Position (Rank)' (primary's own
    ranks) PLUS per-competitor ASIN rank columns in the same file."""
    p = tmp_path / "cerebro_real.csv"
    p.write_text(
        '﻿"Keyword Phrase","Keyword Sales","Search Volume","Position (Rank)",'
        f'"{C1}","{C2}"\n'
        '"rescue tube",120,5000,3,"12","-"\n'
        '"lifeguard buoy",40,2000,"-","7",20\n',
        encoding="utf-8")
    pid = tk.create_project(db, "Real", YOU)["project_id"]
    r = tk.import_cerebro(db, pid, str(p), date(2026, 7, 12))   # no asin arg
    assert set(r["asins"]) == {YOU, C1, C2}
    assert r["keywords_added"] == 2 and r["rank_rows"] == 4     # 3 + '-' skipped x2
    m = tk.matrix(db, pid)
    row = next(x for x in m["rows"] if x["keyword"] == "rescue tube")
    # matrix columns need competitors imported; check snapshots directly instead
    snaps = {(s.asin, s.organic_rank) for s in db.query(md.RankSnapshot)}
    assert (YOU, 3) in snaps and (C1, 12) in snaps and (C2, 20) in snaps
    k = db.query(md.TrackedKeyword).filter_by(keyword="rescue tube").one()
    assert k.search_volume == 5000 and k.keyword_sales == 120


def test_relevancy_computed_from_listing_texts(db, tmp_path):
    """REL = # of ASINs whose listing (Title/Bullets/Description) uses the
    keyword — computed live, not the sheet's manual tag."""
    pid = tk.create_project(db, "Rel", YOU)["project_id"]
    tk.import_cerebro(db, pid, _cerebro_csv(tmp_path, [
        {"Keyword Phrase": "rescue tube", "Search Volume": "5000", "Organic Rank": "3", C1: "9", C2: "4"},
        {"Keyword Phrase": "lifeguard float", "Search Volume": "900", "Organic Rank": "8", C1: "2", C2: "7"},
        {"Keyword Phrase": "unrelated thing", "Search Volume": "100", "Organic Rank": "50", C1: "60", C2: "70"},
    ]), date(2026, 7, 1))
    tk.import_xray(db, pid, _xray_csv(tmp_path, [
        {"ASIN": YOU, "Product Details": "Kemp Rescue Tube for Lifeguards", "Revenue": "100"},
        {"ASIN": C1, "Product Details": "Pro Rescue Tube 50in", "Revenue": "200"},
        {"ASIN": C2, "Product Details": "Swim Buoy", "Revenue": "300"},
    ]))
    # primary's pasted copy counts too (bullets mention lifeguard float)
    tk.set_listing_copy(db, pid, "bullet_points", "Durable LIFEGUARD FLOAT for pool safety")
    m = tk.matrix(db, pid)
    rel = {r["keyword"]: r["relevancy"] for r in m["rows"]}
    assert rel["rescue tube"] == 2          # YOU + C1 titles
    assert rel["lifeguard float"] == 1      # YOU via pasted bullets only
    assert rel["unrelated thing"] == 0


def test_import_catalog_copy(db, tmp_path):
    """Product Benchmark -> Listing Audit hand-off: catalog product copy prefills
    OUR listing elements; empty catalog elements never wipe existing pastes."""
    pid = tk.migrate(db, _sheet_path(tmp_path), "Fixture", date(2026, 7, 1))["project_id"]
    tk.set_listing_copy(db, pid, "description", "hand-written keeper")
    product = {"sku": "10-115", "asin": "B0TEST0009",
               "title": "Kemp USA EMS Backpack",
               "bullets": ["First bullet", "Second bullet"],
               "description": "",                       # empty in the report
               "search_terms": ["emt bag", "ems backpack"]}
    r = tk.import_catalog_copy(db, pid, product)
    assert r["imported"] == ["title", "bullet_points", "search_terms"]
    assert r["skipped"] == ["description"] and r["asin"] == "B0TEST0009"
    copy = tk.listing_audit(db, pid)["copy"]   # {element: text}
    assert copy["title"] == "Kemp USA EMS Backpack"
    assert copy["bullet_points"] == "First bullet\nSecond bullet"
    assert copy["search_terms"] == "emt bag ems backpack"
    assert copy["description"] == "hand-written keeper"   # untouched
    with pytest.raises(ValueError):
        tk.import_catalog_copy(db, 99999, product)        # unknown project

    # SEO connection: search terms tracked as keywords. Short segments stay
    # phrases; segments past the soup threshold split into words (a stray comma
    # inside Amazon word-soup must not create a giant "phrase").
    tracked = {k.keyword for k in db.query(md.TrackedKeyword).filter_by(project_id=pid)}
    assert {"emt bag", "ems backpack"} <= tracked            # short blobs = phrases
    srcs = {k.keyword: k.source for k in db.query(md.TrackedKeyword).filter_by(project_id=pid)}
    assert srcs["emt bag"] == "search_terms"
    assert r["keywords_added"] == 2 and r["keywords_known"] == 0
    r2 = tk.import_catalog_copy(db, pid, {**product, "search_terms":
        ["snore strips, mouth tape for sleeping", "emt bag",
         "one two three four five six seven eight, short tail"]})
    kws2 = {k.keyword for k in db.query(md.TrackedKeyword).filter_by(project_id=pid)}
    assert "snore strips" in kws2 and "mouth tape for sleeping" in kws2   # comma -> phrases
    assert "seven" in kws2 and "short tail" in kws2   # >6-word segment -> words; short one stays
    assert "one two three four five six seven eight" not in kws2
    assert r2["keywords_known"] == 1                                      # "emt bag" already tracked
    r3 = tk.import_catalog_copy(db, pid, product)          # full re-import -> no dupes
    assert r3["keywords_added"] == 0

    # primary ASIN: migrated project keeps its own; a blank project adopts the product's
    assert db.get(md.TrackerProject, pid).primary_asin != "B0TEST0009"
    pid2 = tk.create_project(db, "blank")["project_id"]
    tk.import_catalog_copy(db, pid2, product)
    assert db.get(md.TrackerProject, pid2).primary_asin == "B0TEST0009"


def test_seo_recommend(db):
    """SEO recs + backend search-term line from Listing Audit copy vs tracked kws."""
    pid = tk.create_project(db, "seo proj", primary_asin="B0TESTSEO1")["project_id"]
    for kw, sv in [("snore strips", 9000), ("mouth tape", 5000), ("nasal dilator", 3000)]:
        db.add(md.TrackedKeyword(project_id=pid, keyword=kw, search_volume=sv))
    db.commit()

    # nothing pasted -> missing-copy recommendations for title + search_terms (high)
    r = tk.seo_recommend(db, pid)
    by_area = {x["area"]: x for x in r["recommendations"]}
    assert by_area["title"]["severity"] == "high"
    assert by_area["search_terms"]["severity"] == "high"

    tk.set_listing_copy(db, pid, "title", "Premium Snore Strips for Better Sleep")
    tk.set_listing_copy(db, pid, "bullet_points", "Comfortable mouth friendly\nEasy to use")
    tk.set_listing_copy(db, pid, "search_terms", "snore strips tape")
    tk.set_banned(db, "dilator")
    r = tk.seo_recommend(db, pid)

    # title: 'snore strips' is exact -> only the other two flagged, SV order
    t = next(x for x in r["recommendations"] if x["area"] == "title" and x["keywords"])
    assert [k["keyword"] for k in t["keywords"]] == ["mouth tape", "nasal dilator"]
    # bullets: only 2 lines pasted -> medium bullet-count rec
    assert any(x["area"] == "bullet_points" and "bullet point" in x["title"].lower()
               for x in r["recommendations"])
    # banned phrase present in no copy -> no compliance rec (dilator is only a keyword)
    assert not any(x["area"] == "compliance" for x in r["recommendations"])

    st = r["search_terms"]
    words = {w["word"] for w in st["words"]}
    # visible words (snore/strips/mouth) + banned (dilator) excluded; tape + nasal in
    assert words == {"tape", "nasal"}
    assert st["bytes"] <= st["max_bytes"]
    # 'tape' outranks 'nasal' (its keyword has higher SV) -> first in the line
    assert st["suggested"].split()[0] == "tape"
    # current field wastes snore+strips (already indexed by the title)
    assert st["wasted_words"] == ["snore", "strips"]
    assert any(x["area"] == "search_terms" and "already indexed" in x["title"]
               for x in r["recommendations"])

    # unknown project
    with pytest.raises(ValueError):
        tk.seo_recommend(db, 999)


def test_seo_recommend_proposed_variant(db):
    """Proposed draft is scored independently: current copy/markers never leak in."""
    pid = tk.create_project(db, "variant proj", primary_asin="B0TESTSEO2")["project_id"]
    db.add(md.TrackedKeyword(project_id=pid, keyword="snore strips", search_volume=9000))
    db.commit()
    tk.set_listing_copy(db, pid, "title", "Premium Snore Strips for Better Sleep")

    # proposed is empty -> missing-copy recs fire even though current has a title
    r = tk.seo_recommend(db, pid, variant="proposed")
    assert r["variant"] == "proposed"
    assert any(x["area"] == "title" and "No title" in x["title"] for x in r["recommendations"])
    # and the current variant stays clean on that front
    r_cur = tk.seo_recommend(db, pid, variant="current")
    assert not any("No title" in x["title"] for x in r_cur["recommendations"])

    # draft a proposed title WITHOUT the keyword -> flagged on proposed only
    tk.set_listing_copy(db, pid, "title", "Brand New Sleep Aid", variant="proposed")
    r = tk.seo_recommend(db, pid, variant="proposed")
    t = next(x for x in r["recommendations"] if x["area"] == "title" and x["keywords"])
    assert t["keywords"][0]["keyword"] == "snore strips"
    # proposed search-term line treats only the proposed copy as visible:
    # 'snore'/'strips' are NOT in the draft title -> suggested as backend words
    assert {w["word"] for w in r["search_terms"]["words"]} >= {"snore", "strips"}
    # while on current they're visible (in the live title) -> excluded
    assert not {"snore", "strips"} & {w["word"] for w in r_cur["search_terms"]["words"]}


def test_report_xlsx(db):
    """Exec report workbook: 4 sheets mirroring the views, charts attached."""
    import io as _io
    import openpyxl
    pid = tk.create_project(db, "report proj", primary_asin="B0TESTREP1")["project_id"]
    db.add(md.TrackedKeyword(project_id=pid, keyword="snore strips", search_volume=9000))
    db.add(md.TrackedCompetitor(project_id=pid, asin="B0TESTREP1", brand="You",
                                is_primary=True, active=True, revenue=1000.0, review_count=10))
    db.add(md.TrackedCompetitor(project_id=pid, asin="B0TESTCMP9", brand="Rival",
                                is_primary=False, active=True, revenue=5000.0, review_count=99))
    db.commit()
    tk.set_listing_copy(db, pid, "title", "Premium Snore Strips")
    tk.set_cell(db, db.query(md.TrackedKeyword).filter_by(project_id=pid).first().id,
                "B0TESTREP1", 12)      # page-1 rank -> distribution has data

    data = tk.report_xlsx(db, pid)
    wb = openpyxl.load_workbook(_io.BytesIO(data))
    assert wb.sheetnames == ["Overview", "SEO", "Listing Audit", "Product Overview"]
    ov = wb["Overview"]
    assert "report proj" in ov.cell(row=1, column=1).value
    # SEO sheet rank distribution counted the page-1 rank
    seo = wb["SEO"]
    vals = {seo.cell(row=i, column=1).value: seo.cell(row=i, column=2).value for i in range(5, 9)}
    assert vals.get("Page 1") == 1
    # charts attached (native Excel charts, not images)
    assert len(ov._charts) >= 1 and len(seo._charts) >= 1
    assert len(wb["Listing Audit"]._charts) >= 1 and len(wb["Product Overview"]._charts) >= 1
    # unknown project
    with pytest.raises(ValueError):
        tk.report_xlsx(db, 999)

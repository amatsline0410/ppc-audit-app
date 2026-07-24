"""SKU-level transactions (Payments Date Range report) — parse / merge / summary."""
import textwrap

import pytest

from app.pipeline import transactions as t

CSV = textwrap.dedent('''\
    "Includes Amazon Marketplace, Fulfillment by Amazon (FBA), and Amazon Webstore transactions"
    "All amounts in USD, unless specified"
    "Definitions:"
    "date/time: posted date/time of the transaction"
    "date/time","settlement id","type","order id","sku","description","quantity","marketplace","fulfillment","order city","order state","order postal","tax collection model","product sales","product sales tax","shipping credits","shipping credits tax","gift wrap credits","giftwrap credits tax","Regulatory Fee","Tax On Regulatory Fee","promotional rebates","promotional rebates tax","marketplace withheld tax","selling fees","fba fees","other transaction fees","other","total","Transaction Status","Transaction Release Date"
    "Jun 1, 2026 5:43:55 AM PDT","111","Transfer","","","To your account ending in: 429","","","","","","","","0","0","0","0","0","0","0","0","0","0","0","0","0","0","-100.00","-100.00","Released","Jun 1, 2026 5:43:55 AM PDT"
    "Jun 2, 2026 7:11:52 AM PDT","222","Order","112-0000000-0000001","SKU-A","Product A","1","amazon.com","Amazon","AUSTIN","TX","78701","MarketplaceFacilitator","34.99","2.45","0","0","0","0","0","0","0","0","-2.45","-5.25","-3.54","0","0","26.20","Released","Jun 10, 2026 1:00:00 AM PDT"
    "Jun 5, 2026 8:00:00 AM PDT","222","Order","112-0000000-0000002","SKU-A","Product A","2","amazon.com","Amazon","BOSTON","MA","02108","MarketplaceFacilitator","69.98","4.90","0","0","0","0","0","0","-3.50","0","-4.90","-10.50","-7.08","0","0","48.90","Released","Jun 12, 2026 1:00:00 AM PDT"
    "Jun 20, 2026 9:00:00 AM PDT","333","Refund","112-0000000-0000001","SKU-A","Product A","1","amazon.com","Amazon","AUSTIN","TX","78701","MarketplaceFacilitator","-34.99","0","0","0","0","0","0","0","0","0","0","3.99","0","0","0","-31.00","Released","Jun 20, 2026 9:00:00 AM PDT"
    "Jun 25, 2026 1:00:00 PM PDT","333","Order","113-0000000-0000003","SKU-B","Product B","1","amazon.com","Amazon","DENVER","CO","80014","MarketplaceFacilitator","19.99","1.40","0","0","0","0","0","0","0","0","-1.40","-3.00","-3.54","0","0","13.45","Released","Jul 1, 2026 1:00:00 AM PDT"
''')


@pytest.fixture
def rows(tmp_path):
    p = tmp_path / "txn.csv"
    p.write_text(CSV)
    return t.parse_txn(str(p))


def test_parse(rows):
    assert len(rows) == 5
    order = rows[1]
    assert order["sku"] == "SKU-A" and order["type"] == "Order"
    assert order["date"] == "2026-06-02" and order["quantity"] == 1
    assert order["product_sales"] == 34.99 and order["total"] == 26.20
    assert order["selling_fees"] == -5.25 and order["fba_fees"] == -3.54
    transfer = rows[0]
    assert transfer["sku"] == "" and transfer["total"] == -100.00


def test_parse_rejects_non_transaction_file(tmp_path):
    p = tmp_path / "bad.csv"
    p.write_text("a,b,c\n1,2,3\n")
    with pytest.raises(ValueError):
        t.parse_txn(str(p))


def test_merge_dedupes_reupload(rows):
    data, added, updated, dupes = t.merge({}, rows, "june.csv")
    assert added == 5 and updated == 0 and dupes == 0
    data, added2, updated2, dupes2 = t.merge(data, rows, "june.csv")
    assert added2 == 0 and updated2 == 0 and dupes2 == 5
    assert len(data["rows"]) == 5
    assert [f["name"] for f in data["files"]] == ["june.csv"]  # re-upload replaces entry


def test_merge_updates_changed_rows(rows):
    """Re-export of a known transaction with changed data (status flip, fee
    correction) must REPLACE the stored row — never double-count it."""
    data, _, _, _ = t.merge({}, rows, "june.csv")
    changed = [dict(r) for r in rows]
    changed[1]["status"] = "Deferred"        # SKU-A order: status flip
    changed[4]["fba_fees"] = -4.00           # SKU-B order: fee correction
    changed[4]["total"] = 12.99
    data, added, updated, dupes = t.merge(data, changed, "june_v2.csv")
    assert added == 0 and updated == 2 and dupes == 3
    assert len(data["rows"]) == 5            # no double-count
    by = {r["order_id"] + r["type"]: r for r in data["rows"] if r.get("sku")}
    assert by["112-0000000-0000001Order"]["status"] == "Deferred"
    assert by["113-0000000-0000003Order"]["total"] == 12.99
    s = t.summary(data)
    assert s["totals"]["net"] == pytest.approx(26.20 + 48.90 - 31.00 + 12.99)
    assert {f["name"] for f in data["files"]} == {"june.csv", "june_v2.csv"}


def test_summary_rollup(rows):
    data, _, _, _ = t.merge({}, rows, "june.csv")
    s = t.summary(data)
    assert s["range"]["min"] == "2026-06-01" and s["range"]["max"] == "2026-06-25"
    tot = s["totals"]
    assert tot["orders"] == 3 and tot["refunds"] == 1 and tot["units"] == 3
    assert tot["product_sales"] == pytest.approx(89.97)
    assert tot["net"] == pytest.approx(26.20 + 48.90 - 31.00 + 13.45)
    assert tot["transfers"] == -100.00
    by = {p["sku"]: p for p in s["skus"]}
    a = by["SKU-A"]
    assert a["orders"] == 2 and a["refunds"] == 1 and a["units"] == 2
    assert a["net"] == pytest.approx(26.20 + 48.90 - 31.00)
    assert by["SKU-B"]["units"] == 1


def test_summary_date_filter_and_sku_drill(rows):
    data, _, _, _ = t.merge({}, rows, "june.csv")
    s = t.summary(data, start="2026-06-03", end="2026-06-24")
    assert s["totals"]["transactions"] == 2       # 6/5 order + 6/20 refund
    assert {p["sku"] for p in s["skus"]} == {"SKU-A"}
    assert s["totals"]["net"] == pytest.approx(48.90 - 31.00)
    s2 = t.summary(data, sku="SKU-B")
    assert len(s2["transactions"]) == 1
    assert len(s2["skus"]) == 2                   # rollup not narrowed by sku


def test_report_xlsx(rows):
    import io
    from openpyxl import load_workbook
    data, _, _, _ = t.merge({}, rows, "june.csv")
    wb = load_workbook(io.BytesIO(t.report_xlsx(data)))
    assert wb.sheetnames == ["Overview", "Daily Trend", "By SKU", "Transactions"]
    ov = wb["Overview"]
    kpis = {ov.cell(row=r, column=1).value: ov.cell(row=r, column=2).value for r in range(4, 15)}
    assert kpis["Orders"] == 3 and kpis["Refunds"] == 1
    assert kpis["Net proceeds"] == pytest.approx(57.55)
    assert len(ov._charts) == 2                      # net-by-SKU bar + deductions pie
    assert len(wb["Daily Trend"]._charts) == 1       # sales vs net line
    assert wb["Transactions"].max_row == 3 + 4       # header + 4 SKU rows
    # date-window export narrows everything
    wb2 = load_workbook(io.BytesIO(t.report_xlsx(data, start="2026-06-03", end="2026-06-24")))
    assert wb2["Transactions"].max_row == 3 + 2
    with pytest.raises(ValueError):
        t.report_xlsx({})                            # empty ledger


def test_store_roundtrip_and_clear(tmp_path, monkeypatch):
    from app import database as dbmod
    monkeypatch.setattr(dbmod, "_store_dir", lambda sid: str(tmp_path))
    p = tmp_path / "txn.csv"
    p.write_text(CSV)
    data, _, _, _ = t.merge({}, t.parse_txn(str(p)), "june.csv")
    t.write_txn("s1", data)
    assert len(t.read_txn("s1")["rows"]) == 5
    t.delete_all("s1")
    assert t.read_txn("s1") == {}
